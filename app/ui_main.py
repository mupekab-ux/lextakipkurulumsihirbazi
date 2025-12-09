# -*- coding: utf-8 -*-
from PyQt6.QtCore import (
    QAbstractTableModel,
    Qt,
    QModelIndex,
    QSettings,
    QSortFilterProxyModel,
    QDate,
    pyqtSignal,
    QPoint,
    QTimer,
    QThread,
    QByteArray,
    QMarginsF,
    QRect,
    QRectF,
    QSizeF,
    QStringListModel,
)
from PyQt6.QtGui import (
    QColor,
    QFont,
    QIcon,
    QPalette,
    QBrush,
    QKeySequence,
    QShortcut,
    QTextDocument,
    QTextCharFormat,
    QPainter,
    QPageLayout,
    QPageSize,
    QPen,
)
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFrame,
    QLineEdit,
    QPushButton,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QComboBox,
    QCheckBox,
    QLabel,
    QHeaderView,
    QTabWidget,
    QFileDialog,
    QInputDialog,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QStyleOptionHeader,
    QStyle,
    QMenu,
    QDialog,
    QFormLayout,
    QDateEdit,
    QDoubleSpinBox,
    QDialogButtonBox,
    QRadioButton,
    QSpinBox,
    QButtonGroup,
    QTabBar,
    QAbstractItemDelegate,
    QAbstractItemView,
    QGridLayout,
    QToolButton,
    QSizePolicy,
    QCompleter,
    QListWidget,
    QListWidgetItem,
    QGroupBox,
    QSplitter,
    QCalendarWidget,
    QPlainTextEdit,
)
from PyQt6.QtPrintSupport import QPrinter
from datetime import datetime, date, timedelta
from dataclasses import dataclass
import html
import os
import re
import time
import json
import sqlite3
from functools import partial
from typing import Any, Callable, Iterable, List, Literal, Optional
from collections import Counter

try:  # pragma: no cover - runtime import guard
    from app.db import (
        get_connection,
        is_case_closed,
        update_dosya_with_auto_timeline,
        get_manual_tasks_between,
        insert_manual_task,
        update_manual_task,
        delete_manual_task,
        get_case_tasks_between,
        get_all_manual_tasks,
        get_pending_tasks,
        get_completed_tasks,
        mark_task_complete,
    )
except ModuleNotFoundError:  # pragma: no cover
    from db import (
        get_connection,
        is_case_closed,
        update_dosya_with_auto_timeline,
        get_manual_tasks_between,
        insert_manual_task,
        update_manual_task,
        delete_manual_task,
        get_case_tasks_between,
        get_all_manual_tasks,
        get_pending_tasks,
        get_completed_tasks,
        mark_task_complete,
    )

try:  # pragma: no cover - runtime import guard
    from app.models import (
        log_action,
        fetch_dosyalar_by_color_hex,
        set_archive_status,
        get_all_dosyalar,
        export_dosyalar_to_csv,
        export_dosyalar_to_xlsx,
        export_dosyalar_to_docx,
        get_users,
        get_permissions_for_role,
        list_finance_overview,
        mark_next_installment_paid,
        add_partial_payment,
        create_custom_tab,
        rename_custom_tab,
        get_dosya_ids_for_tab,
        list_custom_tabs,
        delete_custom_tab,
        harici_create,
        harici_get_master_list,
        harici_update_quick_info,
        get_setting,
        delete_case_hard,
        summarize_finance_by_ids,
        summarize_harici_finance_by_ids,
        get_statuses,
        get_status_color,
        update_dosya,
    )
except ModuleNotFoundError:  # pragma: no cover
    from models import (
        log_action,
        fetch_dosyalar_by_color_hex,
        set_archive_status,
        get_all_dosyalar,
        export_dosyalar_to_csv,
        export_dosyalar_to_xlsx,
        export_dosyalar_to_docx,
        get_users,
        get_permissions_for_role,
        list_finance_overview,
        mark_next_installment_paid,
        add_partial_payment,
        create_custom_tab,
        rename_custom_tab,
        get_dosya_ids_for_tab,
        list_custom_tabs,
        delete_custom_tab,
        harici_create,
        harici_get_master_list,
        harici_update_quick_info,
        get_setting,
        delete_case_hard,
        summarize_finance_by_ids,
        summarize_harici_finance_by_ids,
        get_statuses,
        get_status_color,
        update_dosya,
    )

try:  # pragma: no cover - runtime import guard
    from app.utils import (
        normalize_hex,
        ROLE_ABBREVIATIONS,
        COLOR_MAP,
        USER_ROLE_LABELS,
        ASSIGNMENT_EDIT_ROLES,
        format_tl,
        tl_to_cents,
        get_durusma_color,
        get_task_color_by_date,
        hex_to_qcolor,
        get_status_text_color,
    )
except ModuleNotFoundError:  # pragma: no cover
    from utils import (
        normalize_hex,
        ROLE_ABBREVIATIONS,
        COLOR_MAP,
        USER_ROLE_LABELS,
        ASSIGNMENT_EDIT_ROLES,
        format_tl,
        tl_to_cents,
        get_durusma_color,
        get_task_color_by_date,
        hex_to_qcolor,
        get_status_text_color,
    )

try:  # pragma: no cover - runtime import guard
    from app.status_helpers import get_dava_durumu_list
except ModuleNotFoundError:  # pragma: no cover
    from status_helpers import get_dava_durumu_list  # type: ignore

try:  # pragma: no cover - runtime import guard
    from app.ui_edit_dialog import EditDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_edit_dialog import EditDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_arabuluculuk_tab import ArabuluculukTab
except ModuleNotFoundError:  # pragma: no cover
    from ui_arabuluculuk_tab import ArabuluculukTab

try:  # pragma: no cover - runtime import guard
    from app.ui_tebligatlar_tab import TebligatlarTab
except ModuleNotFoundError:  # pragma: no cover
    from ui_tebligatlar_tab import TebligatlarTab

try:  # pragma: no cover - runtime import guard
    from app.services.dosya_service import get_dosya_assignees, set_dosya_assignees
except ModuleNotFoundError:  # pragma: no cover
    from services.dosya_service import get_dosya_assignees, set_dosya_assignees

try:  # pragma: no cover - runtime import guard
    from app.ui_attachments_dialog import AttachmentsDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_attachments_dialog import AttachmentsDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_settings_dialog import SettingsDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_settings_dialog import SettingsDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_finance_dialog import FinanceDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_finance_dialog import FinanceDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_finans_harici_quick_dialog import FinansHariciQuickDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_finans_harici_quick_dialog import FinansHariciQuickDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_finans_harici_dialog import FinansHariciDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_finans_harici_dialog import FinansHariciDialog

try:  # pragma: no cover - runtime import guard
    from app.ui_vekalet_dialog import VekaletDialog
except ModuleNotFoundError:  # pragma: no cover
    from ui_vekalet_dialog import VekaletDialog

try:  # pragma: no cover - runtime import guard
    from app.workers import ChangeDetectorWorker
except ModuleNotFoundError:  # pragma: no cover
    from workers import ChangeDetectorWorker

try:  # pragma: no cover - runtime import guard
    from app.demo_manager import get_demo_manager
    from app.ui_demo_dialog import DemoStatusWidget
    from app.db import get_database_path
except ModuleNotFoundError:  # pragma: no cover
    from demo_manager import get_demo_manager
    from ui_demo_dialog import DemoStatusWidget
    from db import get_database_path


def _resource_path(relative_path: str) -> str:
    """
    PyInstaller ile paketlendiğinde dosya yollarını düzgün çözer.
    """
    import sys
    import os
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


ALERT_CATEGORY_META: dict[str, dict[str, str]] = {
    "hearing": {"icon": "⚖️", "accent": "#ffb300", "label": "Duruşma"},
    "notice": {"icon": "✉️", "accent": "#26c6da", "label": "Tebligat"},
    "payment": {"icon": "₺", "accent": "#2e7d32", "label": "Ödeme"},
    "mediation": {"icon": "🤝", "accent": "#8e24aa", "label": "Arabul."},
}

PreparedDosyaRow = tuple[
    dict[str, Any],
    dict[int, dict[Any, object]],
]

_FONT_BOLD = QFont()
_FONT_BOLD.setBold(True)
_ARROW_TEXT = "→"
_CENTER_ALIGNMENT = int(Qt.AlignmentFlag.AlignCenter)
_QCOLOR_CACHE: dict[str, QColor] = {}
_QBRUSH_CACHE: dict[str, QBrush] = {}
STATUS_BRUSHES: dict[str, QBrush | None] = {}
STATUS_FG: dict[str, QColor] = {}
DEFAULT_STATUS_FG = QColor("#000000")
_STATUS_PALETTE_LOADED = False
_DATE_FIELDS = {"durusma_tarihi", "is_tarihi", "is_tarihi_2"}
_STATUS_COLOR_FIELDS = {"dava_durumu", "tekrar_dava_durumu_2"}
_STATUS_COLOR_META = "status_color_key"
OPTIONAL_DATE_MIN = QDate(1900, 1, 1)
OPTIONAL_DATE_MAX = QDate(7999, 12, 31)
_JOB_DATE_BRUSHES = {
    "past": QBrush(QColor("#ff4d4d")),
    "today": QBrush(QColor("#4caf50")),
    "soon": QBrush(QColor("#ffeb3b")),
    "future": QBrush(QColor("#2196f3")),
}
COL_SELECTION = 0
COL_BN = 1
COL_DOSYA_ESAS_NO = 2
COL_MUVEKKIL_ADI = 3
COL_KARSI_TARAF = 4
COL_DOSYA_KONUSU = 5
COL_MAHKEME_ADI = 6
COL_DURUSMA_TARIHI = 7
COL_DAVA_DURUMU = 8
COL_IS_TARIHI = 9
COL_ACIKLAMA = 10
COL_DAVA_DURUMU_2 = 11
COL_IS_TARIHI_2 = 12
COL_ACIKLAMA_2 = 13

DEFAULT_DOSYALAR_COLUMN_WIDTHS = {
    COL_SELECTION: 36,
    COL_BN: 45,
    COL_DOSYA_ESAS_NO: 120,
    COL_MUVEKKIL_ADI: 150,
    COL_KARSI_TARAF: 140,
    COL_DOSYA_KONUSU: 150,
    COL_MAHKEME_ADI: 130,
    COL_DURUSMA_TARIHI: 110,
    COL_DAVA_DURUMU: 220,
    COL_IS_TARIHI: 110,
    COL_ACIKLAMA: 200,
    COL_DAVA_DURUMU_2: 220,
    COL_IS_TARIHI_2: 110,
    COL_ACIKLAMA_2: 200,
}


def _coerce_to_qdate(value: Any) -> QDate | None:
    if isinstance(value, QDate):
        return value if value.isValid() else None
    if isinstance(value, datetime):
        return QDate(value.year, value.month, value.day)
    if isinstance(value, date):
        return QDate(value.year, value.month, value.day)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        parsed = QDate.fromString(text, "yyyy-MM-dd")
        if parsed.isValid():
            return parsed
        parsed = QDate.fromString(text, "dd.MM.yyyy")
        if parsed.isValid():
            return parsed
    return None
def _date_sort_key(value: date | None) -> int:
    if value is None:
        return 0
    return value.year * 10000 + value.month * 100 + value.day


def _int_sort_key(value: Any) -> int:
    try:
        if value in (None, ""):
            return 0
        return int(str(value).strip())
    except (TypeError, ValueError):
        return 0


def _normalize_text_value(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _with_role_abbreviation(base_text: str, role_value: Any) -> str:
    role_key = ""
    if role_value not in (None, ""):
        role_key = str(role_value)
    abbr = ROLE_ABBREVIATIONS.get(role_key)
    if not abbr:
        return base_text
    if base_text:
        return f"{base_text} ({abbr})"
    return f"({abbr})"


def _job_date_background(value: Any) -> QBrush | None:
    date_value = coerce_to_date(value)
    if date_value is None:
        return None
    today = date.today()
    days = (date_value - today).days
    if days < 0:
        return _JOB_DATE_BRUSHES["past"]
    if days == 0:
        return _JOB_DATE_BRUSHES["today"]
    if 1 <= days <= 4:
        return _JOB_DATE_BRUSHES["soon"]
    if days > 4:
        return _JOB_DATE_BRUSHES["future"]
    return None


def load_status_palette(force: bool = False) -> None:
    """Populate shared status brush/foreground caches once."""

    global _STATUS_PALETTE_LOADED
    if _STATUS_PALETTE_LOADED and not force:
        return
    STATUS_BRUSHES.clear()
    STATUS_FG.clear()
    try:
        statuses = get_statuses()
    except Exception as exc:  # pragma: no cover - defensive logging
        print(f"[perf][warn] status palette load failed: {exc}")
        _STATUS_PALETTE_LOADED = False
        return
    for status in statuses:
        name = _normalize_text_value(status.get("ad"))
        if not name:
            continue
        color_hex = normalize_hex(status.get("color_hex"))
        if not color_hex:
            try:
                color_hex = normalize_hex(get_status_color(name))
            except Exception as exc:  # pragma: no cover - defensive logging
                print(f"[perf][warn] status color fetch failed: {exc}")
                color_hex = None
        _apply_status_palette_entry(name, color_hex)
    _STATUS_PALETTE_LOADED = True


def _apply_status_palette_entry(name: str, color_hex: str | None) -> None:
    normalized = _normalize_text_value(name)
    if not normalized:
        return
    brush = _cached_brush(color_hex)
    fg_hex = get_status_text_color(color_hex)
    fg_color = _cached_qcolor(fg_hex) or DEFAULT_STATUS_FG
    STATUS_BRUSHES[normalized] = brush
    STATUS_FG[normalized] = fg_color


def _cached_qcolor(hex_code: str | None) -> QColor | None:
    normalized = normalize_hex(hex_code)
    if not normalized:
        return None
    cached = _QCOLOR_CACHE.get(normalized)
    if cached is not None:
        return cached
    color = hex_to_qcolor(normalized)
    _QCOLOR_CACHE[normalized] = color
    return color


def _cached_brush(hex_code: str | None) -> QBrush | None:
    normalized = normalize_hex(hex_code)
    if not normalized:
        return None
    cached = _QBRUSH_CACHE.get(normalized)
    if cached is not None:
        return cached
    color = _cached_qcolor(normalized)
    if color is None:
        return None
    brush = QBrush(color)
    _QBRUSH_CACHE[normalized] = brush
    return brush


def _ensure_status_palette_entry(name: str) -> tuple[QBrush | None, QColor]:
    normalized = _normalize_text_value(name)
    if not normalized:
        return None, DEFAULT_STATUS_FG
    has_bg = normalized in STATUS_BRUSHES
    has_fg = normalized in STATUS_FG
    if has_bg and has_fg:
        return STATUS_BRUSHES[normalized], STATUS_FG[normalized]
    color_hex: str | None = None
    try:
        color_hex = normalize_hex(get_status_color(normalized))
    except Exception as exc:  # pragma: no cover - defensive logging
        print(f"[perf][warn] status color lookup failed: {exc}")
    _apply_status_palette_entry(normalized, color_hex)
    return STATUS_BRUSHES.get(normalized), STATUS_FG.get(normalized, DEFAULT_STATUS_FG)


def _durusma_value_as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    try:
        parsed = date.fromisoformat(text[:10])
    except ValueError:
        return text
    return parsed.isoformat()


def _durusma_color_roles(value: Any) -> tuple[QBrush | None, QColor | None]:
    raw_text = _durusma_value_as_text(value)
    if not raw_text:
        return None, None
    color_info = get_durusma_color(raw_text)
    if not color_info:
        return None, None
    bg_brush = _cached_brush(color_info.get("bg"))
    fg_color = _cached_qcolor(color_info.get("fg"))
    return bg_brush, fg_color


def _now() -> float:
    return time.perf_counter()


def _ms(start: float, end: float) -> str:
    return f"{(end - start) * 1000:.1f} ms"

_TOKEN_PATTERN = re.compile(
    r"#(?P<key>[a-zA-Z_]+):(?P<start>\d{4}-\d{2}-\d{2})(?:\.\.(?P<end>\d{4}-\d{2}-\d{2}))?"
)


@dataclass(frozen=True)
class AlertTokenRange:
    start: date
    end: date


def parse_alert_tokens(text: str) -> tuple[str, dict[str, AlertTokenRange]]:
    """Extract alert tokens from the given text and return the cleaned text."""

    cleaned_parts: list[str] = []
    tokens: dict[str, AlertTokenRange] = {}
    for part in text.split():
        match = _TOKEN_PATTERN.fullmatch(part.strip())
        if not match:
            cleaned_parts.append(part)
            continue
        key = match.group("key").lower()
        if key == "due":
            key = "payment"
        if key not in ALERT_CATEGORY_META:
            cleaned_parts.append(part)
            continue
        try:
            start_value = date.fromisoformat(match.group("start"))
            end_raw = match.group("end") or match.group("start")
            end_value = date.fromisoformat(end_raw)
        except ValueError:
            cleaned_parts.append(part)
            continue
        if end_value < start_value:
            start_value, end_value = end_value, start_value
        tokens[key] = AlertTokenRange(start=start_value, end=end_value)
    cleaned = " ".join(cleaned_parts).strip()
    return cleaned, tokens


def coerce_to_date(value) -> date | None:
    """Convert various date representations into a ``date`` object."""

    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            return None
        return parsed.date()


class FinanceSummaryBar(QFrame):
    """Kompakt finans özet çubuğu."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("financeSummary")
        self.setMinimumHeight(90)
        self.setMaximumHeight(110)
        self.setSizePolicy(
            QSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        )
        layout = QHBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)
        self._value_labels: dict[str, QLabel] = {}
        metrics = [
            ("contract", "Toplam Sözleşme"),
            ("collected", "Toplam Tahsil"),
            ("expense", "Toplam Masraf"),
            ("balance", "Kalan"),
        ]
        for key, title in metrics:
            card = QFrame(self)
            card.setObjectName("financeSummaryCard")
            card.setMinimumWidth(180)
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(12, 8, 12, 8)
            card_layout.setSpacing(4)
            title_label = QLabel(title, card)
            title_label.setProperty("class", "summaryTitle")
            value_label = QLabel("0,00 ₺", card)
            value_label.setProperty("class", "summaryValue")
            card_layout.addWidget(title_label)
            card_layout.addWidget(value_label)
            card_layout.addStretch(1)
            layout.addWidget(card)
            self._value_labels[key] = value_label
        layout.addStretch(1)
        self.setStyleSheet(
            """
            QFrame#financeSummary {
                background: #1f1f1f;
                border: 1px solid #2c2c2c;
                border-radius: 10px;
            }
            QFrame#financeSummaryCard {
                background-color: rgba(255, 255, 255, 0.02);
                border: 1px solid #2c2c2c;
                border-radius: 8px;
            }
            QLabel.summaryTitle {
                color: #bdbdbd;
                font-size: 10pt;
            }
            QLabel.summaryValue {
                color: #eaeaea;
                font-weight: 600;
                font-size: 12pt;
            }
            """
        )

    def update_totals(
        self,
        *,
        contract: int = 0,
        collected: int = 0,
        expense: int = 0,
        balance: int = 0,
    ) -> None:
        values = {
            "contract": format_tl(contract),
            "collected": format_tl(collected),
            "expense": format_tl(expense),
            "balance": format_tl(balance),
        }
        for key, label in self._value_labels.items():
            label.setText(values.get(key, "0,00 ₺"))


def _install_header_menu(view: QTableView, title_provider: Callable[[int], str]) -> None:
    header = view.horizontalHeader()
    if header is None:
        return
    header.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)

    def _show_menu(position: QPoint) -> None:
        menu = QMenu(view)
        for visual_index in range(header.count()):
            logical_index = header.logicalIndex(visual_index)
            title = title_provider(logical_index) or f"Sütun {logical_index + 1}"
            action = menu.addAction(title)
            action.setCheckable(True)
            action.setChecked(not view.isColumnHidden(logical_index))
            action.toggled.connect(
                lambda checked, col=logical_index: view.setColumnHidden(col, not checked)
            )
        menu.exec(header.mapToGlobal(position))

    header.customContextMenuRequested.connect(_show_menu)


def _install_copy_shortcut(view: QTableView) -> None:
    shortcut = QShortcut(
        QKeySequence(QKeySequence.StandardKey.Copy), view
    )

    def _copy_selection() -> None:
        selection = view.selectionModel()
        if selection is None or not selection.hasSelection():
            return
        indexes = selection.selectedIndexes()
        if not indexes:
            return
        sorted_indexes = sorted(indexes, key=lambda idx: (idx.row(), idx.column()))
        rows: dict[int, list[QModelIndex]] = {}
        for index in sorted_indexes:
            rows.setdefault(index.row(), []).append(index)
        text_rows: list[str] = []
        for row in sorted(rows):
            columns = sorted(rows[row], key=lambda idx: idx.column())
            values = []
            for index in columns:
                data = index.data(Qt.ItemDataRole.DisplayRole)
                values.append(str(data) if data is not None else "")
            text_rows.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(text_rows))

    shortcut.activated.connect(_copy_selection)


class TableExportSelectionDialog(QDialog):
    """Tablo export için satır seçim dialogu.

    Aralık seçimi destekler: "3-6" veya "1,3,5-7" gibi.
    """

    def __init__(
        self,
        view: QTableView,
        parent: QWidget | None = None,
        *,
        title: str = "Dışa Aktarılacak Kayıtları Seç",
        display_columns: list[int] | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(600)
        self.setMinimumHeight(450)

        self._view = view
        self._model = view.model()
        self._display_columns = display_columns or [0, 1, 2]  # İlk 3 sütun varsayılan
        self._selected_rows: list[int] = []

        if self._model is None:
            return

        row_count = self._model.rowCount()

        layout = QVBoxLayout(self)

        # Bilgi etiketi
        info_label = QLabel(f"Toplam {row_count} kayıt bulundu. Dışa aktarmak istediklerinizi seçin:")
        layout.addWidget(info_label)

        # Aralık seçimi
        range_group = QGroupBox("Aralık Seçimi")
        range_layout = QHBoxLayout(range_group)
        range_layout.addWidget(QLabel("Aralık:"))
        self.range_edit = QLineEdit()
        self.range_edit.setPlaceholderText("Örn: 3-6 veya 1,3,5-7")
        self.range_edit.setMinimumWidth(150)
        self.range_edit.returnPressed.connect(self._apply_range)
        range_layout.addWidget(self.range_edit)
        self.apply_range_btn = QPushButton("Uygula")
        self.apply_range_btn.clicked.connect(self._apply_range)
        range_layout.addWidget(self.apply_range_btn)
        range_layout.addStretch()
        layout.addWidget(range_group)

        # Seçim butonları
        select_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("Tümünü Seç")
        self.select_all_btn.clicked.connect(self._select_all)
        select_layout.addWidget(self.select_all_btn)

        self.deselect_all_btn = QPushButton("Hiçbirini Seçme")
        self.deselect_all_btn.clicked.connect(self._deselect_all)
        select_layout.addWidget(self.deselect_all_btn)
        select_layout.addStretch()
        layout.addLayout(select_layout)

        # Kayıt listesi
        self.list_widget = QListWidget()
        for row_idx in range(row_count):
            # Display columns'dan metin oluştur
            parts = [str(row_idx + 1) + "."]
            for col in self._display_columns:
                if col < self._model.columnCount():
                    index = self._model.index(row_idx, col)
                    data = index.data(Qt.ItemDataRole.DisplayRole)
                    if data:
                        parts.append(str(data)[:30])  # Max 30 karakter
            item_text = " - ".join(parts)
            item = QListWidgetItem(item_text)
            item.setCheckState(Qt.CheckState.Checked)
            item.setData(Qt.ItemDataRole.UserRole, row_idx)
            self.list_widget.addItem(item)

        layout.addWidget(self.list_widget)

        # Dialog butonları
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def _select_all(self) -> None:
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item:
                item.setCheckState(Qt.CheckState.Checked)

    def _deselect_all(self) -> None:
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item:
                item.setCheckState(Qt.CheckState.Unchecked)

    def _apply_range(self) -> None:
        """Aralık metnini parse edip ilgili kayıtları seçer."""
        range_text = self.range_edit.text().strip()
        if not range_text:
            return

        self._deselect_all()

        selected_indices: set[int] = set()
        total = self.list_widget.count()

        parts = range_text.replace(" ", "").split(",")
        for part in parts:
            if "-" in part:
                try:
                    range_parts = part.split("-")
                    if len(range_parts) == 2:
                        start = int(range_parts[0])
                        end = int(range_parts[1])
                        if start > end:
                            start, end = end, start
                        for i in range(start, end + 1):
                            if 1 <= i <= total:
                                selected_indices.add(i - 1)
                except ValueError:
                    continue
            else:
                try:
                    num = int(part)
                    if 1 <= num <= total:
                        selected_indices.add(num - 1)
                except ValueError:
                    continue

        for idx in selected_indices:
            item = self.list_widget.item(idx)
            if item:
                item.setCheckState(Qt.CheckState.Checked)

        if selected_indices:
            self.range_edit.setStyleSheet("")
        else:
            self.range_edit.setStyleSheet("border: 1px solid red;")

    def get_selected_row_indices(self) -> list[int]:
        """Seçilen satır indekslerini döndürür."""
        selected = []
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item and item.checkState() == Qt.CheckState.Checked:
                row_idx = item.data(Qt.ItemDataRole.UserRole)
                if isinstance(row_idx, int):
                    selected.append(row_idx)
        return selected


def export_table_to_excel_with_selection(
    view: QTableView,
    filename: str,
    selected_rows: list[int] | None = None,
) -> None:
    """Export selected rows from a QTableView to Excel.

    If selected_rows is None, exports all rows.
    """
    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    columns = _gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return

    # Eğer seçim yoksa tüm satırları al
    if selected_rows is None:
        selected_rows = list(range(model.rowCount()))

    if not selected_rows:
        QMessageBox.warning(view, "Uyarı", "Dışa aktarılacak satır seçilmedi.")
        return

    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Veri"

        # Stiller
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="Calibri", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        col_widths = {}

        # Başlık satırı
        for col_idx, column in enumerate(columns, start=1):
            header_text = model.headerData(
                column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole
            )
            header_str = str(header_text or "")
            cell = sheet.cell(row=1, column=col_idx, value=header_str)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_widths[col_idx] = len(header_str) + 2

        # Seçili veri satırları
        for output_row, source_row in enumerate(selected_rows):
            for col_offset, column in enumerate(columns, start=1):
                index = model.index(source_row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                value_str = str(data or "")
                cell = sheet.cell(row=output_row + 2, column=col_offset, value=value_str)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                if output_row % 2 == 1:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

                current_width = col_widths.get(col_offset, 0)
                text_width = min(len(value_str), 50) + 2
                if text_width > current_width:
                    col_widths[col_offset] = text_width

        # Sütun genişlikleri
        for col_idx, width in col_widths.items():
            adjusted_width = max(8, min(width, 50))
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        sheet.row_dimensions[1].height = 25

        if selected_rows:
            last_col = get_column_letter(len(columns))
            sheet.auto_filter.ref = f"A1:{last_col}{len(selected_rows) + 1}"

        sheet.freeze_panes = "A2"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        workbook.save(filename)
        QMessageBox.information(
            view, "Başarılı", f"Excel dosyası oluşturuldu:\n{filename}\n({len(selected_rows)} kayıt)"
        )
    except Exception as exc:
        QMessageBox.critical(view, "Hata", f"Excel çıktısı oluşturulamadı:\n{exc}")


def export_table_to_pdf_with_selection(
    view: QTableView,
    filename: str,
    *,
    title: str,
    subtitle: str | None = None,
    selected_rows: list[int] | None = None,
) -> None:
    """Export selected rows from a QTableView to PDF."""
    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return

    columns = _gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return

    if selected_rows is None:
        selected_rows = list(range(model.rowCount()))

    if not selected_rows:
        QMessageBox.warning(view, "Uyarı", "Dışa aktarılacak satır seçilmedi.")
        return

    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)
        layout = QPageLayout(
            QPageSize(QPageSize.PageSizeId.A4),
            QPageLayout.Orientation.Landscape,
            QMarginsF(15, 15, 15, 15),
        )
        printer.setPageLayout(layout)
        page_rect = printer.pageRect(QPrinter.Unit.Point)
        page_width = float(page_rect.width())

        parts = [
            "<html><head>",
            "<style>",
            "body { font-family: 'DejaVu Sans', 'Arial', sans-serif; font-size: 10pt; color: #111; }",
            ".header { margin-bottom: 12px; }",
            ".header .title { font-size: 14pt; font-weight: bold; }",
            ".header .subtitle { font-size: 11pt; color: #444; }",
            ".header .meta { font-size: 10pt; color: #555; }",
            "table { border-collapse: collapse; width: 100%; font-size: 9pt; }",
            "th, td { border: 1px solid #666; padding: 5px 8px; text-align: left; }",
            "th { background-color: #2c3e50; color: white; font-weight: bold; }",
            "tr:nth-child(even) { background-color: #f9f9f9; }",
            "</style></head><body>",
        ]
        timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
        parts.append("<div class='header'>")
        parts.append(f"<div class='title'>{html.escape(title)}</div>")
        if subtitle:
            parts.append(f"<div class='subtitle'>{html.escape(subtitle)}</div>")
        parts.append(f"<div class='meta'>Oluşturma: {html.escape(timestamp)}</div>")
        parts.append("</div>")
        parts.append("<table><thead><tr>")

        for column in columns:
            header_text = model.headerData(
                column, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole
            )
            parts.append(f"<th>{html.escape(str(header_text or ''))}</th>")

        parts.append("</tr></thead><tbody>")

        for row in selected_rows:
            parts.append("<tr>")
            for column in columns:
                index = model.index(row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                parts.append(f"<td>{html.escape(str(data or ''))}</td>")
            parts.append("</tr>")

        parts.append("</tbody></table>")
        parts.append(f"<p style='font-size: 9pt; color: #666; margin-top: 10px;'>Toplam: {len(selected_rows)} kayıt</p>")
        parts.append("</body></html>")

        document = QTextDocument()
        document.setDocumentMargin(0)
        document.setHtml("".join(parts))
        document.setPageSize(QSizeF(page_width, float(page_rect.height())))
        document.print(printer)

        QMessageBox.information(
            view, "Başarılı", f"PDF dosyası oluşturuldu:\n{filename}\n({len(selected_rows)} kayıt)"
        )
    except Exception as exc:
        QMessageBox.critical(view, "Hata", f"PDF çıktısı oluşturulamadı:\n{exc}")


def _gather_visible_columns(view: QTableView) -> list[int]:
    model = view.model()
    if model is None:
        return []
    header = view.horizontalHeader()
    column_count = model.columnCount()
    if header is None:
        return [
            column
            for column in range(column_count)
            if not view.isColumnHidden(column)
        ]
    visible: list[int] = []
    for visual_index in range(header.count()):
        logical_index = header.logicalIndex(visual_index)
        if logical_index < 0 or logical_index >= column_count:
            continue
        if view.isColumnHidden(logical_index):
            continue
        visible.append(logical_index)
    return visible


def export_table_to_excel(view: QTableView, filename: str) -> None:
    """Export the visible contents of ``view`` into a well-formatted Excel workbook.

    Özellikler:
    - Otomatik sütun genişliği (içeriğe göre)
    - Başlık satırı stilleri (koyu arka plan, beyaz yazı, kalın)
    - Tablo çizgileri ve kenarlıklar
    - Otomatik filtre
    - Başlık satırı dondurma (freeze panes)
    """

    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    columns = _gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return
    rows = model.rowCount()
    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Veri"

        # Stiller
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="Calibri", size=10)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        # Sütun genişliklerini hesaplamak için maksimum uzunlukları takip et
        col_widths = {}

        # Başlık satırı
        for col_idx, column in enumerate(columns, start=1):
            header_text = model.headerData(
                column,
                Qt.Orientation.Horizontal,
                Qt.ItemDataRole.DisplayRole,
            )
            header_str = str(header_text or "")
            cell = sheet.cell(row=1, column=col_idx, value=header_str)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_widths[col_idx] = len(header_str) + 2

        # Veri satırları
        for row in range(rows):
            for col_offset, column in enumerate(columns, start=1):
                index = model.index(row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                value_str = str(data or "")
                cell = sheet.cell(row=row + 2, column=col_offset, value=value_str)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                # Satır renklendirme (alternatif satırlar)
                if row % 2 == 1:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

                # Sütun genişliği güncelle
                current_width = col_widths.get(col_offset, 0)
                # Uzun metinleri 50 karakter ile sınırla
                text_width = min(len(value_str), 50) + 2
                if text_width > current_width:
                    col_widths[col_offset] = text_width

        # Sütun genişliklerini uygula
        for col_idx, width in col_widths.items():
            # Minimum 8, maksimum 50 karakter genişlik
            adjusted_width = max(8, min(width, 50))
            sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Başlık satırı yüksekliği
        sheet.row_dimensions[1].height = 25

        # Otomatik filtre ekle
        if rows > 0:
            last_col = get_column_letter(len(columns))
            sheet.auto_filter.ref = f"A1:{last_col}{rows + 1}"

        # Başlık satırını dondur
        sheet.freeze_panes = "A2"

        # Sayfa ayarları
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        workbook.save(filename)
        QMessageBox.information(
            view,
            "Başarılı",
            f"Excel dosyası oluşturuldu:\n{filename}",
        )
    except Exception as exc:  # pragma: no cover - file system safety
        QMessageBox.critical(
            view,
            "Hata",
            f"Excel çıktısı oluşturulamadı:\n{exc}",
        )


def export_table_to_pdf(
    view: QTableView,
    filename: str,
    *,
    title: str,
    subtitle: str | None = None,
) -> None:
    """Render the table contents to a landscaped PDF document with header.

    Çok sayfalı PDF desteği ile okunabilir boyutta çıktı üretir.
    """

    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return
    columns = _gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return
    rows = model.rowCount()
    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        # Printer ayarları
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)
        layout = QPageLayout(
            QPageSize(QPageSize.PageSizeId.A4),
            QPageLayout.Orientation.Landscape,
            QMarginsF(15, 15, 15, 15),
        )
        printer.setPageLayout(layout)
        page_rect = printer.pageRect(QPrinter.Unit.Point)
        page_width = float(page_rect.width())

        # HTML içeriği oluştur
        parts = [
            "<html><head>",
            "<style>",
            "body { font-family: 'DejaVu Sans', 'Arial', sans-serif; font-size: 10pt; color: #111; }",
            ".header { margin-bottom: 12px; }",
            ".header .title { font-size: 14pt; font-weight: bold; }",
            ".header .subtitle { font-size: 11pt; color: #444; }",
            ".header .meta { font-size: 10pt; color: #555; }",
            "table { border-collapse: collapse; width: 100%; font-size: 9pt; }",
            "th, td { border: 1px solid #666; padding: 5px 8px; text-align: left; }",
            "th { background-color: #2c3e50; color: white; font-weight: bold; }",
            "tr:nth-child(even) { background-color: #f9f9f9; }",
            "</style></head><body>",
        ]
        timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
        parts.append("<div class='header'>")
        parts.append(f"<div class='title'>{html.escape(title)}</div>")
        if subtitle:
            parts.append(f"<div class='subtitle'>{html.escape(subtitle)}</div>")
        parts.append(f"<div class='meta'>Oluşturma: {html.escape(timestamp)}</div>")
        parts.append("</div>")
        parts.append("<table><thead><tr>")
        for column in columns:
            header_text = model.headerData(
                column,
                Qt.Orientation.Horizontal,
                Qt.ItemDataRole.DisplayRole,
            )
            parts.append(
                f"<th>{html.escape(str(header_text or ''))}</th>"
            )
        parts.append("</tr></thead><tbody>")
        for row in range(rows):
            parts.append("<tr>")
            for column in columns:
                index = model.index(row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                parts.append(f"<td>{html.escape(str(data or ''))}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table>")
        parts.append(f"<p style='font-size: 9pt; color: #666; margin-top: 10px;'>Toplam: {rows} kayıt</p>")
        parts.append("</body></html>")

        document = QTextDocument()
        document.setDocumentMargin(0)
        document.setHtml("".join(parts))
        # Sayfa genişliğini ayarla - bu otomatik sayfalama sağlar
        document.setPageSize(QSizeF(page_width, float(page_rect.height())))

        # PDF'e yazdır (çok sayfalı)
        document.print(printer)

        QMessageBox.information(
            view,
            "Başarılı",
            f"PDF dosyası oluşturuldu:\n{filename}",
        )
    except Exception as exc:  # pragma: no cover - file system safety
        QMessageBox.critical(
            view,
            "Hata",
            f"PDF çıktısı oluşturulamadı:\n{exc}",
        )


class OptionalDateEdit(QDateEdit):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setCalendarPopup(True)
        self.setDisplayFormat("dd.MM.yyyy")
        self.setSpecialValueText("—")
        self.setDateRange(OPTIONAL_DATE_MIN, OPTIONAL_DATE_MAX)
        self.clear()

    def keyPressEvent(self, event):  # type: ignore[override]
        if (
            event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace)
            and event.modifiers() == Qt.KeyboardModifier.NoModifier
        ):
            self.clear()
            event.accept()
            return
        super().keyPressEvent(event)


class OptionalDateDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):  # type: ignore[override]
        return OptionalDateEdit(parent)

    def setEditorData(self, editor, index):  # type: ignore[override]
        raw_value = index.data(Qt.ItemDataRole.EditRole)
        target: QDate | None = None
        if isinstance(raw_value, QDate):
            target = raw_value if raw_value.isValid() else None
        elif isinstance(raw_value, datetime):
            target = QDate(raw_value.year, raw_value.month, raw_value.day)
        elif isinstance(raw_value, date):
            target = QDate(raw_value.year, raw_value.month, raw_value.day)
        elif isinstance(raw_value, str):
            text = raw_value.strip()
            if text:
                parsed = QDate.fromString(text, "yyyy-MM-dd")
                if not parsed.isValid():
                    parsed = QDate.fromString(text, "dd.MM.yyyy")
                if parsed.isValid():
                    target = parsed
        if target is None:
            editor.setDate(QDate.currentDate())
        else:
            editor.setDate(target)

    def setModelData(self, editor, model, index):  # type: ignore[override]
        model.setData(index, editor.date(), Qt.ItemDataRole.EditRole)


class TodoDateDelegate(QStyledItemDelegate):
    """Görevler tablosunda tarih düzenleme için takvim popup ile delegate."""

    def createEditor(self, parent, option, index):  # type: ignore[override]
        editor = QDateEdit(parent)
        editor.setCalendarPopup(True)
        editor.setDisplayFormat("dd.MM.yyyy")
        editor.setDateRange(QDate(2000, 1, 1), QDate(2100, 12, 31))
        return editor

    def setEditorData(self, editor, index):  # type: ignore[override]
        raw_value = index.data(Qt.ItemDataRole.DisplayRole)
        target: QDate | None = None

        if isinstance(raw_value, QDate):
            target = raw_value if raw_value.isValid() else None
        elif isinstance(raw_value, datetime):
            target = QDate(raw_value.year, raw_value.month, raw_value.day)
        elif isinstance(raw_value, date):
            target = QDate(raw_value.year, raw_value.month, raw_value.day)
        elif isinstance(raw_value, str):
            text = raw_value.strip()
            if text and text != "Tarihsiz":
                # Türkçe format: dd.MM.yyyy
                parsed = QDate.fromString(text, "dd.MM.yyyy")
                if not parsed.isValid():
                    # ISO format: yyyy-MM-dd
                    parsed = QDate.fromString(text, "yyyy-MM-dd")
                if parsed.isValid():
                    target = parsed

        if target is None or not target.isValid():
            editor.setDate(QDate.currentDate())
        else:
            editor.setDate(target)

    def setModelData(self, editor, model, index):  # type: ignore[override]
        qdate = editor.date()
        if qdate.isValid():
            # Türkçe formatında göster
            display_text = qdate.toString("dd.MM.yyyy")
            model.setData(index, display_text, Qt.ItemDataRole.DisplayRole)

    def updateEditorGeometry(self, editor, option, index):  # type: ignore[override]
        editor.setGeometry(option.rect)


class QuickDatePopup(QFrame):
    dateSelected = pyqtSignal(QDate)
    cleared = pyqtSignal()
    cancelled = pyqtSignal()

    def __init__(self, parent: QWidget | None = None, *, initial_date: QDate | None = None):
        super().__init__(parent, Qt.WindowType.Popup)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setFrameShadow(QFrame.Shadow.Raised)
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self._closing = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(4)

        self.date_edit = QDateEdit(self)
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        self.date_edit.setDateRange(OPTIONAL_DATE_MIN, OPTIONAL_DATE_MAX)
        today = QDate.currentDate()
        if initial_date and initial_date.isValid():
            self.date_edit.setDate(initial_date)
        else:
            self.date_edit.setDate(today)
        layout.addWidget(self.date_edit)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(0, 0, 0, 0)
        btn_row.setSpacing(4)
        layout.addLayout(btn_row)

        def make_button(text: str, handler: Callable[[], None]) -> QPushButton:
            button = QPushButton(text, self)
            button.clicked.connect(handler)
            btn_row.addWidget(button)
            return button

        make_button("Yok", self._on_clear)
        make_button("Yarın", lambda: self._emit_with_offset(1))
        make_button("3 gün sonra", lambda: self._emit_with_offset(3))
        make_button("Haftaya", lambda: self._emit_with_offset(7))

        self.date_edit.dateChanged.connect(self._on_date_changed)

    def _emit_with_offset(self, days: int) -> None:
        self._closing = True
        self.dateSelected.emit(QDate.currentDate().addDays(days))
        self.close()

    def _on_clear(self) -> None:
        self._closing = True
        self.cleared.emit()
        self.close()

    def _on_date_changed(self, date: QDate) -> None:
        if not date or not date.isValid():
            return
        self._closing = True
        self.dateSelected.emit(date)
        self.close()

    def focusOutEvent(self, event):  # type: ignore[override]
        super().focusOutEvent(event)
        if not self._closing:
            self.cancelled.emit()
            self.close()

    def keyPressEvent(self, event):  # type: ignore[override]
        if event.key() == Qt.Key.Key_Escape:
            self._closing = True
            self.cancelled.emit()
            self.close()
            return
        super().keyPressEvent(event)


class TurkishFilterProxyModel(QSortFilterProxyModel):
    """Türkçe karakter duyarlı filtreleme yapan proxy model."""

    @staticmethod
    def turkish_normalize(text: str) -> str:
        """Türkçe karakterleri arama için normalize eder.

        i/İ/ı/I hepsini aynı karaktere dönüştürür böylece
        kullanıcı 'i' yazsa da 'ı' yazsa da 'İstinaf' bulunur.
        """
        # Önce büyük Türkçe harfleri küçüğe çevir
        result = text.replace("İ", "i").replace("I", "i")  # I ve İ → i
        result = result.replace("Ğ", "ğ").replace("Ü", "ü").replace("Ş", "ş")
        result = result.replace("Ö", "ö").replace("Ç", "ç")
        result = result.lower()
        # ı'yı da i'ye çevir (arama kolaylığı için)
        result = result.replace("ı", "i")
        return result

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        """Türkçe karakter duyarlı filtreleme."""
        pattern = self.filterRegularExpression().pattern()
        if not pattern:
            return True

        index = self.sourceModel().index(source_row, 0, source_parent)
        item_text = index.data(Qt.ItemDataRole.DisplayRole) or ""

        pattern_norm = self.turkish_normalize(pattern)
        item_norm = self.turkish_normalize(item_text)

        return pattern_norm in item_norm


class StatusDelegate(QStyledItemDelegate):
    def __init__(
        self,
        parent: QWidget | None = None,
        *,
        items: list[str] | None = None,
    ) -> None:
        super().__init__(parent)
        self._items = items or []

    def createEditor(self, parent, option, index):  # type: ignore[override]
        combo = QComboBox(parent)
        combo.setEditable(True)
        combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        combo.addItems(self._items)
        # Dropdown okunu görünür yap
        combo.setStyleSheet("""
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid #999;
            }
            QComboBox::down-arrow {
                width: 10px;
                height: 10px;
            }
        """)

        # Türkçe karakter duyarlı completer (utils.py'deki TurkishCompleter mantığı)
        source_model = QStringListModel(self._items, combo)
        proxy_model = TurkishFilterProxyModel(combo)
        proxy_model.setSourceModel(source_model)

        completer = QCompleter(combo)
        completer.setModel(proxy_model)
        # UnfilteredPopupCompletion: Qt'nin kendi filtrelemesini devre dışı bırak
        # Sadece bizim TurkishFilterProxyModel filtrelesin
        completer.setCompletionMode(QCompleter.CompletionMode.UnfilteredPopupCompletion)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)

        # Kullanıcı yazarken filtrelemeyi güncelle
        # NOT: textEdited kullanılıyor (textChanged değil) çünkü ok tuşuyla
        # seçenekler arasında gezinirken filtreleme güncellenmemeli
        line_edit = combo.lineEdit()
        if line_edit:
            def update_filter(text: str) -> None:
                proxy_model.setFilterRegularExpression(text)
                # Qt'nin kendi prefix filtrelemesini bypass et
                completer.setCompletionPrefix("")
            line_edit.textEdited.connect(update_filter)

        combo.setCompleter(completer)
        self._connect_return_key(combo)
        return combo

    def setEditorData(self, editor, index):  # type: ignore[override]
        value = index.data(Qt.ItemDataRole.DisplayRole) or ""
        editor.setCurrentText(str(value))

    def setModelData(self, editor, model, index):  # type: ignore[override]
        text = editor.currentText().strip()
        model.setData(index, text, Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):  # type: ignore[override]
        editor.setGeometry(option.rect)

    def _connect_return_key(self, combo: QComboBox) -> None:
        """Commit inline status edits when Enter is pressed.

        Otomatik tamamlama: Tek eşleşen seçenek varsa onu seç.
        Türkçe karakter desteği: i/İ ve ı/I doğru şekilde eşleştirilir.
        """

        line_edit = combo.lineEdit()
        if line_edit is None:
            return

        delegate = self  # Closure için

        def _turkish_lower(text: str) -> str:
            """Türkçe karakterleri arama için normalize eder."""
            result = text.replace("İ", "i").replace("I", "i")
            result = result.replace("Ğ", "ğ").replace("Ü", "ü").replace("Ş", "ş")
            result = result.replace("Ö", "ö").replace("Ç", "ç")
            result = result.lower()
            result = result.replace("ı", "i")
            return result

        def _find_best_match(search_text: str) -> str | None:
            """Aranan metne en uygun eşleşmeyi bul."""
            if not search_text:
                return None

            search_lower = _turkish_lower(search_text)

            # 1. Tam eşleşme kontrol et
            for item in delegate._items:
                if _turkish_lower(item) == search_lower:
                    return item

            # 2. İçeren eşleşmeleri bul
            contains_matches = [
                item for item in delegate._items
                if search_lower in _turkish_lower(item)
            ]

            # 3. Prefix eşleşmeleri bul
            prefix_matches = [
                item for item in delegate._items
                if _turkish_lower(item).startswith(search_lower)
            ]

            # Öncelik: tek eşleşme > ilk prefix > ilk contains
            if len(contains_matches) == 1:
                return contains_matches[0]
            if len(prefix_matches) == 1:
                return prefix_matches[0]
            if prefix_matches:
                return prefix_matches[0]
            if contains_matches:
                return contains_matches[0]

            return None

        def _do_commit() -> None:
            """Enter tuşuna basıldığında seçimi uygula."""
            current_text = line_edit.text().strip()

            # Önce manuel eşleşmeyi dene - bu her zaman çalışır
            if current_text:
                best_match = _find_best_match(current_text)
                if best_match:
                    combo.setCurrentText(best_match)

            delegate.commitData.emit(combo)
            delegate.closeEditor.emit(combo, QAbstractItemDelegate.EndEditHint.NoHint)

        # Completer popup'ından seçim yapıldığında
        completer = combo.completer()
        if completer:
            def _on_activated(text: str) -> None:
                combo.setCurrentText(text)
                delegate.commitData.emit(combo)
                delegate.closeEditor.emit(combo, QAbstractItemDelegate.EndEditHint.NoHint)
            completer.activated.connect(_on_activated)

        # returnPressed sinyalini bağla
        line_edit.returnPressed.connect(_do_commit)

class LexTabBar(QTabBar):
    tabDoubleClicked = pyqtSignal(int)

    def mouseDoubleClickEvent(self, event):  # type: ignore[override]
        index = self.tabAt(event.position().toPoint())
        if index >= 0:
            self.tabDoubleClicked.emit(index)
        super().mouseDoubleClickEvent(event)


class DosyaTableModel(QAbstractTableModel):
    DATA_CALLS = 0
    ROLE_HITS: Counter = Counter()
    COL_SELECTION = COL_SELECTION
    COL_BN = COL_BN
    COL_DURUSMA_TARIHI = COL_DURUSMA_TARIHI
    COL_DAVA_DURUMU = COL_DAVA_DURUMU
    COL_IS_TARIHI = COL_IS_TARIHI
    COL_DAVA_DURUMU_2 = COL_DAVA_DURUMU_2
    COL_IS_TARIHI_2 = COL_IS_TARIHI_2
    DATE_EDITABLE_COLUMNS = (
        COL_DURUSMA_TARIHI,
        COL_IS_TARIHI,
        COL_IS_TARIHI_2,
    )
    _STATUS_COLUMNS = (COL_DAVA_DURUMU, COL_DAVA_DURUMU_2)
    EDITABLE_FIELDS = {
        COL_DURUSMA_TARIHI: "durusma_tarihi",
        COL_IS_TARIHI: "is_tarihi",
        COL_IS_TARIHI_2: "is_tarihi_2",
        COL_DAVA_DURUMU: "dava_durumu",
        COL_DAVA_DURUMU_2: "tekrar_dava_durumu_2",
        # NOT: Açıklama sütunları Qt editor çakışması nedeniyle inline edit'ten çıkarıldı
        # Açıklamalar diyalog üzerinden düzenlenebilir
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        load_status_palette()
        self.headers = [
            "",
            "BN",
            "Dosya Esas No",
            "Müvekkil Adı",
            "Karşı Taraf",
            "Dosya Konusu",
            "Mahkeme Adı",
            "Duruşma Tarihi",
            "Dava Durumu",
            "İş Tarihi",
            "Açıklama",
            "Dava Durumu 2",
            "İş Tarihi 2",
            "Açıklama 2",
        ]
        self.keys = [
            None,
            "buro_takip_no",
            "dosya_esas_no",
            "muvekkil_adi",
            "karsi_taraf",
            "dosya_konusu",
            "mahkeme_adi",
            "durusma_tarihi",
            "dava_durumu",
            "is_tarihi",
            "aciklama",
            "tekrar_dava_durumu_2",
            "is_tarihi_2",
            "aciklama_2",
        ]
        self.records: list[dict[str, Any]] = []
        self._rows: list[dict[int, dict[Any, object]]] = []
        self._attached_view: QAbstractItemView | None = None

    def attach_view(self, view: QAbstractItemView | None) -> None:
        self._attached_view = view

    def set_records(self, records: list[dict[str, Any]]) -> None:
        prepared = self.prepare_records(records)
        self.apply_prepared_records(prepared)

    def prepare_records(self, records: list[dict[str, Any]]) -> list[PreparedDosyaRow]:
        prepared: list[PreparedDosyaRow] = []
        today = date.today()
        for record in records:
            record_copy = dict(record)
            row_cache = self._build_row_cache(record_copy, today=today)
            prepared.append((record_copy, row_cache))
        return prepared

    def apply_prepared_records(self, prepared: list[PreparedDosyaRow]) -> None:
        view = self._attached_view
        table_view = view if isinstance(view, QTableView) else None
        sorting_was_enabled = table_view.isSortingEnabled() if table_view else False
        updates_disabled = False
        if view is not None:
            try:
                view.setUpdatesEnabled(False)
                updates_disabled = True
            except Exception:
                pass
        if table_view is not None:
            table_view.setSortingEnabled(False)
        self.beginResetModel()
        self.records = [record for record, _ in prepared]
        self._rows = [cells for _, cells in prepared]
        self.endResetModel()
        if table_view is not None:
            table_view.setSortingEnabled(sorting_was_enabled)
        if view is not None and updates_disabled:
            try:
                view.setUpdatesEnabled(True)
            except Exception:
                pass

    def record_at(self, row: int) -> dict | None:
        if 0 <= row < len(self.records):
            return self.records[row]
        return None

    def rowCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self._rows)

    def columnCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self.headers)

    def flags(self, index):  # type: ignore[override]
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
        base_flags = super().flags(index)
        if index.column() in self.EDITABLE_FIELDS:
            return base_flags | Qt.ItemFlag.ItemIsEditable
        return base_flags

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        type(self).DATA_CALLS += 1
        type(self).ROLE_HITS[role] += 1
        dc = type(self).DATA_CALLS
        if dc % 5000 == 0:
            print(f"[perf] data() calls: {dc}")
        display_hits = type(self).ROLE_HITS[Qt.ItemDataRole.DisplayRole]
        if (
            role == Qt.ItemDataRole.DisplayRole
            and display_hits % 20000 == 0
            and display_hits
        ):
            print(f"[perf] DisplayRole hits: {display_hits}")
        if not index.isValid():
            return None
        if not (0 <= index.row() < len(self._rows)):
            return None
        row_cache = self._rows[index.row()]
        cell = row_cache.get(index.column())
        if not cell:
            return None
        if role == Qt.ItemDataRole.BackgroundRole:
            bg_value = self._background_data(index, cell)
            if bg_value is not None:
                return bg_value
        if role == Qt.ItemDataRole.ForegroundRole:
            fg_value = self._foreground_data(index, cell)
            if fg_value is not None:
                return fg_value
        if role in cell:
            return cell[role]
        return None

    def _field_to_column(self, field_name: str) -> int | None:
        """Alan adından sütun numarasını bul."""
        try:
            return self.keys.index(field_name)
        except ValueError:
            return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):  # type: ignore[override]
        if role != Qt.ItemDataRole.EditRole or not index.isValid():
            return False
        field = self.EDITABLE_FIELDS.get(index.column())
        if not field:
            return False
        record = self.record_at(index.row())
        if not record:
            return False
        record_id = record.get("id")
        try:
            normalized_id = int(record_id)
        except (TypeError, ValueError):
            return False
        iso_value = self._normalize_editor_value(value)
        existing_value = record.get(field) or None
        if existing_value == iso_value:
            return False
        payload = {
            field: iso_value,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        tracked_state = {
            "dava_durumu": record.get("dava_durumu"),
            "dava_durumu_2": record.get("tekrar_dava_durumu_2"),
            "aciklama": record.get("aciklama"),
            "aciklama_2": record.get("aciklama_2"),
            "is_tarihi": record.get("is_tarihi"),
            "is_tarihi_2": record.get("is_tarihi_2"),
            "durusma_tarihi": record.get("durusma_tarihi"),
        }
        parent = self.parent()
        current_user = getattr(parent, "current_user", {}) or {}
        user_name = (
            current_user.get("username")
            or current_user.get("kullanici_adi")
            or ""
        )
        try:
            update_dosya_with_auto_timeline(
                normalized_id, payload, user_name, original_state=tracked_state
            )
        except RuntimeError as exc:
            self._show_update_error(str(exc))
            return False
        except Exception as exc:
            self._show_update_error(f"Alan güncellenemedi:\n{exc}")
            return False
        record[field] = iso_value
        record["updated_at"] = payload["updated_at"]

        # ADIM 1 & 2: Dava durumu değiştiğinde is_tarihi ve aciklama sıfırlandıysa tabloya yansıt
        extra_updated_columns: list[int] = []
        if field == "dava_durumu":
            old_dava = tracked_state.get("dava_durumu") or ""
            new_dava = iso_value or ""
            # Boşa çekildi veya değişti → is_tarihi ve aciklama sıfırlandı
            if not new_dava or (old_dava and old_dava != new_dava):
                record["is_tarihi"] = None
                record["aciklama"] = None
                is_tarihi_col = self._field_to_column("is_tarihi")
                if is_tarihi_col is not None:
                    extra_updated_columns.append(is_tarihi_col)
                aciklama_col = self._field_to_column("aciklama")
                if aciklama_col is not None:
                    extra_updated_columns.append(aciklama_col)
        elif field == "tekrar_dava_durumu_2":
            old_dava2 = tracked_state.get("dava_durumu_2") or ""
            new_dava2 = iso_value or ""
            if not new_dava2 or (old_dava2 and old_dava2 != new_dava2):
                record["is_tarihi_2"] = None
                record["aciklama_2"] = None
                is_tarihi_2_col = self._field_to_column("is_tarihi_2")
                if is_tarihi_2_col is not None:
                    extra_updated_columns.append(is_tarihi_2_col)
                aciklama_2_col = self._field_to_column("aciklama_2")
                if aciklama_2_col is not None:
                    extra_updated_columns.append(aciklama_2_col)

        cell = self._build_cell_roles(field, record, today=date.today())
        self._rows[index.row()][index.column()] = cell

        # Ekstra güncellenen sütunları da yenile
        for col in extra_updated_columns:
            extra_field = self.keys[col] if col < len(self.keys) else None
            if extra_field:
                extra_cell = self._build_cell_roles(extra_field, record, today=date.today())
                self._rows[index.row()][col] = extra_cell

        self.dataChanged.emit(
            index,
            index,
            [
                Qt.ItemDataRole.DisplayRole,
                Qt.ItemDataRole.UserRole,
                Qt.ItemDataRole.EditRole,
                Qt.ItemDataRole.BackgroundRole,
                Qt.ItemDataRole.ForegroundRole,
            ],
        )

        # Ekstra sütunlar için dataChanged'i ertele (editor kapatıldıktan sonra)
        if extra_updated_columns:
            def emit_extra_changes():
                for col in extra_updated_columns:
                    extra_index = self.index(index.row(), col)
                    self.dataChanged.emit(
                        extra_index,
                        extra_index,
                        [
                            Qt.ItemDataRole.DisplayRole,
                            Qt.ItemDataRole.UserRole,
                            Qt.ItemDataRole.EditRole,
                            Qt.ItemDataRole.BackgroundRole,
                            Qt.ItemDataRole.ForegroundRole,
                        ],
                    )
            QTimer.singleShot(0, emit_extra_changes)

        return True

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.headers[section]
        return super().headerData(section, orientation, role)

    def sort(self, column, order):  # type: ignore[override]
        if not (0 <= column < len(self.keys)):
            return
        reverse = order == Qt.SortOrder.DescendingOrder
        combined = list(zip(self.records, self._rows))
        sort_role = Qt.ItemDataRole.UserRole
        display_role = Qt.ItemDataRole.DisplayRole

        def sort_key(item: tuple[dict[str, Any], dict[int, dict[Qt.ItemDataRole, object]]]):
            cell = item[1].get(column) or {}
            value = cell.get(sort_role)
            if isinstance(value, (int, float)):
                return value
            if isinstance(value, str):
                return value
            return str(cell.get(display_role, ""))

        self.layoutAboutToBeChanged.emit()
        combined.sort(key=sort_key, reverse=reverse)
        self.records = [record for record, _ in combined]
        self._rows = [cells for _, cells in combined]
        self.layoutChanged.emit()

    def _background_data(
        self, index: QModelIndex, cell: dict[Any, object]
    ) -> QBrush | None:
        column = index.column()
        record = self.record_at(index.row())
        # NOTE: restored duruşma/dava status coloring logic using get_durusma_color/get_status_color
        if column == self.COL_DURUSMA_TARIHI and record:
            bg_brush, _ = _durusma_color_roles(record.get("durusma_tarihi"))
            if bg_brush is not None:
                return bg_brush
        if record and column in (self.COL_IS_TARIHI, self.COL_IS_TARIHI_2):
            key = self.keys[column]
            brush = _job_date_background(record.get(key))
            if brush is not None:
                return brush
        if column in self._STATUS_COLUMNS:
            status_key = cell.get(_STATUS_COLOR_META)
            if status_key:
                if status_key in STATUS_BRUSHES:
                    return STATUS_BRUSHES.get(status_key)
                brush, _ = _ensure_status_palette_entry(status_key)
                return brush
        return None

    def _foreground_data(
        self, index: QModelIndex, cell: dict[Any, object]
    ) -> QColor | QBrush | None:
        column = index.column()
        record = self.record_at(index.row())
        if column in (self.COL_IS_TARIHI, self.COL_IS_TARIHI_2):
            return QBrush(Qt.GlobalColor.black)
        if column == self.COL_DURUSMA_TARIHI and record:
            _, fg_color = _durusma_color_roles(record.get("durusma_tarihi"))
            if fg_color is not None:
                return fg_color
        if column in self._STATUS_COLUMNS:
            status_key = cell.get(_STATUS_COLOR_META)
            if status_key:
                if status_key not in STATUS_FG:
                    _, fg_color = _ensure_status_palette_entry(status_key)
                    return fg_color
                return STATUS_FG.get(status_key, DEFAULT_STATUS_FG)
        return None

    def _build_row_cache(
        self, record: dict[str, Any], *, today: date | None = None
    ) -> dict[int, dict[Any, object]]:
        row_cache: dict[int, dict[Any, object]] = {}
        for column, key in enumerate(self.keys):
            cell = self._build_cell_roles(key, record, today=today)
            if cell:
                row_cache[column] = cell
        return row_cache

    def _build_cell_roles(
        self, key: str | None, record: dict[str, Any], *, today: date | None = None
    ) -> dict[Any, object]:
        cell: dict[Any, object] = {}
        if key is None:
            cell[Qt.ItemDataRole.DisplayRole] = _ARROW_TEXT
            cell[Qt.ItemDataRole.TextAlignmentRole] = _CENTER_ALIGNMENT
            cell[Qt.ItemDataRole.FontRole] = _FONT_BOLD
            cell[Qt.ItemDataRole.UserRole] = ""
            return cell
        raw_value = record.get(key)
        if key == "buro_takip_no":
            bn_text = _normalize_text_value(raw_value)
            cell[Qt.ItemDataRole.DisplayRole] = bn_text
            cell[Qt.ItemDataRole.UserRole] = _int_sort_key(raw_value)
            # NOTE: ensure BN sorts numerically via UserRole
            return cell
        if key in _DATE_FIELDS:
            date_value = coerce_to_date(raw_value)
            cell[Qt.ItemDataRole.DisplayRole] = (
                date_value.strftime("%d.%m.%Y") if date_value else ""
            )
            cell[Qt.ItemDataRole.UserRole] = _date_sort_key(date_value)
            cell[Qt.ItemDataRole.EditRole] = raw_value or ""
        else:
            text_value = _normalize_text_value(raw_value)
            if key == "muvekkil_adi":
                text_value = _with_role_abbreviation(
                    text_value, record.get("muvekkil_rolu")
                )
            cell[Qt.ItemDataRole.DisplayRole] = text_value
            cell[Qt.ItemDataRole.UserRole] = text_value.casefold() if text_value else ""
        if key in _STATUS_COLOR_FIELDS:
            status_label = _normalize_text_value(record.get(key))
            if status_label:
                cell[_STATUS_COLOR_META] = status_label
        return cell

    def _normalize_editor_value(self, value: Any) -> Optional[str]:
        if isinstance(value, QDate):
            if not value.isValid() or value == OPTIONAL_DATE_MIN:
                return None
            return value.toString("yyyy-MM-dd")
        if value in (None, ""):
            return None
        text = str(value).strip()
        if not text:
            return None
        parsed = QDate.fromString(text, "yyyy-MM-dd")
        if parsed.isValid():
            return parsed.toString("yyyy-MM-dd")
        return text

    def _show_update_error(self, message: str) -> None:
        view = self._attached_view
        parent = view.window() if isinstance(view, QWidget) else view
        QMessageBox.critical(parent or None, "Hata", message)


class CustomTabProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.setSortCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.setDynamicSortFilter(True)
        self.setSortRole(Qt.ItemDataRole.UserRole)
        self.allowed_ids: set[int] | None = None
        self.col_id: int | None = None
        self._open_only: bool = False

    def set_allowed_ids(self, ids: Optional[Iterable[int]]) -> None:
        if ids is None:
            self.allowed_ids = None
        else:
            normalized: set[int] = set()
            for value in ids:
                try:
                    normalized.add(int(value))
                except (TypeError, ValueError):
                    continue
            self.allowed_ids = normalized
        self.invalidateFilter()

    def set_id_column(self, column: int | None) -> None:
        self.col_id = column

    def set_open_only(self, enabled: bool) -> None:
        if self._open_only == bool(enabled):
            return
        self._open_only = bool(enabled)
        self.invalidateFilter()

    def filterAcceptsRow(
        self, source_row: int, source_parent: QModelIndex
    ) -> bool:  # type: ignore[override]
        model = self.sourceModel()
        record = None
        if model is not None and hasattr(model, "record_at"):
            record = model.record_at(source_row)  # type: ignore[attr-defined]
        if self.allowed_ids is not None:
            record_id: Optional[int] = None
            if record is not None:
                record_id = record.get("id")
            elif model is not None and self.col_id is not None:
                index = model.index(source_row, self.col_id, source_parent)
                value = model.data(index, Qt.ItemDataRole.DisplayRole)
                try:
                    record_id = int(value)
                except (TypeError, ValueError):
                    record_id = None
            if record_id is None:
                return False
            try:
                normalized_id = int(record_id)
            except (TypeError, ValueError):
                return False
            if normalized_id not in self.allowed_ids:
                return False
        if self._open_only and record is not None:
            if is_case_closed(record):
                return False
        return super().filterAcceptsRow(source_row, source_parent)

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:  # type: ignore[override]
        model = self.sourceModel()
        if model is None:
            return super().lessThan(left, right)
        left_value = model.data(left, Qt.ItemDataRole.UserRole)
        right_value = model.data(right, Qt.ItemDataRole.UserRole)
        if isinstance(left_value, (int, float)) and isinstance(right_value, (int, float)):
            return left_value < right_value
        left_display = model.data(left, Qt.ItemDataRole.DisplayRole) or ""
        right_display = model.data(right, Qt.ItemDataRole.DisplayRole) or ""
        return str(left_display) < str(right_display)


class DosyalarTab(QWidget):
    new_requested = pyqtSignal()
    edit_requested = pyqtSignal()
    archive_requested = pyqtSignal()
    attachments_requested = pyqtSignal()
    vekalet_requested = pyqtSignal()
    settings_requested = pyqtSignal()
    row_double_clicked = pyqtSignal(QModelIndex)
    filters_changed = pyqtSignal()
    clear_filters_requested = pyqtSignal()
    refresh_requested = pyqtSignal()

    def __init__(
        self,
        current_user: dict,
        *,
        mode: Literal["main", "custom"] = "main",
        custom_tab_id: Optional[int] = None,
        only_own_records: bool = False,
        can_manage_assignments: bool = False,
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(parent)
        self.current_user = current_user or {}
        self.current_user_id = self.current_user.get("id")
        self.permissions: dict[str, bool] = (
            self.current_user.get("permissions", {}) if self.current_user else {}
        ) or {}
        self.mode = mode
        self.custom_tab_id = custom_tab_id
        self.only_own_records = only_own_records
        self.can_manage_assignments = can_manage_assignments

        self.search_input: QLineEdit | None = None
        self.status_filter_combo: QComboBox | None = None
        self.durusma_filter_combo: QComboBox | None = None
        self.open_only_checkbox: QCheckBox | None = None
        self.assignment_checkbox: QCheckBox | None = None
        self.user_filter_combo: QComboBox | None = None
        self.clear_filters_button: QToolButton | None = None
        self._date_delegate: OptionalDateDelegate | None = None
        self._quick_date_popup: QuickDatePopup | None = None
        self._status_delegate: StatusDelegate | None = None

        self.table_model = DosyaTableModel(self)
        self.proxy = CustomTabProxyModel(self)
        self.proxy.setSourceModel(self.table_model)
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(300)
        self._search_timer.timeout.connect(self._emit_filters_changed)

        self._setup_ui()
        self.update_assignment_checkbox_state()
        if self.status_filter_combo is not None:
            self.populate_status_filter()
        if self.user_filter_combo is not None:
            self.populate_user_filter()

    def update_user_context(
        self,
        current_user: dict | None = None,
        permissions: dict[str, bool] | None = None,
        only_own_records: bool | None = None,
        can_manage_assignments: bool | None = None,
        **kwargs,
    ) -> None:
        """Refresh the cached user/permission context for the tab."""

        if current_user is not None:
            self.current_user = current_user or {}
            self.current_user_id = self.current_user.get("id")
        if permissions is not None:
            self.permissions = permissions or {}
        if only_own_records is not None:
            self.only_own_records = only_own_records
        if can_manage_assignments is not None:
            self.can_manage_assignments = can_manage_assignments

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)

        show_filters = self.mode == "main"
        if show_filters:
            filters_layout = QHBoxLayout()

            self.search_input = QLineEdit()
            self.search_input.setPlaceholderText("Ara…")
            self.search_input.textChanged.connect(self._on_search_text_changed)
            filters_layout.addWidget(QLabel("Ara:"))
            filters_layout.addWidget(self.search_input)

            # NOTE: toolbar modernization with inline clear + alerts controls
            self.clear_filters_button = QToolButton()
            self.clear_filters_button.setObjectName("clearFiltersButton")
            self.clear_filters_button.setText("Temizle")
            self.clear_filters_button.setCursor(
                Qt.CursorShape.PointingHandCursor
            )
            self.clear_filters_button.setToolTip(
                "Filtreleri temizle (Ctrl+Backspace)"
            )
            self.clear_filters_button.clicked.connect(self.clear_all_filters)
            filters_layout.addWidget(self.clear_filters_button)

            self.refresh_button = QToolButton()
            self.refresh_button.setObjectName("refreshButton")
            self.refresh_button.setText("Yenile")
            self.refresh_button.setCursor(Qt.CursorShape.PointingHandCursor)
            self.refresh_button.setToolTip("Tabloyu yenile (F5)")
            self.refresh_button.clicked.connect(self._on_refresh_clicked)
            filters_layout.addWidget(self.refresh_button)

            filters_layout.addSpacing(12)

            self.status_filter_combo = QComboBox()
            self.status_filter_combo.currentIndexChanged.connect(
                self._emit_filters_changed
            )
            filters_layout.addWidget(QLabel("Durum:"))
            filters_layout.addWidget(self.status_filter_combo)


            self.durusma_filter_combo = QComboBox()
            self.durusma_filter_combo.addItem("Hepsi", None)
            self.durusma_filter_combo.addItem("Bu Hafta", "bu_hafta")
            self.durusma_filter_combo.addItem("Önümüzdeki Hafta", "gelecek_hafta")
            self.durusma_filter_combo.addItem("Bu Ay", "bu_ay")
            self.durusma_filter_combo.currentIndexChanged.connect(
                self._emit_filters_changed
            )
            filters_layout.addWidget(QLabel("Duruşma:"))
            filters_layout.addWidget(self.durusma_filter_combo)

            self.open_only_checkbox = QCheckBox("Sadece açık dosyalar")
            self.open_only_checkbox.toggled.connect(self._emit_filters_changed)
            filters_layout.addWidget(self.open_only_checkbox)

            self.assignment_checkbox = QCheckBox("Sadece bana atananlar")
            self.assignment_checkbox.toggled.connect(self._emit_filters_changed)
            filters_layout.addWidget(self.assignment_checkbox)

            if self.can_manage_assignments:
                self.user_filter_combo = QComboBox()
                self.user_filter_combo.currentIndexChanged.connect(
                    self._emit_filters_changed
                )
                filters_layout.addWidget(QLabel("Kullanıcı:"))
                filters_layout.addWidget(self.user_filter_combo)

            filters_layout.addStretch(1)
            layout.addLayout(filters_layout)

        self.table_view = QTableView(self)
        self.table_model.attach_view(self.table_view)
        self.table_view.setModel(self.proxy)
        self.table_view.setSelectionBehavior(
            QTableView.SelectionBehavior.SelectRows
        )
        self.table_view.setSelectionMode(
            QTableView.SelectionMode.SingleSelection
        )
        self.table_view.setSortingEnabled(True)
        self.table_view.setWordWrap(False)
        # Dava durumu gibi düzenlenebilir alanlar için düzenlemeye izin ver
        self.table_view.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.table_view.setVerticalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.table_view.setHorizontalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        header = self.table_view.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(False)
            header.setSectionResizeMode(
                QHeaderView.ResizeMode.Interactive
            )
        self.table_view.verticalHeader().setVisible(False)
        self._date_delegate = OptionalDateDelegate(self.table_view)
        for column in self.table_model.DATE_EDITABLE_COLUMNS:
            self.table_view.setItemDelegateForColumn(column, self._date_delegate)
        status_items = get_dava_durumu_list()
        self._status_delegate = StatusDelegate(self.table_view, items=status_items)
        self.table_view.setItemDelegateForColumn(
            self.table_model.COL_DAVA_DURUMU, self._status_delegate
        )
        self.table_view.setItemDelegateForColumn(
            self.table_model.COL_DAVA_DURUMU_2, self._status_delegate
        )
        self.table_view.doubleClicked.connect(self._handle_table_double_click)
        _install_header_menu(
            self.table_view,
            lambda col: self.table_model.headers[col]
            if 0 <= col < len(self.table_model.headers)
            else "",
        )
        _install_copy_shortcut(self.table_view)
        self._apply_initial_column_widths()
        layout.addWidget(self.table_view)

        button_layout = QHBoxLayout()
        self.new_button = QPushButton("Yeni")
        self.new_button.clicked.connect(self.new_requested.emit)
        button_layout.addWidget(self.new_button)

        self.edit_button = QPushButton("Düzenle")
        self.edit_button.clicked.connect(self.edit_requested.emit)
        button_layout.addWidget(self.edit_button)

        self.archive_button = QPushButton("Arşiv")
        self.archive_button.clicked.connect(self.archive_requested.emit)
        button_layout.addWidget(self.archive_button)

        self.vekalet_button = QPushButton("Vekaletnameler")
        self.vekalet_button.clicked.connect(self.vekalet_requested.emit)
        button_layout.addWidget(self.vekalet_button)

        self.settings_button = QPushButton("Ayarlar")
        self.settings_button.clicked.connect(self.settings_requested.emit)
        button_layout.addWidget(self.settings_button)

        self.export_excel_button = QPushButton("Excel'e Aktar")
        self.export_excel_button.clicked.connect(self._on_export_excel_clicked)
        button_layout.addWidget(self.export_excel_button)

        self.export_pdf_button = QPushButton("PDF'e Aktar")
        self.export_pdf_button.clicked.connect(self._on_export_pdf_clicked)
        button_layout.addWidget(self.export_pdf_button)

        button_layout.addStretch()
        layout.addLayout(button_layout)

    def _apply_initial_column_widths(self) -> None:
        if self.table_view is None:
            return
        header = self.table_view.horizontalHeader()
        if header is None:
            return
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        def _enforce_selection_column() -> None:
            if COL_SELECTION < header.count():
                header.setSectionResizeMode(COL_SELECTION, QHeaderView.ResizeMode.Fixed)
                width = DEFAULT_DOSYALAR_COLUMN_WIDTHS.get(COL_SELECTION)
                if width:
                    header.resizeSection(COL_SELECTION, width)

        # Kaydedilmiş sütun genişliklerini yükle
        restored = False
        try:
            settings = QSettings("MyCompany", "TakibiEsasi")
            stored = settings.value("dosyalar/col_widths")
            widths: list[int] = []
            if isinstance(stored, (list, tuple)):
                for value in stored:
                    try:
                        widths.append(int(value))
                    except (TypeError, ValueError):
                        continue
            elif isinstance(stored, str):
                for part in stored.split(","):
                    part = part.strip()
                    if part:
                        try:
                            widths.append(int(part))
                        except ValueError:
                            continue

            if widths:
                for col, width in enumerate(widths):
                    if 0 <= col < header.count() and width > 0:
                        header.resizeSection(col, width)
                restored = True
        except Exception:
            pass

        if not restored:
            # Varsayılan genişlikler
            for column, width in DEFAULT_DOSYALAR_COLUMN_WIDTHS.items():
                if column >= header.count():
                    continue
                if column == COL_SELECTION:
                    header.setSectionResizeMode(column, QHeaderView.ResizeMode.Fixed)
                header.resizeSection(column, width)

        _enforce_selection_column()
        setattr(self.table_view, "_lex_skip_main_defaults", True)

        # Sütun genişliği değiştiğinde otomatik kaydet
        self._column_save_timer = QTimer(self)
        self._column_save_timer.setSingleShot(True)
        self._column_save_timer.setInterval(500)  # 500ms bekle
        self._column_save_timer.timeout.connect(self._delayed_save_column_widths)
        header.sectionResized.connect(self._on_section_resized)

    def _on_section_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Sütun genişliği değiştiğinde timer'ı yeniden başlat."""
        if hasattr(self, "_column_save_timer"):
            self._column_save_timer.start()

    def _delayed_save_column_widths(self) -> None:
        """Sütun genişliklerini QSettings'e kaydet."""
        header = self.table_view.horizontalHeader() if self.table_view else None
        if header is None:
            return
        try:
            widths = [header.sectionSize(col) for col in range(header.count())]
            settings = QSettings("MyCompany", "TakibiEsasi")
            settings.setValue("dosyalar/col_widths", widths)
            settings.sync()
        except Exception:
            pass

    def _handle_table_double_click(self, index: QModelIndex) -> None:
        if not index.isValid():
            return
        if self.proxy is None:
            return
        view = getattr(self, "table_view", None)
        source_index = self.proxy.mapToSource(index)
        if not source_index.isValid():
            return
        column = source_index.column()
        date_columns = (
            self.table_model.COL_IS_TARIHI,
            self.table_model.COL_IS_TARIHI_2,
            self.table_model.COL_DURUSMA_TARIHI,
        )
        if column in date_columns:
            self._show_quick_date_popup(index)
            return
        if column in (
            self.table_model.COL_DAVA_DURUMU,
            self.table_model.COL_DAVA_DURUMU_2,
        ):
            if isinstance(view, QTableView):
                view.edit(index)
            return
        self.row_double_clicked.emit(index)

    def _on_export_excel_clicked(self) -> None:
        if self.table_view is None:
            return
        model = self.table_view.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak kayıt bulunamadı.")
            return

        # Seçim dialogu göster
        dialog = TableExportSelectionDialog(
            self.table_view, self,
            title="Dışa Aktarılacak Dosyaları Seç",
            display_columns=[0, 1, 2],  # Büro No, Esas No, Müvekkil
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        selected_rows = dialog.get_selected_row_indices()
        if not selected_rows:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel olarak kaydet",
            "dosyalar.xlsx",
            "Excel Dosyaları (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        export_table_to_excel_with_selection(self.table_view, path, selected_rows)

    def _on_export_pdf_clicked(self) -> None:
        if self.table_view is None:
            return
        model = self.table_view.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak kayıt bulunamadı.")
            return

        # Seçim dialogu göster
        dialog = TableExportSelectionDialog(
            self.table_view, self,
            title="Dışa Aktarılacak Dosyaları Seç",
            display_columns=[0, 1, 2],
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        selected_rows = dialog.get_selected_row_indices()
        if not selected_rows:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "PDF olarak kaydet",
            "dosyalar.pdf",
            "PDF Dosyaları (*.pdf)",
        )
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        export_table_to_pdf_with_selection(
            self.table_view,
            path,
            title="TakibiEsasi – Dosyalar Listesi",
            subtitle=None,
            selected_rows=selected_rows,
        )

    def _close_quick_date_popup(self) -> None:
        if self._quick_date_popup is None:
            return
        popup = self._quick_date_popup
        self._quick_date_popup = None
        popup.hide()
        popup.deleteLater()

    def _show_quick_date_popup(self, view_index: QModelIndex) -> None:
        view = getattr(self, "table_view", None)
        if view is None:
            return
        model = view.model()
        if model is None:
            return
        self._close_quick_date_popup()
        raw_value = model.data(view_index, Qt.ItemDataRole.EditRole)
        initial_date = _coerce_to_qdate(raw_value)
        popup = QuickDatePopup(view, initial_date=initial_date)
        self._quick_date_popup = popup

        def apply_date(date: QDate) -> None:
            model.setData(view_index, date, Qt.ItemDataRole.EditRole)
            self._close_quick_date_popup()

        def clear_date() -> None:
            model.setData(view_index, None, Qt.ItemDataRole.EditRole)
            self._close_quick_date_popup()

        popup.dateSelected.connect(apply_date)
        popup.cleared.connect(clear_date)
        popup.cancelled.connect(self._close_quick_date_popup)
        rect = view.visualRect(view_index)
        popup_size = popup.sizeHint()
        top_left = view.viewport().mapToGlobal(rect.topLeft())
        popup.move(top_left.x(), top_left.y() - popup_size.height())
        popup.show()
        popup.raise_()
        popup.activateWindow()
        popup.date_edit.setFocus()

    def _on_search_text_changed(self, _text: str) -> None:
        if self._search_timer.isActive():
            self._search_timer.stop()
        self._search_timer.start()

    def _emit_filters_changed(self) -> None:
        if self.mode == "main":
            self.filters_changed.emit()

    # NOTE: added clear_all_filters helper to keep toolbar actions self-contained
    def clear_all_filters(self) -> None:
        """Reset every search/filter widget and refresh the proxy/table."""

        if self.search_input is not None and self.search_input.text():
            self.search_input.clear()
        if self.status_filter_combo is not None:
            self.status_filter_combo.blockSignals(True)
            if self.status_filter_combo.count() > 0:
                self.status_filter_combo.setCurrentIndex(0)
            self.status_filter_combo.blockSignals(False)
        if self.durusma_filter_combo is not None:
            self.durusma_filter_combo.blockSignals(True)
            self.durusma_filter_combo.setCurrentIndex(0)
            self.durusma_filter_combo.blockSignals(False)
        if self.open_only_checkbox is not None:
            self.open_only_checkbox.blockSignals(True)
            self.open_only_checkbox.setChecked(False)
            self.open_only_checkbox.blockSignals(False)
        if self.assignment_checkbox is not None:
            self.assignment_checkbox.blockSignals(True)
            self.assignment_checkbox.setChecked(False)
            self.assignment_checkbox.blockSignals(False)
        if self.user_filter_combo is not None:
            self.user_filter_combo.blockSignals(True)
            if self.user_filter_combo.count() > 0:
                self.user_filter_combo.setCurrentIndex(0)
            self.user_filter_combo.blockSignals(False)
        self.update_assignment_checkbox_state()
        if self.proxy is not None:
            try:
                self.proxy.invalidateFilter()
            except Exception:
                pass
        table_view = getattr(self, "table_view", None)
        if isinstance(table_view, QTableView):
            table_view.scrollToTop()
        self._emit_filters_changed()

    def reset_filters(self) -> None:
        self.clear_all_filters()

    def _on_refresh_clicked(self) -> None:
        """Yenile butonuna tıklandığında tabloyu yenile."""
        self.refresh_requested.emit()

    def set_records(self, records: list[dict[str, Any]]) -> None:
        built = self.build_model_rows(records)
        self.apply_model_rows(built)
        self.apply_proxy_sort_filter()

    def build_model_rows(self, records: list[dict[str, Any]]) -> list[PreparedDosyaRow]:
        return self.table_model.prepare_records(records)

    def apply_model_rows(self, built_rows: list[PreparedDosyaRow]) -> None:
        self.table_model.apply_prepared_records(built_rows)

    def apply_proxy_sort_filter(self) -> None:
        if self.proxy is None:
            return
        if self.open_only_checkbox is not None:
            self.proxy.set_open_only(self.open_only_checkbox.isChecked())
        else:
            self.proxy.set_open_only(False)
        self.proxy.invalidate()

    def collect_filters(self) -> dict[str, object | None]:
        hex6: str | None = None
        if self.status_filter_combo is not None:
            data = self.status_filter_combo.currentData()
            hex6 = data if data else None

        search_text: str | None = None
        if self.search_input is not None:
            text = self.search_input.text().strip()
            if text:
                search_text = text

        open_only = False
        if self.open_only_checkbox is not None:
            open_only = self.open_only_checkbox.isChecked()

        other_filters: dict[str, object] | None = None
        if self.durusma_filter_combo is not None:
            period = self.durusma_filter_combo.currentData()
            if period:
                other_filters = {"durusma_period": period}

        assigned_user_id: int | None = None
        if self.only_own_records and self.current_user_id is not None:
            try:
                assigned_user_id = int(self.current_user_id)
            except (TypeError, ValueError):
                assigned_user_id = None
        elif self.assignment_checkbox is not None and self.assignment_checkbox.isChecked():
            if self.current_user_id is not None:
                try:
                    assigned_user_id = int(self.current_user_id)
                except (TypeError, ValueError):
                    assigned_user_id = None
        elif self.user_filter_combo is not None:
            data = self.user_filter_combo.currentData()
            if data not in (None, ""):
                try:
                    assigned_user_id = int(data)
                except (TypeError, ValueError):
                    assigned_user_id = None

        return {
            "hex6": hex6,
            "search_text": search_text,
            "open_only": open_only,
            "other_filters": other_filters,
            "assigned_user_id": assigned_user_id,
        }

    def populate_status_filter(self) -> None:
        if self.status_filter_combo is None:
            return
        current = self.status_filter_combo.currentData()
        self.status_filter_combo.blockSignals(True)
        self.status_filter_combo.clear()
        self.status_filter_combo.addItem("Hepsi", None)
        for label, code in COLOR_MAP.items():
            self.status_filter_combo.addItem(label, normalize_hex(code))
        if current is not None:
            index = self.status_filter_combo.findData(current)
            if index >= 0:
                self.status_filter_combo.setCurrentIndex(index)
        self.status_filter_combo.blockSignals(False)
        self._emit_filters_changed()

    def populate_user_filter(self) -> None:
        if self.user_filter_combo is None:
            return
        current = self.user_filter_combo.currentData()
        self.user_filter_combo.blockSignals(True)
        self.user_filter_combo.clear()
        self.user_filter_combo.addItem("Tümü", None)
        users = [user for user in get_users() if user.get("active")]
        users.sort(key=lambda item: (item.get("username") or "").lower())
        for user in users:
            self.user_filter_combo.addItem(user.get("username", ""), user.get("id"))
        if current is not None:
            index = self.user_filter_combo.findData(current)
            if index >= 0:
                self.user_filter_combo.setCurrentIndex(index)
        self.user_filter_combo.blockSignals(False)
        self._emit_filters_changed()

    def update_assignment_checkbox_state(self) -> None:
        if self.assignment_checkbox is not None:
            if self.only_own_records:
                self.assignment_checkbox.setChecked(True)
                self.assignment_checkbox.setEnabled(False)
            else:
                self.assignment_checkbox.setEnabled(True)
        if self.user_filter_combo is not None:
            allow_user_selection = self.can_manage_assignments and not self.only_own_records
            self.user_filter_combo.setVisible(allow_user_selection)
            self.user_filter_combo.setEnabled(allow_user_selection)
            if not allow_user_selection:
                self.user_filter_combo.setCurrentIndex(0)



class FinanceTableModel(QAbstractTableModel):
    headers = [
        "BN",
        "Dosya Esas No",
        "Müvekkil Adı",
        "Sabit Ücret",
        "Toplam Ücret",
        "Tahsil Edilen",
        "% Oran",
        "Masraf Toplam",
        "Masraf Tahsil",
        "Kalan Bakiye",
    ]

    currency_columns = {3, 4, 5, 7, 8, 9}
    COL_KALAN_BAKIYE = 9

    def __init__(self, parent=None):
        super().__init__(parent)
        self.records: list[dict] = []

    def set_records(self, records: list[dict]) -> None:
        self.beginResetModel()
        prepared: list[dict] = []
        for record in records:
            item = dict(record)
            display: dict[int, str] = {}
            user_roles: dict[int, object] = {}
            bn_value = item.get("buro_takip_no")
            # NOTE: numeric BN sorting for finance tables via cached UserRole
            display[0] = str(bn_value or "")
            user_roles[0] = _int_sort_key(bn_value)
            display[1] = str(item.get("dosya_esas_no") or "")
            display[2] = str(item.get("muvekkil_adi") or "")
            display[3] = format_tl(item.get("sozlesme_ucreti_cents") or 0)
            display[4] = format_tl(item.get("toplam_ucret_cents", 0))
            display[5] = format_tl(item.get("tahsil_edilen_cents", 0))
            percent_raw = item.get("sozlesme_yuzdesi")
            percent_value: float | None
            try:
                percent_value = float(percent_raw) if percent_raw else None
            except (TypeError, ValueError):
                percent_value = None
            if percent_value is not None:
                base = item.get("tahsil_hedef_cents") or 0
                base_text = f" / {format_tl(base)}" if base else ""
                display[6] = f"{percent_value:.2f} %{base_text}"
            else:
                display[6] = ""
            display[7] = format_tl(item.get("masraf_toplam_cents", 0))
            display[8] = format_tl(item.get("masraf_tahsil_cents", 0))
            display[9] = format_tl(item.get("kalan_bakiye_cents", 0))
            item["_display"] = display
            item["_user_roles"] = user_roles
            prepared.append(item)
        self.records = prepared
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self.records)

    def columnCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self.headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if not index.isValid() or index.row() >= len(self.records):
            return None
        record = self.records[index.row()]
        column = index.column()
        if role == Qt.ItemDataRole.DisplayRole:
            display_map = record.get("_display", {})
            return display_map.get(column, "")
        if role == Qt.ItemDataRole.UserRole:
            sort_map = record.get("_user_roles", {})
            return sort_map.get(column)
        if role == Qt.ItemDataRole.BackgroundRole:
            if (
                column == self.COL_KALAN_BAKIYE
                and record.get("has_overdue_installment")
            ):
                return QColor("#FF6666")
        if role == Qt.ItemDataRole.ForegroundRole:
            if (
                column == self.COL_KALAN_BAKIYE
                and record.get("has_overdue_installment")
            ):
                return QBrush(Qt.GlobalColor.black)
            return None
        if role == Qt.ItemDataRole.TextAlignmentRole:
            if column in self.currency_columns:
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            if column == 6:
                return Qt.AlignmentFlag.AlignCenter
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.headers[section]
        return super().headerData(section, orientation, role)

    def record_at(self, row: int) -> dict | None:
        if 0 <= row < len(self.records):
            return self.records[row]
        return None


class FinanceProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._search_text = ""
        self.setDynamicSortFilter(True)
        self.setSortRole(Qt.ItemDataRole.UserRole)
        self._due_categories: set[str] | None = None
        self._user_filter: int | None = None
        self._payment_range: AlertTokenRange | None = None

    def set_search_text(self, text: str) -> None:
        cleaned, tokens = parse_alert_tokens(text or "")
        self._search_text = cleaned.strip().lower()
        self._payment_range = tokens.get("payment")
        self.invalidateFilter()

    def set_due_category_filter(self, category: str | None) -> None:
        mapping: dict[str, set[str]] = {
            "overdue": {"overdue"},
            "this_week": {"due_today", "due_1_3", "due_4_7"},
            "this_month": {"due_future"},
        }
        new_value: set[str] | None
        if not category:
            new_value = None
        else:
            new_value = mapping.get(category, {category})
        if new_value == self._due_categories:
            return
        self._due_categories = new_value
        self.invalidateFilter()

    def set_user_filter(self, user_id: int | None) -> None:
        if user_id == self._user_filter:
            return
        self._user_filter = user_id
        self.invalidateFilter()

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:  # type: ignore[override]
        model = self.sourceModel()
        if model is None:
            return super().lessThan(left, right)
        left_value = model.data(left, Qt.ItemDataRole.UserRole)
        right_value = model.data(right, Qt.ItemDataRole.UserRole)
        if isinstance(left_value, (int, float)) and isinstance(right_value, (int, float)):
            return left_value < right_value
        left_display = model.data(left, Qt.ItemDataRole.DisplayRole) or ""
        right_display = model.data(right, Qt.ItemDataRole.DisplayRole) or ""
        return str(left_display) < str(right_display)

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:  # type: ignore[override]
        model = self.sourceModel()
        if model is None:
            return True
        record = None
        if hasattr(model, "record_at"):
            record = model.record_at(source_row)
        if self._due_categories is not None:
            due_category = record.get("due_category") if record else None
            if due_category not in self._due_categories:
                return False
        if self._user_filter is not None:
            assigned_ids = []
            if record is not None:
                assigned_ids = record.get("assigned_user_ids") or []
            if self._user_filter not in assigned_ids:
                return False
        if self._payment_range is not None:
            if record is None:
                return False
            start, end = self._payment_range.start, self._payment_range.end
            match_found = False
            for key in ("next_due_date", "due_date", "vade_tarihi"):
                due_date = coerce_to_date(record.get(key))
                if due_date is not None and start <= due_date <= end:
                    match_found = True
                    break
            if not match_found:
                return False
        if not self._search_text:
            return True
        for column in (0, 1, 2):
            index = model.index(source_row, column, source_parent)
            value = model.data(index, Qt.ItemDataRole.DisplayRole)
            if value and self._search_text in str(value).lower():
                return True
        return False

    def record_at(self, row: int) -> dict | None:
        if row < 0 or row >= self.rowCount():
            return None
        source_model = self.sourceModel()
        if source_model is None or not hasattr(source_model, "record_at"):
            return None
        proxy_index = self.index(row, 0)
        source_index = self.mapToSource(proxy_index)
        if not source_index.isValid():
            return None
        return source_model.record_at(source_index.row())  # type: ignore[attr-defined]


class HariciFinanceTableModel(QAbstractTableModel):
    headers = [
        "BN",
        "Dosya Esas No",
        "Müvekkil Adı",
        "Sabit Ücret",
        "Toplam Ücret",
        "Tahsil Edilen",
        "% Oran",
        "Masraf Toplam",
        "Masraf Tahsil",
        "Kalan Bakiye",
    ]

    currency_columns = {3, 4, 5, 7, 8, 9}
    COL_KALAN_BAKIYE = 9

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.records: list[dict] = []

    def set_records(self, records: list[dict]) -> None:
        self.beginResetModel()
        prepared: list[dict] = []
        for record in records:
            item = dict(record)
            item["has_overdue_installment"] = bool(item.get("has_overdue_installment"))
            display: dict[int, str] = {}
            user_roles: dict[int, object] = {}
            bn_value = item.get("harici_bn")
            display[0] = str(bn_value or "")
            user_roles[0] = _int_sort_key(bn_value)
            display[1] = str(item.get("harici_esas_no") or "")
            display[2] = str(item.get("harici_muvekkil") or "")
            display[3] = format_tl(item.get("sabit_ucret_cents") or 0)
            display[4] = format_tl(item.get("toplam_ucret_cents") or 0)
            display[5] = format_tl(item.get("tahsil_edilen_cents") or 0)
            percent_raw = item.get("yuzde_orani")
            percent_value: float | None
            try:
                percent_value = float(percent_raw) if percent_raw else None
            except (TypeError, ValueError):
                percent_value = None
            if percent_value is not None:
                base = item.get("tahsil_hedef_cents") or 0
                base_text = f" / {format_tl(base)}" if base else ""
                display[6] = f"{percent_value:.2f} %{base_text}"
            else:
                display[6] = ""
            display[7] = format_tl(item.get("masraf_toplam_cents") or 0)
            display[8] = format_tl(item.get("masraf_tahsil_cents") or 0)
            display[9] = format_tl(item.get("kalan_bakiye_cents") or 0)
            item["_display"] = display
            item["_user_roles"] = user_roles
            prepared.append(item)
        self.records = prepared
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self.records)

    def columnCount(self, parent=QModelIndex()):  # type: ignore[override]
        return len(self.headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if not index.isValid() or index.row() >= len(self.records):
            return None
        record = self.records[index.row()]
        column = index.column()
        if role == Qt.ItemDataRole.DisplayRole:
            return record.get("_display", {}).get(column, "")
        if role == Qt.ItemDataRole.UserRole:
            return record.get("_user_roles", {}).get(column)
        if role == Qt.ItemDataRole.BackgroundRole:
            if column == self.COL_KALAN_BAKIYE and record.get("has_overdue_installment"):
                return QColor("#FF6666")
        if role == Qt.ItemDataRole.ForegroundRole:
            if column == self.COL_KALAN_BAKIYE and record.get("has_overdue_installment"):
                return QBrush(Qt.GlobalColor.black)
            return None
        if role == Qt.ItemDataRole.TextAlignmentRole:
            if column in self.currency_columns:
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            if column == 6:
                return Qt.AlignmentFlag.AlignCenter
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):  # type: ignore[override]
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.headers[section]
        return super().headerData(section, orientation, role)

    def record_at(self, row: int) -> dict | None:
        if 0 <= row < len(self.records):
            return self.records[row]
        return None

    def row_id(self, row: int) -> int | None:
        record = self.record_at(row)
        if not record:
            return None
        try:
            return int(record.get("id"))
        except (TypeError, ValueError):
            return None

    def find_row_by_id(self, record_id: int) -> int | None:
        for idx, record in enumerate(self.records):
            try:
                if int(record.get("id")) == record_id:
                    return idx
            except (TypeError, ValueError):
                continue
        return None


class FinanceRowDelegate(QStyledItemDelegate):
    COLORS = {
        "overdue": QColor("#4a1f1f"),
        "due_today": QColor("#5c4714"),
        "due_1_3": QColor("#3a3f1a"),
        "due_4_7": QColor("#2a3d44"),
        "due_future": QColor("#1f2c32"),
    }

    def __init__(self, parent=None):
        super().__init__(parent)

    def paint(self, painter, option, index):  # type: ignore[override]
        opt = QStyleOptionViewItem(option)
        color = self._background_color(index)
        cell_background = index.data(Qt.ItemDataRole.BackgroundRole)
        background_color: QColor | None = None
        if cell_background is not None and not (opt.state & QStyle.StateFlag.State_Selected):
            brush = self._as_brush(cell_background)
            opt.backgroundBrush = brush
            background_color = brush.color()
        elif color is not None and not (opt.state & QStyle.StateFlag.State_Selected):
            opt.backgroundBrush = QBrush(color)
            background_color = color
        custom_foreground = index.data(Qt.ItemDataRole.ForegroundRole)
        text_color = None
        if isinstance(custom_foreground, QBrush):
            text_color = custom_foreground.color()
        elif isinstance(custom_foreground, QColor):
            text_color = custom_foreground
        elif isinstance(custom_foreground, Qt.GlobalColor):
            text_color = QColor(custom_foreground)
        if text_color is not None:
            for group in (
                QPalette.ColorGroup.Active,
                QPalette.ColorGroup.Inactive,
                QPalette.ColorGroup.Disabled,
            ):
                opt.palette.setColor(group, QPalette.ColorRole.Text, text_color)
        super().paint(painter, opt, index)

    def _background_color(self, index):
        record = self._record_for_index(index)
        if not record:
            return None
        category = record.get("due_category")
        return self.COLORS.get(category)

    def _as_brush(self, value):
        if isinstance(value, QBrush):
            return value
        if isinstance(value, QColor):
            return QBrush(value)
        return QBrush(QColor(value))

    def _record_for_index(self, index):
        model = index.model()
        if isinstance(model, QSortFilterProxyModel):
            source_index = model.mapToSource(index)
            source_model = model.sourceModel()
            if hasattr(source_model, "record_at"):
                return source_model.record_at(source_index.row())
            return None
        if hasattr(model, "record_at"):
            return model.record_at(index.row())
        return None


class PartialPaymentDialog(QDialog):
    """Kısmi ödeme girişi için basit diyalog."""

    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Kısmi ödeme")
        layout = QVBoxLayout(self)

        form = QFormLayout()
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        form.addRow("Tarih", self.date_edit)

        self.amount_edit = QDoubleSpinBox()
        self.amount_edit.setDecimals(2)
        self.amount_edit.setMaximum(1_000_000_000)
        self.amount_edit.setSuffix(" ₺")
        form.addRow("Tutar", self.amount_edit)

        self.method_edit = QLineEdit()
        self.method_edit.setPlaceholderText("Örn. Havale, Nakit")
        form.addRow("Yöntem", self.method_edit)

        self.note_edit = QLineEdit()
        self.note_edit.setPlaceholderText("Açıklama")
        form.addRow("Açıklama", self.note_edit)

        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def accept(self) -> None:  # type: ignore[override]
        if self.amount_edit.value() <= 0:
            QMessageBox.warning(self, "Geçersiz tutar", "Tutar sıfırdan büyük olmalıdır.")
            return
        super().accept()

    def result_data(self) -> dict:
        return {
            "tarih": self.date_edit.date().toString("yyyy-MM-dd"),
            "tutar_cents": tl_to_cents(self.amount_edit.value()),
            "yontem": self.method_edit.text().strip() or None,
            "aciklama": self.note_edit.text().strip() or None,
        }


class GorevDialog(QDialog):
    """Görev ekleme/düzenleme diyaloğu."""

    def __init__(
        self,
        *,
        parent: QWidget | None = None,
        mode: str = "manual",
        task: dict | None = None,
    ) -> None:
        super().__init__(parent)
        self.mode = mode
        self.task = task or {}
        self._is_readonly_task = False
        self.setWindowTitle("Görev")

        layout = QVBoxLayout(self)
        form = QFormLayout()

        # Tarih alanı - takvim ve erteleme butonları ile
        date_container = QWidget()
        date_layout = QVBoxLayout(date_container)
        date_layout.setContentsMargins(0, 0, 0, 0)
        date_layout.setSpacing(4)

        # Tarihsiz görev seçeneği
        self.no_date_checkbox = QCheckBox("Tarih belirleme (sadece yapılacaklar listesinde görünsün)")
        self.no_date_checkbox.stateChanged.connect(self._on_no_date_changed)
        date_layout.addWidget(self.no_date_checkbox)

        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd.MM.yyyy")
        date_layout.addWidget(self.date_edit)

        # Erteleme butonları
        self.postpone_container = QWidget()
        postpone_layout = QHBoxLayout(self.postpone_container)
        postpone_layout.setContentsMargins(0, 0, 0, 0)
        postpone_layout.setSpacing(6)
        btn_tomorrow = QPushButton("Yarın")
        btn_tomorrow.clicked.connect(lambda: self._set_date_offset(1))
        btn_3days = QPushButton("Üç gün sonra")
        btn_3days.clicked.connect(lambda: self._set_date_offset(3))
        btn_week = QPushButton("Haftaya")
        btn_week.clicked.connect(lambda: self._set_date_offset(7))
        postpone_layout.addWidget(btn_tomorrow)
        postpone_layout.addWidget(btn_3days)
        postpone_layout.addWidget(btn_week)
        postpone_layout.addStretch()
        date_layout.addWidget(self.postpone_container)

        form.addRow("Tarih", date_container)

        # Konu alanı - düzenlenebilir combo box
        self.type_combo = QComboBox()
        self.type_combo.setEditable(True)
        self.type_combo.addItems(["Duruşma", "Değişik İş"])
        self.type_combo.setCurrentText("Değişik İş")
        form.addRow("Konu", self.type_combo)

        self.bn_edit = QLineEdit()
        self.bn_edit.setPlaceholderText("BN opsiyonel")
        form.addRow("BN", self.bn_edit)

        self.muvekkil_edit = QLineEdit()
        self.muvekkil_edit.setPlaceholderText("Müvekkil")
        form.addRow("Müvekkil", self.muvekkil_edit)

        self.aciklama_edit = QPlainTextEdit()
        form.addRow("Açıklama", self.aciklama_edit)

        # Dosya ID'sini sakla - dosya_atamalar senkronizasyonu için
        self._dosya_id: int | None = self.task.get("dosya_id")

        # Atanan kullanıcılar - çoklu seçim listesi
        self.assignees_list = QListWidget()
        self.assignees_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.assignees_list.setMinimumHeight(100)
        self._assignment_user_items: dict[int, QListWidgetItem] = {}
        self._populate_assignees_list()
        form.addRow("Atanan Kullanıcılar", self.assignees_list)

        meta = self._parse_meta(self.task.get("aciklama", ""))
        # Tarihsiz görev kontrolü
        task_date = self.task.get("date") or self.task.get("tarih")
        if task_date is None or task_date == "":
            self.no_date_checkbox.setChecked(True)
            self._on_no_date_changed(Qt.CheckState.Checked.value)
        else:
            self.date_edit.setDate(self._qdate_from_task_date(task_date))
        if meta.get("notes"):
            self.aciklama_edit.setPlainText(meta.get("notes", ""))
        elif self.task.get("aciklama"):
            self.aciklama_edit.setPlainText(self.task.get("aciklama", ""))
        if meta.get("bn"):
            self.bn_edit.setText(meta.get("bn", ""))
        if meta.get("muvekkil"):
            self.muvekkil_edit.setText(meta.get("muvekkil", ""))
        type_value = meta.get("type") or self.task.get("type")
        if type_value:
            idx = self.type_combo.findText(type_value, Qt.MatchFlag.MatchFixedString)
            if idx >= 0:
                self.type_combo.setCurrentIndex(idx)
            else:
                self.type_combo.setCurrentText(type_value)
        # Atanan kullanıcıları yükle
        self._load_assignees()

        if self.mode == "auto":
            self.aciklama_edit.setReadOnly(True)
            self.type_combo.setEnabled(False)
            self.bn_edit.setReadOnly(True)
            self.muvekkil_edit.setReadOnly(True)

        # TEBLIGAT ve ARABULUCULUK görevleri için tamamen salt okunur yap
        task_type = self.task.get("type", "")
        self._is_readonly_task = task_type in ("TEBLIGAT", "ARABULUCULUK")
        if self._is_readonly_task:
            self.setWindowTitle("Görev Detayı (Salt Okunur)")
            self.no_date_checkbox.setEnabled(False)
            self.date_edit.setEnabled(False)
            self.postpone_container.setEnabled(False)
            self.type_combo.setEnabled(False)
            self.bn_edit.setReadOnly(True)
            self.muvekkil_edit.setReadOnly(True)
            self.aciklama_edit.setReadOnly(True)
            self.assignees_list.setEnabled(False)

            # Uyarı etiketi ekle
            readonly_label = QLabel(
                "⚠️ Bu görev {} sekmesinden yönetilmektedir. Değişiklikler yalnızca o sekmeden yapılabilir.".format(
                    "Tebligatlar" if task_type == "TEBLIGAT" else "Arabuluculuk"
                )
            )
            readonly_label.setStyleSheet("color: #856404; background-color: #fff3cd; padding: 8px; border-radius: 4px;")
            readonly_label.setWordWrap(True)
            layout.addWidget(readonly_label)

        layout.addLayout(form)

        buttons = QDialogButtonBox(parent=self)
        if self._is_readonly_task:
            # Salt okunur görevler için sadece Kapat butonu
            close_btn = buttons.addButton("Kapat", QDialogButtonBox.ButtonRole.RejectRole)
            buttons.rejected.connect(self.reject)
        else:
            save_btn = buttons.addButton("Kaydet", QDialogButtonBox.ButtonRole.AcceptRole)
            cancel_btn = buttons.addButton("Çık", QDialogButtonBox.ButtonRole.RejectRole)
            buttons.accepted.connect(self.accept)
            buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        # Kaydedilmiş pencere boyutunu yükle
        self._restore_dialog_size()

    def _restore_dialog_size(self) -> None:
        """Kaydedilmiş pencere boyutunu yükle."""
        settings = QSettings("TakibiEsasi", "TakibiEsasi")
        size = settings.value("GorevDialog/size")
        if size:
            self.resize(size)

    def closeEvent(self, event) -> None:
        """Pencere boyutunu kaydet ve kapat."""
        settings = QSettings("TakibiEsasi", "TakibiEsasi")
        settings.setValue("GorevDialog/size", self.size())
        super().closeEvent(event)

    def _set_date_offset(self, days: int) -> None:
        """Tarihi bugünden belirtilen gün kadar ileri ayarla."""
        self.no_date_checkbox.setChecked(False)
        self.date_edit.setDate(QDate.currentDate().addDays(days))

    def _on_no_date_changed(self, state: int) -> None:
        """Tarihsiz görev seçeneği değiştiğinde çağrılır."""
        is_no_date = state == Qt.CheckState.Checked.value
        self.date_edit.setEnabled(not is_no_date)
        self.postpone_container.setEnabled(not is_no_date)

    def accept(self) -> None:  # type: ignore[override]
        super().accept()

    def _populate_assignees_list(self) -> None:
        """Kullanıcı listesini doldur."""
        self.assignees_list.clear()
        self._assignment_user_items.clear()
        users = get_users()
        users.sort(key=lambda item: item.get("username", "").lower())
        for user in users:
            role_label = USER_ROLE_LABELS.get(user.get("role"), user.get("role", ""))
            username = user.get("username", "")
            display = username
            if role_label:
                display = f"{username} ({role_label})"
            if not user.get("active"):
                display = f"{display} [Pasif]"
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, user.get("id"))
            if not user.get("active"):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEnabled)
            self.assignees_list.addItem(item)
            try:
                user_id = int(user.get("id"))
                self._assignment_user_items[user_id] = item
            except (TypeError, ValueError):
                continue

    def _load_assignees(self) -> None:
        """Mevcut atanan kullanıcıları yükle - dosya varsa dosya_atamalar'dan, yoksa gorevler'den."""
        if self._dosya_id:
            # Dosya bağlantısı varsa dosya_atamalar'dan yükle
            try:
                assignees = get_dosya_assignees(int(self._dosya_id))
                user_ids = [a.get("id") for a in assignees if a.get("id")]
                self._select_assignees(user_ids)
            except Exception:
                pass
        else:
            # Dosya bağlantısı yoksa gorevler.atanan_kullanicilar'dan yükle
            atanan = self.task.get("atanan_kullanicilar", "")
            if atanan:
                # Virgülle ayrılmış kullanıcı adlarını bul
                usernames = [u.strip() for u in atanan.split(",") if u.strip()]
                users = get_users()
                user_id_map = {u.get("username"): u.get("id") for u in users}
                user_ids = [user_id_map.get(uname) for uname in usernames if user_id_map.get(uname)]
                self._select_assignees(user_ids)

    def _select_assignees(self, user_ids: list) -> None:
        """Belirtilen kullanıcıları seç."""
        target_ids = {int(uid) for uid in user_ids if uid is not None}
        self.assignees_list.blockSignals(True)
        for user_id, item in self._assignment_user_items.items():
            item.setSelected(user_id in target_ids)
        self.assignees_list.blockSignals(False)

    def _get_selected_assignee_ids(self) -> list[int]:
        """Seçili kullanıcı ID'lerini döndür."""
        ids: list[int] = []
        for item in self.assignees_list.selectedItems():
            user_id = item.data(Qt.ItemDataRole.UserRole)
            try:
                ids.append(int(user_id))
            except (TypeError, ValueError):
                continue
        return ids

    def _get_selected_assignee_usernames(self) -> str:
        """Seçili kullanıcı adlarını virgülle ayrılmış string olarak döndür."""
        usernames: list[str] = []
        users = get_users()
        user_map = {u.get("id"): u.get("username") for u in users}
        for user_id in self._get_selected_assignee_ids():
            username = user_map.get(user_id)
            if username:
                usernames.append(username)
        return ", ".join(usernames)

    def save_assignees_if_linked_to_dosya(self) -> None:
        """Eğer dosya bağlantısı varsa atanan kullanıcıları dosya_atamalar'a kaydet."""
        if self._dosya_id:
            try:
                set_dosya_assignees(int(self._dosya_id), self._get_selected_assignee_ids())
            except Exception:
                pass

    def result_data(self) -> dict:
        meta_payload = {
            "bn": self.bn_edit.text().strip(),
            "muvekkil": self.muvekkil_edit.text().strip(),
            "type": self.type_combo.currentText(),
            "notes": self.aciklama_edit.toPlainText().strip(),
        }
        aciklama_value = json.dumps(meta_payload, ensure_ascii=False)
        # Tarihsiz görev kontrolü
        if self.no_date_checkbox.isChecked():
            tarih_value = None
        else:
            qdate = self.date_edit.date()
            tarih_value = date(qdate.year(), qdate.month(), qdate.day())
        # Seçili kullanıcıları string olarak al
        atanan_str = self._get_selected_assignee_usernames()
        return {
            "tarih": tarih_value,
            "konu": self.type_combo.currentText().strip(),
            "aciklama": f"__META__{aciklama_value}",
            "atanan_kullanicilar": atanan_str,
        }

    @staticmethod
    def _parse_meta(raw: str) -> dict:
        if not raw or not raw.startswith("__META__"):
            return {}
        try:
            return json.loads(raw.replace("__META__", "", 1))
        except json.JSONDecodeError:
            return {}

    @staticmethod
    def _qdate_from_task_date(value: Any) -> QDate:
        if isinstance(value, date):
            return QDate(value.year, value.month, value.day)
        if isinstance(value, str) and value:
            parsed = QDate.fromString(value, "yyyy-MM-dd")
            if parsed.isValid():
                return parsed
        return QDate.currentDate()


class WeekNumberHeader(QHeaderView):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(Qt.Orientation.Vertical, parent)
        self.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setSectionsClickable(False)
        self.setHighlightSections(False)

    def paintSection(self, painter: QPainter, rect: QRect, logicalIndex: int) -> None:  # type: ignore[override]
        if not rect.isValid() or self.model() is None:
            return
        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        option = QStyleOptionHeader()
        self.initStyleOption(option)
        option.rect = rect
        option.text = f"Hafta {self.model().headerData(logicalIndex, Qt.Orientation.Vertical, Qt.ItemDataRole.DisplayRole)}"
        option.palette.setColor(QPalette.ColorRole.ButtonText, QColor("#888888"))
        option.font.setPointSize(option.font.pointSize() - 1)
        self.style().drawControl(QStyle.ControlElement.CE_Header, option, painter, self)
        painter.restore()


class UserAssignmentDelegate(QStyledItemDelegate):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._users = self._load_users()

    @staticmethod
    def _load_users() -> list[str]:
        users = [user for user in get_users() if user.get("active")]
        users.sort(key=lambda item: (item.get("username") or "").lower())
        return [user.get("username", "") for user in users]

    def createEditor(self, parent: QWidget, option: QStyleOptionViewItem, index: QModelIndex):  # type: ignore[override]
        combo = QComboBox(parent)
        combo.addItem("", "")
        for username in self._users:
            combo.addItem(username, username)
        return combo

    def setEditorData(self, editor: QWidget, index: QModelIndex) -> None:  # type: ignore[override]
        if isinstance(editor, QComboBox):
            current_value = index.data(Qt.ItemDataRole.DisplayRole) or ""
            match = editor.findData(current_value)
            if match >= 0:
                editor.setCurrentIndex(match)
            else:
                editor.setCurrentIndex(0)
        else:
            super().setEditorData(editor, index)

    def setModelData(self, editor: QWidget, model: QAbstractTableModel, index: QModelIndex) -> None:  # type: ignore[override]
        if isinstance(editor, QComboBox):
            model.setData(index, editor.currentData() or editor.currentText())
        else:
            super().setModelData(editor, model, index)


class TaskBadgeCalendar(QCalendarWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.tasks_by_date: dict[str, list[dict[str, Any]]] = {}
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.setGridVisible(True)
        self.go_to_today_handler: Callable[[], None] | None = None

    def set_tasks(self, tasks_by_date: dict[str, list[dict[str, Any]]]) -> None:
        self.tasks_by_date = tasks_by_date
        self._update_tooltips()
        self.update()

    def _update_tooltips(self) -> None:
        self.setDateTextFormat(QDate(), QTextCharFormat())
        for date_str, tasks in self.tasks_by_date.items():
            qdate = QDate.fromString(date_str, "yyyy-MM-dd")
            fmt = QTextCharFormat()
            tooltip_lines = []
            for task in tasks:
                bn_text = f"BN {task.get('bn')}" if task.get("bn") else "BN -"
                subject = task.get("konu") or task.get("description") or ""
                type_text = task.get("type_label", "") or task.get("type", "")
                assigned = task.get("atanan_kullanicilar") or "Otomatik"
                tooltip_lines.append(
                    f"{bn_text} – {subject} – {type_text} – atanan: {assigned}"
                )
            fmt.setToolTip("\n".join(tooltip_lines))
            self.setDateTextFormat(qdate, fmt)

    def paintCell(self, painter: QPainter, rect: QRect, date: QDate) -> None:  # type: ignore[override]
        painter.save()
        is_selected = date == self.selectedDate()
        in_month = date.month() == self.monthShown() and date.year() == self.yearShown()
        background = QColor("#ffffff") if in_month else QColor(210, 210, 210)  # Önceki/sonraki ay daha koyu
        if is_selected:
            background = self.palette().highlight().color()
        painter.fillRect(rect, background)

        day_text = str(date.day())
        if is_selected:
            text_color = QColor("#000000") if background.lightness() > 140 else QColor("#ffffff")
        elif not in_month:
            text_color = QColor("#707070")  # Önceki/sonraki ay metni daha koyu
        else:
            text_color = QColor("#000000") if background.lightness() > 140 else QColor("#ffffff")
        painter.setPen(text_color)
        font = painter.font()
        font.setBold(True)  # Tarih rakamları her zaman kalın
        painter.setFont(font)
        painter.drawText(rect.adjusted(4, 2, -4, -rect.height() // 2), Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop, day_text)

        date_str = date.toString("yyyy-MM-dd")
        tasks = self.tasks_by_date.get(date_str, [])
        if tasks:
            self._draw_badges(painter, rect, tasks, date_str)
        if date == QDate.currentDate():
            pen = QPen(QColor("#2c3e50"))
            pen.setWidth(2)
            painter.setPen(pen)
            painter.drawRect(rect.adjusted(1, 1, -1, -1))
        painter.restore()

    def keyPressEvent(self, event):  # type: ignore[override]
        if event is None:
            return
        if event.matches(QKeySequence.StandardKey.Refresh):
            self.update()
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier and event.key() in (
            Qt.Key.Key_Left,
            Qt.Key.Key_Right,
        ):
            delta = -7 if event.key() == Qt.Key.Key_Left else 7
            self.setSelectedDate(self.selectedDate().addDays(delta))
            self.showSelectedDate()
            return
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier and event.key() == Qt.Key.Key_G:
            if self.go_to_today_handler:
                self.go_to_today_handler()
            else:
                today = QDate.currentDate()
                self.setSelectedDate(today)
                self.setCurrentPage(today.year(), today.month())
            return
        if event.key() == Qt.Key.Key_PageUp:
            self.showPreviousMonth()
            return
        if event.key() == Qt.Key.Key_PageDown:
            self.showNextMonth()
            return
        super().keyPressEvent(event)

    def _draw_badges(self, painter: QPainter, rect: QRect, tasks: list[dict[str, Any]], date_str: str) -> None:
        """
        Aynı güne ait birden fazla görevi daha anlaşılır göstermek için
        rozetleri birden fazla satıra yayar ve sığmayanlar için
        '+N görev' şeklinde özet rozet çizer.
        """
        if not tasks:
            return

        fm = painter.fontMetrics()
        padding = 6
        spacing = 4  # rozetler arası yatay boşluk
        row_spacing = 2  # satırlar arası dikey boşluk
        badge_height = fm.height() + 4
        min_badge_width = 60  # Minimum rozet genişliği

        # Üst tarafta gün numarası için ayrılan tahmini alan
        day_label_height = fm.height() + 6
        top_reserved = rect.top() + day_label_height
        bottom_margin = 4

        max_width = rect.width() - 8
        if max_width <= 0:
            return

        # Rozetler için kullanılabilir dikey alan
        available_height = rect.bottom() - bottom_margin - top_reserved
        if available_height < badge_height:
            available_height = badge_height

        max_rows = max(1, available_height // (badge_height + row_spacing))

        # Her satırda kaç rozet sığacağını hesapla
        badges_per_row = max(1, (max_width + spacing) // (min_badge_width + spacing))

        drawn = 0
        index = 0

        # En alttan yukarıya doğru satır satır rozet çiz
        for row_idx in range(max_rows):
            if index >= len(tasks):
                break

            y = rect.bottom() - bottom_margin - badge_height - row_idx * (badge_height + row_spacing)
            if y < top_reserved:
                break

            # Bu satırda kaç rozet çizeceğimizi hesapla
            row_badge_count = min(badges_per_row, len(tasks) - index)
            if row_badge_count <= 0:
                break

            # Rozet genişliğini satırdaki rozet sayısına göre eşit böl
            total_spacing = (row_badge_count - 1) * spacing
            badge_width = (max_width - total_spacing) // row_badge_count
            badge_width = max(min_badge_width, min(badge_width, max_width))

            x = rect.left() + 4
            for _ in range(row_badge_count):
                if index >= len(tasks):
                    break

                task = tasks[index]
                text = (task.get("konu") or "").strip() or self._badge_text(task.get("type"))
                bg_color, fg_color = self._badge_colors(task)

                # Kalan alana göre genişliği ayarla
                remaining_space = rect.right() - 4 - x
                actual_width = min(badge_width, remaining_space)
                if actual_width < 30:
                    break

                elided = fm.elidedText(
                    text,
                    Qt.TextElideMode.ElideRight,
                    max(10, actual_width - padding * 2),
                )
                badge_rect = QRect(int(x), int(y), int(actual_width), badge_height)
                painter.fillRect(badge_rect, bg_color)
                painter.setPen(fg_color)
                painter.drawRect(badge_rect)
                painter.drawText(badge_rect, Qt.AlignmentFlag.AlignCenter, elided)

                x += actual_width + spacing
                drawn += 1
                index += 1

        remaining = len(tasks) - drawn
        if remaining > 0:
            # Sığmayan görevler için '+N görev' şeklinde özet rozet
            summary_text = f"+{remaining} görev"
            summary_width = fm.horizontalAdvance(summary_text) + padding * 2
            summary_width = min(max_width, max(60, summary_width))

            summary_y = top_reserved
            summary_x = rect.right() - 4 - summary_width

            badge_rect = QRect(int(summary_x), int(summary_y), int(summary_width), badge_height)
            painter.fillRect(badge_rect, QColor("#dddddd"))
            painter.setPen(QColor("#000000"))
            painter.drawRect(badge_rect)
            painter.drawText(badge_rect, Qt.AlignmentFlag.AlignCenter, summary_text)

    @staticmethod
    def _badge_text(task_type: str | None) -> str:
        mapping = {
            "DURUSMA": "Duruşma",
            "IS_TARIHI": "İş Tarihi",
            "IS_TARIHI_2": "İş Tarihi 2",
            "MANUAL": "Görev",
            "TEBLIGAT": "Tebligat",
            "ARABULUCULUK": "Arabuluculuk",
        }
        return mapping.get(task_type or "", task_type or "Görev")

    def _badge_colors(self, task: dict[str, Any]) -> tuple[QColor, QColor]:
        """Badge renklerini görev türüne veya dava durumuna göre döndürür.

        Takvim üzerindeki rozetler:
        - DURUSMA: pembe (#c2185b)
        - TEBLIGAT: mor (#9b59b6)
        - ARABULUCULUK: turkuaz (#1abc9c)
        - Diğerleri: dava durumu renginde gösterilir
        Eğer dava durumu rengi yoksa varsayılan renk kullanılır.
        """
        # Önce görev türüne göre özel renk kontrol et
        task_type = task.get("type", "")
        if task_type == "DURUSMA":
            return QColor("#c2185b"), QColor("#ffffff")  # Pembe, beyaz metin
        if task_type == "TEBLIGAT":
            return QColor("#9b59b6"), QColor("#ffffff")  # Mor, beyaz metin
        if task_type == "ARABULUCULUK":
            return QColor("#1abc9c"), QColor("#000000")  # Turkuaz, siyah metin

        try:
            # Önce direkt renk alanlarını kontrol et
            status_color = task.get("dava_durumu_color") or task.get("status_color")

            # Renk yoksa dava_durumu adından rengi al
            if not status_color:
                dava_durumu = task.get("dava_durumu")
                if dava_durumu and isinstance(dava_durumu, str) and dava_durumu.strip():
                    status_color = get_status_color(dava_durumu.strip())

            if status_color:
                normalized = normalize_hex(status_color)
                if normalized:
                    bg_color = QColor(f"#{normalized}")
                    fg_color = QColor(get_status_text_color(normalized))
                    return bg_color, fg_color
        except Exception:
            pass  # Hata durumunda varsayılan renk kullan

        # Varsayılan renk (dava durumu rengi yoksa veya hata varsa)
        return QColor("#f5f5f5"), QColor("#333333")


class GorevlerTab(QWidget):
    """Görevler sekmesi: Yapılacaklar (To-Do) + Takvim sekmeleri."""

    def __init__(self, current_user: dict, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.current_user = current_user or {}
        self.selected_date = QDate.currentDate()
        self.current_filter = "today"
        self._tasks: list[dict[str, Any]] = []
        self._all_tasks: list[dict[str, Any]] = []
        self._todo_tasks: list[dict[str, Any]] = []
        self._auto_assignees: dict[str, str] = {}
        self._use_selected_date_filter = True
        self._block_item_changed = False
        self._block_todo_item_changed = False
        self._task_data_role = Qt.ItemDataRole.UserRole + 1

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Ana sekme yapısı
        self.main_tabs = QTabWidget()
        layout.addWidget(self.main_tabs)

        # 1. Yapılacaklar (To-Do) sekmesi
        self._build_todo_tab()

        # 2. Takvim sekmesi
        self._build_calendar_tab()

        # Kısayollar
        self._reset_filters_shortcut = QShortcut(QKeySequence("Ctrl+Shift+F"), self)
        self._reset_filters_shortcut.activated.connect(self._clear_filters)
        self._today_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self._today_shortcut.activated.connect(self._go_to_today)
        self._refresh_shortcut = QShortcut(QKeySequence.StandardKey.Refresh, self)
        self._refresh_shortcut.activated.connect(self.refresh_tasks)

        # İlk yükleme
        self._activate_filter("today")
        self._set_month_label()
        self.refresh_tasks()

    def _build_todo_tab(self) -> None:
        """Yapılacaklar (To-Do) sekmesini oluştur."""
        todo_widget = QWidget()
        todo_layout = QVBoxLayout(todo_widget)
        todo_layout.setContentsMargins(8, 8, 8, 8)

        # Başlık ve açıklama
        header_label = QLabel("Tüm görevleriniz burada listelenir. Tamamlanan görevler listenin altına iner.")
        header_label.setStyleSheet("color: #666; font-style: italic; margin-bottom: 8px;")
        todo_layout.addWidget(header_label)

        # Kullanıcı filtresi
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Kullanıcı:"))
        self.todo_user_filter = QComboBox()
        self.todo_user_filter.setMinimumWidth(180)
        self.todo_user_filter.addItem("Tümü", "")
        for username in self._load_users_for_filter():
            self.todo_user_filter.addItem(username, username)
        self.todo_user_filter.currentIndexChanged.connect(self._on_todo_user_filter_changed)
        filter_layout.addWidget(self.todo_user_filter)
        filter_layout.addStretch(1)
        todo_layout.addLayout(filter_layout)

        # Görev tablosu - 7 sütun (Checkbox, Tarih, Konu, BN, Müvekkil, Atanan, İçerik)
        self.todo_table = QTableWidget(0, 7)
        self.todo_table.setHorizontalHeaderLabels(
            ["✓", "Tarih", "Konu", "BN", "Müvekkil", "Atanan", "İçerik"]
        )
        self.todo_table.horizontalHeader().setStretchLastSection(True)
        self.todo_table.setColumnWidth(0, 30)  # Checkbox sütunu
        self.todo_table.setSortingEnabled(False)  # Manuel sıralama yapacağız
        self.todo_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.todo_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.todo_table.doubleClicked.connect(self._on_todo_item_double_clicked)
        self.todo_table.setItemDelegateForColumn(1, TodoDateDelegate(self.todo_table))  # Tarih sütunu
        self.todo_table.setItemDelegateForColumn(5, UserAssignmentDelegate(self.todo_table))
        self.todo_table.itemChanged.connect(self._on_todo_item_changed)

        # Sütun genişliklerini kaydet/yükle
        self._load_todo_column_widths()
        self._todo_column_save_timer = QTimer(self)
        self._todo_column_save_timer.setSingleShot(True)
        self._todo_column_save_timer.setInterval(500)
        self._todo_column_save_timer.timeout.connect(self._save_todo_column_widths)
        self.todo_table.horizontalHeader().sectionResized.connect(self._on_todo_column_resized)

        todo_layout.addWidget(self.todo_table)

        # Aksiyon butonları
        todo_actions = QHBoxLayout()
        self.todo_add_button = QPushButton("Görev Ekle")
        self.todo_add_button.clicked.connect(self._add_todo_task)
        todo_actions.addWidget(self.todo_add_button)

        self.todo_edit_button = QPushButton("Düzenle")
        self.todo_edit_button.clicked.connect(self._edit_todo_task)
        todo_actions.addWidget(self.todo_edit_button)

        self.todo_complete_button = QPushButton("Tamamlandı")
        self.todo_complete_button.clicked.connect(self._toggle_todo_complete)
        todo_actions.addWidget(self.todo_complete_button)

        self.todo_delete_button = QPushButton("Sil")
        self.todo_delete_button.clicked.connect(self._delete_todo_task)
        todo_actions.addWidget(self.todo_delete_button)

        self.todo_refresh_button = QPushButton("Yenile")
        self.todo_refresh_button.clicked.connect(self.refresh_tasks)
        todo_actions.addWidget(self.todo_refresh_button)

        todo_actions.addStretch(1)
        todo_layout.addLayout(todo_actions)

        # Legend - Tarih bazlı renklendirme
        todo_legend = QHBoxLayout()
        todo_legend.addWidget(self._legend_label("#ff4d4d", "Geçmiş"))
        todo_legend.addWidget(self._legend_label("#4caf50", "Bugün"))
        todo_legend.addWidget(self._legend_label("#ffeb3b", "1-3 Gün"))
        todo_legend.addWidget(self._legend_label("#ff9800", "4-7 Gün"))
        todo_legend.addWidget(self._legend_label("#2196f3", "8-14 Gün"))
        todo_legend.addWidget(self._legend_label("#e8e8e8", "Tamamlandı"))
        todo_legend.addStretch(1)
        todo_layout.addLayout(todo_legend)

        self.main_tabs.addTab(todo_widget, "Yapılacaklar")

    def _build_calendar_tab(self) -> None:
        """Takvim sekmesini oluştur."""
        calendar_widget = QWidget()
        calendar_layout = QVBoxLayout(calendar_widget)
        calendar_layout.setContentsMargins(8, 8, 8, 8)

        # Splitter: Takvim üstte daha büyük, tablo altta daha küçük
        splitter = QSplitter(Qt.Orientation.Vertical)

        # === ÜST KISIM: Takvim ve kontrolleri ===
        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)

        # Ay navigasyonu
        header_layout = QHBoxLayout()
        self.prev_month_btn = QToolButton()
        self.prev_month_btn.setText("<<")
        self.prev_month_btn.clicked.connect(partial(self._shift_month, -1))
        self.today_btn = QPushButton("Bugün")
        self.today_btn.clicked.connect(self._go_to_today)
        self.next_month_btn = QToolButton()
        self.next_month_btn.setText(">>")
        self.next_month_btn.clicked.connect(partial(self._shift_month, 1))
        header_layout.addWidget(self.prev_month_btn)
        header_layout.addWidget(self.today_btn)
        header_layout.addWidget(self.next_month_btn)
        header_layout.addStretch(1)
        self.month_label = QLabel()
        self.month_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(self.month_label, 1)
        header_layout.addStretch(1)
        top_layout.addLayout(header_layout)

        # Takvim
        self.calendar = TaskBadgeCalendar()
        self.calendar.go_to_today_handler = self._go_to_today
        self.calendar.setMinimumHeight(300)
        calendar_nav = self.calendar.findChild(QWidget, "qt_calendar_navigationbar")
        if calendar_nav is not None:
            calendar_nav.hide()
        view = self.calendar.findChild(QTableView, "qt_calendar_calendarview")
        if view is not None:
            view.verticalHeader().hide()
            # Gün adları satırı yüksekliğini en aza indir (Pzt, Sal, Çar, vb.)
            h_header = view.horizontalHeader()
            h_header.setFixedHeight(8)  # 8px yükseklik
            h_header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
            h_header.setStyleSheet(
                "QHeaderView::section { font-weight: bold; padding: 0px; font-size: 7px; margin: 0px; }"
            )
        self.calendar.selectionChanged.connect(self._on_calendar_changed)
        self.calendar.activated.connect(self._on_calendar_double_clicked)
        top_layout.addWidget(self.calendar, 1)  # Stretch factor 1

        # Legend - Tarih bazlı renklendirme
        legend = QHBoxLayout()
        legend.addWidget(self._legend_label("#c2185b", "Duruşma"))
        legend.addWidget(self._legend_label("#9b59b6", "Tebligat"))
        legend.addWidget(self._legend_label("#1abc9c", "Arabuluculuk"))
        legend.addWidget(self._legend_label("#ff4d4d", "Geçmiş"))
        legend.addWidget(self._legend_label("#4caf50", "Bugün"))
        legend.addWidget(self._legend_label("#ffeb3b", "1-3 Gün"))
        legend.addWidget(self._legend_label("#ff9800", "4-7 Gün"))
        legend.addWidget(self._legend_label("#2196f3", "8-14 Gün"))
        legend.addStretch(1)
        top_layout.addLayout(legend)

        splitter.addWidget(top_widget)

        # === ALT KISIM: Filtreler ve görev tablosu ===
        bottom_widget = QWidget()
        bottom_layout = QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)

        # Filtreler
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(6)
        self.filter_buttons: dict[str, QToolButton] = {}
        for key, label in (
            ("all", "Tümü"),
            ("today", "Bugün"),
            ("week", "Bu Hafta"),
            ("month", "Bu Ay"),
        ):
            btn = QToolButton()
            btn.setText(label)
            btn.setCheckable(True)
            btn.clicked.connect(partial(self._on_filter_changed, key))
            self.filter_buttons[key] = btn
            filter_layout.addWidget(btn)
        filter_layout.addSpacing(8)
        filter_layout.addWidget(QLabel("Kullanıcı:"))
        self.calendar_user_filter = QComboBox()
        self.calendar_user_filter.setMinimumWidth(180)
        self.calendar_user_filter.addItem("Tümü", "")
        for username in self._load_users_for_filter():
            self.calendar_user_filter.addItem(username, username)
        self.calendar_user_filter.currentIndexChanged.connect(self._on_calendar_user_changed)
        filter_layout.addWidget(self.calendar_user_filter)
        self.refresh_button = QToolButton()
        self.refresh_button.setText("Yenile")
        self.refresh_button.setToolTip("Takvimi ve görev listesini yenile")
        self.refresh_button.setMinimumWidth(90)
        self.refresh_button.setMinimumHeight(28)
        self.refresh_button.clicked.connect(self.refresh_tasks)
        filter_layout.addWidget(self.refresh_button)
        filter_layout.addStretch(1)
        bottom_layout.addLayout(filter_layout)

        # Görev tablosu
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(
            ["Tarih", "Konu", "BN", "Müvekkil", "Atanan", "İçerik"]
        )
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSortingEnabled(True)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_item_double_clicked)
        self.table.setItemDelegateForColumn(4, UserAssignmentDelegate(self.table))
        self.table.itemChanged.connect(self._on_item_changed)

        # Sütun genişliklerini kaydet/yükle
        self._load_calendar_task_column_widths()
        self._calendar_task_column_save_timer = QTimer(self)
        self._calendar_task_column_save_timer.setSingleShot(True)
        self._calendar_task_column_save_timer.setInterval(500)
        self._calendar_task_column_save_timer.timeout.connect(self._save_calendar_task_column_widths)
        self.table.horizontalHeader().sectionResized.connect(self._on_calendar_task_column_resized)

        bottom_layout.addWidget(self.table, 1)  # Stretch factor 1

        # Aksiyon butonları
        actions = QHBoxLayout()
        self.add_button = QPushButton("Görev Ekle")
        self.add_button.clicked.connect(self._add_task)
        actions.addWidget(self.add_button)

        self.edit_button = QPushButton("Seçili Görevi Düzenle")
        self.edit_button.clicked.connect(self._edit_task)
        actions.addWidget(self.edit_button)

        self.delete_button = QPushButton("Sil")
        self.delete_button.clicked.connect(self._delete_task)
        actions.addWidget(self.delete_button)

        actions.addStretch(1)
        bottom_layout.addLayout(actions)

        # Kısayol etiketi
        shortcut_label = QLabel(
            "Kısayollar: Ctrl+G → Bugün, Ctrl+Shift+F → Filtreleri Temizle, F5 → Yenile"
        )
        shortcut_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        bottom_layout.addWidget(shortcut_label)

        splitter.addWidget(bottom_widget)

        # Splitter oranları: Takvim %65, Tablo %35
        splitter.setSizes([650, 350])
        splitter.setStretchFactor(0, 2)  # Üst kısım 2x esner
        splitter.setStretchFactor(1, 1)  # Alt kısım 1x esner

        calendar_layout.addWidget(splitter)

        self.main_tabs.addTab(calendar_widget, "Takvim")

    def _legend_label(self, color: str, text: str) -> QLabel:
        label = QLabel(text)
        bg = QColor(color)
        r, g, b, _ = bg.getRgb()
        luma = 0.299 * r + 0.587 * g + 0.114 * b
        fg = "black" if luma > 160 else "white"
        label.setStyleSheet(
            f"QLabel {{ background-color: {color}; color: {fg}; padding: 2px 6px; border-radius: 4px; }}"
        )
        return label

    def _activate_filter(self, key: str) -> None:
        for name, btn in self.filter_buttons.items():
            btn.setChecked(name == key)
        self.current_filter = key

    def _range_for_filter(self) -> tuple[QDate, QDate]:
        today = QDate.currentDate()
        if self.current_filter == "today":
            return today, today
        if self.current_filter == "week":
            start = today.addDays(-(today.dayOfWeek() - 1))
            end = start.addDays(6)
            return start, end
        if self.current_filter == "month":
            start = QDate(self.calendar.yearShown(), self.calendar.monthShown(), 1)
            end = start.addMonths(1).addDays(-1)
            return start, end
        return today.addYears(-1), today.addYears(1)

    def _calendar_range(self) -> tuple[QDate, QDate]:
        today = QDate.currentDate()
        return today.addYears(-2), today.addYears(2)

    def _on_filter_changed(self, key: str) -> None:
        self._activate_filter(key)
        if key == "today":
            self._go_to_today()
            self._use_selected_date_filter = True
        else:
            self._use_selected_date_filter = False
        self.refresh_tasks()

    def _clear_filters(self) -> None:
        self._use_selected_date_filter = False
        self._activate_filter("all")
        self.refresh_tasks()

    def _on_calendar_user_changed(self) -> None:
        self.refresh_tasks()

    def _on_todo_user_filter_changed(self) -> None:
        """Yapılacaklar sekmesinde kullanıcı filtresi değiştiğinde."""
        self._populate_todo_table()

    @staticmethod
    def _load_users_for_filter() -> list[str]:
        users = [user for user in get_users() if user.get("active")]
        users.sort(key=lambda item: (item.get("username") or "").lower())
        return [user.get("username", "") for user in users]

    @staticmethod
    def _safe_date(value: Any) -> tuple[date | None, str]:
        if not value:
            return None, ""
        if isinstance(value, date):
            return value, value.isoformat()
        text = str(value).strip()
        if not text:
            return None, ""
        try:
            parsed = date.fromisoformat(text.split("T")[0])
            return parsed, parsed.isoformat()
        except ValueError:
            try:
                parsed_dt = datetime.strptime(text, "%Y-%m-%d").date()
                return parsed_dt, parsed_dt.isoformat()
            except ValueError:
                return None, text

    def _user_id_lookup(self) -> dict[str, Any]:
        if not hasattr(self, "_user_id_cache"):
            self._user_id_cache = {
                user.get("username", ""): user.get("id") for user in get_users()
            }
        return self._user_id_cache

    def _shift_month(self, delta: int) -> None:
        month = self.calendar.monthShown() + delta
        year = self.calendar.yearShown()
        while month < 1:
            month += 12
            year -= 1
        while month > 12:
            month -= 12
            year += 1
        self.calendar.setCurrentPage(year, month)
        self._set_month_label()
        self.refresh_tasks()

    def _set_month_label(self) -> None:
        month_names = [
            "Ocak",
            "Şubat",
            "Mart",
            "Nisan",
            "Mayıs",
            "Haziran",
            "Temmuz",
            "Ağustos",
            "Eylül",
            "Ekim",
            "Kasım",
            "Aralık",
        ]
        month = self.calendar.monthShown()
        year = self.calendar.yearShown()
        name = month_names[month - 1] if 1 <= month <= 12 else ""
        self.month_label.setText(f"{name} {year}")

    def _go_to_today(self) -> None:
        today = QDate.currentDate()
        self.calendar.setSelectedDate(today)
        self.calendar.setCurrentPage(today.year(), today.month())
        self._set_month_label()
        self._activate_filter("today")
        self._use_selected_date_filter = True
        self.refresh_tasks()

    # -------------------- Takvim Görev Tablosu Sütun Genişlikleri --------------------

    def _load_calendar_task_column_widths(self) -> None:
        """Takvim görev tablosu için kayıtlı sütun genişliklerini yükle."""
        try:
            settings = QSettings("MyCompany", "TakibiEsasi")
            widths = settings.value("calendar_tasks/col_widths", None)
            if widths:
                header = self.table.horizontalHeader()
                for col, width in enumerate(widths):
                    if col < header.count():
                        try:
                            header.resizeSection(col, int(width))
                        except (TypeError, ValueError):
                            pass
        except Exception:
            pass

    def _save_calendar_task_column_widths(self) -> None:
        """Takvim görev tablosu sütun genişliklerini kaydet."""
        try:
            header = self.table.horizontalHeader()
            widths = [header.sectionSize(col) for col in range(header.count())]
            settings = QSettings("MyCompany", "TakibiEsasi")
            settings.setValue("calendar_tasks/col_widths", widths)
            settings.sync()
        except Exception:
            pass

    def _on_calendar_task_column_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Takvim görev tablosu sütun genişliği değiştiğinde kaydetmeyi tetikle."""
        if hasattr(self, "_calendar_task_column_save_timer"):
            self._calendar_task_column_save_timer.start()

    def _on_calendar_changed(self) -> None:
        self.selected_date = self.calendar.selectedDate()
        self._use_selected_date_filter = True
        self._set_month_label()
        self.refresh_tasks()

    def _on_calendar_double_clicked(self) -> None:
        self._add_task(for_date=self.calendar.selectedDate())

    def _collect_tasks(self, start: QDate, end: QDate) -> list[dict[str, Any]]:
        start_str = start.toString("yyyy-MM-dd")
        end_str = end.toString("yyyy-MM-dd")
        user_ids = self._user_id_lookup()
        manual_rows = get_manual_tasks_between(start_str, end_str)
        manual_tasks: list[dict[str, Any]] = []

        def _clean(value: Any) -> str:
            return (value or "").strip()

        for row in manual_rows:
            row_dict = dict(row)
            meta = GorevDialog._parse_meta(row_dict.get("aciklama", ""))
            date_obj, date_str = self._safe_date(row_dict.get("tarih"))
            # Dosyaya atanan kullanıcıları tercih et, yoksa görevin kendi atamasını kullan
            atanan_label = _clean(
                row_dict.get("dosya_atanan_kullanicilar")
                or row_dict.get("atanan_kullanicilar", "")
            )
            gorev_turu = row_dict.get("gorev_turu", "")
            manual_type_label = _clean(meta.get("type") or gorev_turu or row_dict.get("type") or "Değişik İş")
            konu_text = _clean(row_dict.get("konu", ""))

            # İçerik metnini görev türüne göre al
            if gorev_turu == "TEBLIGAT":
                icerik_text = _clean(meta.get("icerik", ""))
            elif gorev_turu == "ARABULUCULUK":
                icerik_text = _clean(meta.get("konu", ""))
            else:
                icerik_text = _clean(meta.get("notes", ""))

            # aciklama alanı __META__ ile başlıyorsa ham veriyi kullanma
            if not icerik_text:
                raw_aciklama = row_dict.get("aciklama", "")
                if raw_aciklama and not raw_aciklama.startswith("__META__"):
                    icerik_text = _clean(raw_aciklama)

            if not konu_text:
                konu_text = icerik_text or manual_type_label
            if not icerik_text:
                icerik_text = konu_text
            manual_tasks.append(
                {
                    "date": date_obj,
                    "date_str": date_str,
                    "tur_label": manual_type_label,
                    "bn": meta.get("bn", ""),
                    "muvekkil": meta.get("muvekkil", ""),
                    "konu_text": konu_text,
                    "icerik_text": icerik_text,
                    "calendar_text": konu_text,
                    "atanan_label": atanan_label,
                    "atanan_id": user_ids.get(atanan_label),
                    "atanan_kullanicilar": atanan_label,
                    "kaynak": "manuel",
                    "kaynak_field": "manuel",
                    "case_id": None,
                    "manual_id": row_dict.get("id"),
                    "dosya_id": None,
                    "source": "manual",
                    "type": gorev_turu or meta.get("type") or "MANUAL",
                    "task_id": row_dict.get("id"),
                    "id": row_dict.get("id"),
                }
            )

        case_tasks: list[dict[str, Any]] = []
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        case_meta_cache: dict[int, dict[str, Any]] = {}

        def _fetch_case_meta(dosya_id: Any) -> dict[str, Any]:
            if not dosya_id:
                return {}
            if dosya_id not in case_meta_cache:
                cur.execute(
                    """
                    SELECT dava_durumu, aciklama, tekrar_dava_durumu_2 AS dava_durumu_2, aciklama_2
                    FROM dosyalar
                    WHERE id = ?
                    """,
                    (dosya_id,),
                )
                row = cur.fetchone()
                case_meta_cache[dosya_id] = dict(row) if row else {}
            return case_meta_cache.get(dosya_id, {})

        for task in get_case_tasks_between(start_str, end_str):
            base = task | {"source": "auto"}
            meta_fields = _fetch_case_meta(base.get("dosya_id"))
            date_obj, date_str = self._safe_date(base.get("date"))
            key = base.get("task_id") or f"{base.get('dosya_id')}-{base.get('type')}"
            # Veritabanından gelen atamayı öncelikli kullan (dosya_atamalar'dan)
            # Cache sadece veritabanında değer yoksa kullanılsın
            db_assigned = _clean(base.get("atanan_kullanicilar", ""))
            assigned_label = db_assigned or _clean(self._auto_assignees.get(key, ""))
            task_type = base.get("type")
            tur_label = {
                "DURUSMA": "Duruşma",
                "IS_TARIHI": "İş Tarihi",
                "IS_TARIHI_2": "İş Tarihi 2",
            }.get(task_type, task_type or "Görev")

            def _field(meta_key: str, base_key: str = "") -> str:
                keys = [meta_key]
                if base_key:
                    keys.append(base_key)
                for key_name in keys:
                    if key_name and meta_fields.get(key_name):
                        return _clean(meta_fields.get(key_name))
                    if key_name and base.get(key_name):
                        return _clean(base.get(key_name))
                return ""

            konu_text = ""
            icerik_text = ""
            calendar_text = ""
            if task_type == "IS_TARIHI":
                konu_text = _field("dava_durumu")
                icerik_text = _field("aciklama")
                calendar_text = konu_text
            elif task_type == "IS_TARIHI_2":
                konu_text = _field("dava_durumu_2")
                icerik_text = _field("aciklama_2")
                calendar_text = konu_text
            elif task_type == "DURUSMA":
                konu_text = "Duruşma"
                icerik_text = "Duruşma"
                calendar_text = _field("dava_durumu") or konu_text
            else:
                konu_text = _field("description")
                icerik_text = _field("description")
                calendar_text = konu_text

            fallback_konu = next(
                (
                    cand
                    for cand in (
                        _field("dava_durumu"),
                        _field("dava_durumu_2"),
                        _field("aciklama"),
                        _field("description"),
                        tur_label,
                    )
                    if cand
                ),
                "",
            )
            if not konu_text:
                konu_text = fallback_konu or "Görev"
            if not icerik_text:
                icerik_text = _field("aciklama") or _field("description") or konu_text
            if not calendar_text:
                calendar_text = konu_text

            case_tasks.append(
                {
                    "date": date_obj,
                    "date_str": date_str,
                    "tur_label": tur_label,
                    "bn": base.get("bn", ""),
                    "muvekkil": base.get("muvekkil_adi", ""),
                    "konu_text": konu_text or "",
                    "icerik_text": icerik_text or "",
                    "calendar_text": calendar_text or "",
                    "atanan_label": assigned_label,
                    "atanan_id": user_ids.get(assigned_label),
                    "atanan_kullanicilar": assigned_label,
                    "kaynak": "otomatik",
                    "kaynak_field": {
                        "DURUSMA": "durusma_tarihi",
                        "IS_TARIHI": "is_tarihi",
                        "IS_TARIHI_2": "is_tarihi_2",
                    }.get(task_type, ""),
                    "case_id": base.get("dosya_id"),
                    "dosya_id": base.get("dosya_id"),
                    "manual_id": None,
                    "source": "auto",
                    "type": task_type,
                    "task_id": key,
                    "dava_durumu": base.get("dava_durumu"),  # Renklendirme için
                    "dava_durumu_color": base.get("dava_durumu_color"),  # Direkt renk
                }
            )
        conn.close()
        return manual_tasks + case_tasks

    def _add_task(self, *, for_date: QDate | None = None) -> None:
        target_date = for_date or self.calendar.selectedDate()
        dialog = GorevDialog(
            parent=self,
            mode="manual",
            task={"date": target_date.toString("yyyy-MM-dd")},
        )
        dialog.date_edit.setDate(target_date)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        data = dialog.result_data()
        insert_manual_task(
            data["tarih"],
            data["konu"],
            data.get("aciklama"),
            data.get("atanan_kullanicilar"),
            self.current_user.get("username"),
        )
        self._use_selected_date_filter = True
        self.refresh_tasks()

    def _apply_filters(self) -> list[dict[str, Any]]:
        start, end = self._range_for_filter()
        if self._use_selected_date_filter:
            start = self.calendar.selectedDate()
            end = start
        start_dt = start.toPyDate()
        end_dt = end.toPyDate()
        selected_user = ""
        if hasattr(self, "calendar_user_filter") and self.calendar_user_filter is not None:
            selected_user = self.calendar_user_filter.currentData() or ""
        filtered: list[dict[str, Any]] = []
        for task in self._all_tasks:
            task_date = task.get("date")
            if not isinstance(task_date, date):
                continue
            if task_date < start_dt or task_date > end_dt:
                continue
            # Çoklu kullanıcı desteği: virgülle ayrılmış listede seçili kullanıcıyı ara
            if selected_user:
                atanan_label = task.get("atanan_label", "")
                # Virgülle ayrılmış kullanıcıları listeye çevir
                atanan_list = [u.strip() for u in atanan_label.split(",") if u.strip()]
                if selected_user not in atanan_list:
                    continue
            filtered.append(task)
        return filtered

    def _edit_task(self) -> None:
        row = self.table.currentRow()
        if row < 0:
            return
        data = self.table.item(row, 0).data(self._task_data_role)
        if not isinstance(data, dict):
            return
        if data.get("source") == "auto" and not data.get("dosya_id"):
            return
        if data.get("source") == "manual":
            dialog = GorevDialog(parent=self, mode="manual", task=data)
            date_q = QDate.fromString(data.get("date_str", ""), "yyyy-MM-dd")
            if not date_q.isValid() and isinstance(data.get("date"), date):
                dt = data.get("date")
                date_q = QDate(dt.year, dt.month, dt.day)
            dialog.date_edit.setDate(date_q)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            # Dosya bağlantısı varsa atanan kullanıcıları dosya_atamalar'a kaydet
            dialog.save_assignees_if_linked_to_dosya()
            result = dialog.result_data()
            update_manual_task(
                int(data.get("manual_id") or data.get("id")),
                result["tarih"],
                result["konu"],
                result.get("aciklama"),
                result.get("atanan_kullanicilar"),
            )
        else:
            dialog = GorevDialog(parent=self, mode="auto", task=data)
            date_q = QDate.fromString(data.get("date_str", ""), "yyyy-MM-dd")
            if not date_q.isValid() and isinstance(data.get("date"), date):
                dt = data.get("date")
                date_q = QDate(dt.year, dt.month, dt.day)
            dialog.date_edit.setDate(date_q)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            # Dosya bağlantısı varsa atanan kullanıcıları dosya_atamalar'a kaydet
            dialog.save_assignees_if_linked_to_dosya()
            result = dialog.result_data()
            assigned_user = result.get("atanan_kullanicilar", "")
            data["atanan_kullanicilar"] = assigned_user
            data["atanan_label"] = assigned_user
            key = data.get("task_id") or f"{data.get('dosya_id')}-{data.get('type')}"
            if key:
                self._auto_assignees[key] = assigned_user
            field = {
                "DURUSMA": "durusma_tarihi",
                "IS_TARIHI": "is_tarihi",
                "IS_TARIHI_2": "is_tarihi_2",
            }.get(data.get("type"))
            if field:
                update_dosya_with_auto_timeline(
                    int(data.get("dosya_id")),
                    {field: result["tarih"]},
                    self.current_user.get("username", ""),
                )
        self.refresh_tasks()

    def _delete_task(self) -> None:
        row = self.table.currentRow()
        if row < 0:
            return
        data = self.table.item(row, 0).data(self._task_data_role)
        if not isinstance(data, dict) or data.get("source") != "manual":
            QMessageBox.information(self, "Silinemiyor", "Otomatik görevler buradan silinemez.")
            return
        # TEBLIGAT ve ARABULUCULUK görevleri kendi sekmelerinden silinmeli
        gorev_turu = data.get("type", "")
        if gorev_turu == "TEBLIGAT":
            QMessageBox.information(
                self,
                "Silinemiyor",
                "Tebligat görevleri sadece Tebligatlar sekmesinden silinebilir.",
            )
            return
        if gorev_turu == "ARABULUCULUK":
            QMessageBox.information(
                self,
                "Silinemiyor",
                "Arabuluculuk görevleri sadece Arabuluculuk sekmesinden silinebilir.",
            )
            return
        delete_manual_task(int(data["id"]))
        self.refresh_tasks()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._block_item_changed or item is None:
            return
        if item.column() != 4:
            return
        task = item.data(self._task_data_role)
        if not isinstance(task, dict):
            return
        new_value = item.text().strip()
        task["atanan_kullanicilar"] = new_value
        task["atanan_label"] = new_value
        if task.get("source") == "manual" and (task.get("manual_id") or task.get("id")):
            update_manual_task(
                int(task.get("manual_id") or task.get("id")),
                task.get("date_str", "") or (task.get("date") or ""),
                task.get("konu_text", ""),
                task.get("icerik_text") or task.get("description"),
                new_value,
            )
        else:
            key = task.get("task_id") or f"{task.get('dosya_id')}-{task.get('type')}"
            if key:
                self._auto_assignees[key] = new_value
        self.refresh_tasks()

    def _on_item_double_clicked(self) -> None:
        self._edit_task()

    def refresh_tasks(self) -> None:
        self._all_tasks = self._collect_tasks(*self._calendar_range())
        self._tasks = self._apply_filters()
        self._populate_table()
        self._refresh_calendar_formats()
        # Yapılacaklar sekmesini de güncelle
        self._collect_todo_tasks()
        self._populate_todo_table()

    def _populate_table(self) -> None:
        self._block_item_changed = True
        # Sıralama etkinken setItem() otomatik sıralama tetikleyebilir,
        # bu da satır indekslerinin geçersiz olmasına neden olur.
        # Önce sıralamayı devre dışı bırakıyoruz.
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        rows = list(self._tasks)
        rows.sort(key=lambda t: (t.get("date") or date.max, t.get("tur_label") or ""))
        for task in rows:
            row_idx = self.table.rowCount()
            self.table.insertRow(row_idx)
            task_date = task.get("date")
            date_q = (
                QDate(task_date.year, task_date.month, task_date.day)
                if isinstance(task_date, date)
                else QDate()
            )
            date_item = QTableWidgetItem(
                date_q.toString("dd.MM.yyyy") if date_q.isValid() else task.get("date_str", "")
            )
            if date_q.isValid():
                date_item.setData(Qt.ItemDataRole.UserRole, date_q.toJulianDay())
            konu_display = task.get("konu_text", "") or ""
            if task.get("type") == "DURUSMA":
                konu_display = "Duruşma"
            row_items = [
                date_item,
                QTableWidgetItem(konu_display),
                QTableWidgetItem(task.get("bn", "")),
                QTableWidgetItem(task.get("muvekkil", "")),
                QTableWidgetItem(task.get("atanan_label", "")),
                QTableWidgetItem(task.get("icerik_text", "")),
            ]
            for col, item in enumerate(row_items):
                item.setData(self._task_data_role, task)
                flags = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                if col == 4:
                    flags |= Qt.ItemFlag.ItemIsEditable
                item.setFlags(flags)
                self.table.setItem(row_idx, col, item)
            self._apply_row_color(row_idx, task)
        # Sıralamayı yeniden etkinleştir ve tarihe göre sırala
        self.table.setSortingEnabled(True)
        self.table.sortItems(0, Qt.SortOrder.AscendingOrder)
        self._block_item_changed = False

    def _apply_row_color(self, row: int, task: dict[str, Any]) -> None:
        """Takvim tablosunda satır stilini uygula.

        Sütunlar: [0: Tarih, 1: Konu, 2: BN, 3: Müvekkil, 4: Atanan, 5: İçerik]
        - Tarih sütunu (col=0): tarihe göre renklendirilir
        - Konu sütunu (col=1): dava durumuna göre renklendirilir (dosyalar gibi)
        - Diğer sütunlar: tema rengi (transparent)
        """
        task_date = task.get("date")

        # Tarih sütunu için renk hesapla
        color_info = get_task_color_by_date(task_date, is_completed=False)
        date_bg = QColor(color_info["bg"])
        date_fg = QColor(color_info["fg"])

        # Konu sütunu için renk al
        status_bg = None
        status_fg = None
        # Önce görev türüne göre özel renk kontrol et
        task_type = task.get("type", "")
        if task_type == "DURUSMA":
            status_bg = QColor("#c2185b")  # Pembe
            status_fg = QColor("#ffffff")  # Beyaz metin
        elif task_type == "TEBLIGAT":
            status_bg = QColor("#9b59b6")  # Mor
            status_fg = QColor("#ffffff")  # Beyaz metin
        elif task_type == "ARABULUCULUK":
            status_bg = QColor("#1abc9c")  # Turkuaz
            status_fg = QColor("#000000")  # Siyah metin (turkuaz açık renk)
        else:
            try:
                status_color = task.get("dava_durumu_color") or task.get("status_color")
                # Renk yoksa dava_durumu adından rengi al
                if not status_color:
                    dava_durumu = task.get("dava_durumu")
                    if dava_durumu and isinstance(dava_durumu, str) and dava_durumu.strip():
                        status_color = get_status_color(dava_durumu.strip())

                if status_color:
                    normalized = normalize_hex(status_color)
                    if normalized:
                        status_bg = QColor(f"#{normalized}")
                        status_fg = QColor(get_status_text_color(normalized))
            except Exception:
                pass  # Hata durumunda konu sütunu renksiz kalır

        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item is not None:
                font = item.font()
                if col == 0:  # Tarih sütunu
                    item.setBackground(date_bg)
                    item.setForeground(date_fg)
                    font.setBold(True)
                    item.setFont(font)
                elif col == 1 and status_bg:  # Konu sütunu - dava durumu rengi
                    item.setBackground(status_bg)
                    item.setForeground(status_fg)
                    font.setBold(True)
                    item.setFont(font)
                else:
                    # Diğer sütunlar tema renginde (transparent)
                    item.setData(Qt.ItemDataRole.BackgroundRole, None)
                    item.setData(Qt.ItemDataRole.ForegroundRole, None)

    def _refresh_calendar_formats(self) -> None:
        """
        Takvimdeki rozetler için tarih -> görev listesi sözlüğünü hazırlar.

        ÖNEMLİ DEĞİŞİKLİK:
        - Daha önce sadece self._tasks (filtrelenmiş görevler) kullanılıyordu.
        - Artık self._all_tasks kullanılıyor ki takvimde her zaman tüm görevler
          (filtre ne olursa olsun) işaretli kalsın.
        """
        tasks_by_date: dict[str, list[dict[str, Any]]] = {}

        # Takvim her zaman tüm görevleri bilsin:
        source_tasks = getattr(self, "_all_tasks", []) or []

        for task in source_tasks:
            # Önce varsa hazır date_str kullan, yoksa date alanını isoformat ile string'e çevir.
            date_val = task.get("date_str") or (
                task.get("date").isoformat() if isinstance(task.get("date"), date) else ""
            )
            if not date_val:
                continue

            calendar_subject = (
                task.get("calendar_text")
                or task.get("konu_text")
                or task.get("icerik_text")
                or ""
            )

            tasks_by_date.setdefault(date_val, []).append(
                {
                    "bn": task.get("bn", ""),
                    "konu": calendar_subject,
                    "description": task.get("icerik_text", ""),
                    "type": task.get("type"),
                    "type_label": task.get("tur_label", ""),
                    "atanan_kullanicilar": task.get("atanan_label", ""),
                    "dava_durumu": task.get("dava_durumu"),  # Rozet renklendirmesi için
                    "dava_durumu_color": task.get("dava_durumu_color"),  # Direkt renk
                }
            )

        self.calendar.set_tasks(tasks_by_date)

    # ==================== YAPILACAKLAR (TO-DO) SEKMESİ METODLARI ====================

    def _load_todo_column_widths(self) -> None:
        """Kayıtlı sütun genişliklerini yükle."""
        try:
            settings = QSettings("MyCompany", "TakibiEsasi")
            widths = settings.value("todo/col_widths", None)
            if widths:
                header = self.todo_table.horizontalHeader()
                for col, width in enumerate(widths):
                    if col < header.count() and col > 0:  # 0. sütun (checkbox) sabit
                        try:
                            header.resizeSection(col, int(width))
                        except (TypeError, ValueError):
                            pass
        except Exception:
            pass

    def _save_todo_column_widths(self) -> None:
        """Sütun genişliklerini kaydet."""
        try:
            header = self.todo_table.horizontalHeader()
            widths = [header.sectionSize(col) for col in range(header.count())]
            settings = QSettings("MyCompany", "TakibiEsasi")
            settings.setValue("todo/col_widths", widths)
            settings.sync()
        except Exception:
            pass

    def _on_todo_column_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Sütun genişliği değiştiğinde kaydetmeyi tetikle."""
        if hasattr(self, "_todo_column_save_timer"):
            self._todo_column_save_timer.start()

    def _collect_todo_tasks(self) -> None:
        """Tüm görevleri (manuel + otomatik) yapılacaklar listesi için topla."""
        user_ids = self._user_id_lookup()
        todo_list: list[dict[str, Any]] = []

        def _clean(value: Any) -> str:
            return (value or "").strip()

        # Manuel görevleri getir (tamamlanmamış + tamamlanmış)
        manual_rows = get_all_manual_tasks()
        for row in manual_rows:
            row_dict = dict(row)
            meta = GorevDialog._parse_meta(row_dict.get("aciklama", ""))
            date_obj, date_str = self._safe_date(row_dict.get("tarih"))
            # Dosyaya atanan kullanıcıları tercih et, yoksa görevin kendi atamasını kullan
            atanan_label = _clean(
                row_dict.get("dosya_atanan_kullanicilar")
                or row_dict.get("atanan_kullanicilar", "")
            )
            gorev_turu = row_dict.get("gorev_turu", "")
            manual_type_label = _clean(meta.get("type") or gorev_turu or row_dict.get("type") or "Değişik İş")
            konu_text = _clean(row_dict.get("konu", ""))

            # İçerik metnini görev türüne göre al
            if gorev_turu == "TEBLIGAT":
                icerik_text = _clean(meta.get("icerik", ""))
            elif gorev_turu == "ARABULUCULUK":
                icerik_text = _clean(meta.get("konu", ""))
            else:
                icerik_text = _clean(meta.get("notes", ""))

            # aciklama alanı __META__ ile başlıyorsa ham veriyi kullanma
            if not icerik_text:
                raw_aciklama = row_dict.get("aciklama", "")
                if raw_aciklama and not raw_aciklama.startswith("__META__"):
                    icerik_text = _clean(raw_aciklama)

            if not konu_text:
                konu_text = icerik_text or manual_type_label
            if not icerik_text:
                icerik_text = konu_text

            todo_list.append({
                "date": date_obj,
                "date_str": date_str,
                "tur_label": manual_type_label,
                "bn": meta.get("bn", ""),
                "muvekkil": meta.get("muvekkil", ""),
                "konu_text": konu_text,
                "icerik_text": icerik_text,
                "atanan_label": atanan_label,
                "atanan_id": user_ids.get(atanan_label),
                "atanan_kullanicilar": atanan_label,
                "kaynak": "manuel",
                "source": "manual",
                "type": gorev_turu or meta.get("type") or "MANUAL",
                "task_id": row_dict.get("id"),
                "id": row_dict.get("id"),
                "manual_id": row_dict.get("id"),
                "dosya_id": row_dict.get("dosya_id"),
                "tamamlandi": bool(row_dict.get("tamamlandi", 0)),
                "tamamlanma_zamani": row_dict.get("tamamlanma_zamani"),
            })

        # Otomatik görevleri de ekle (takvim aralığından)
        for task in self._all_tasks:
            if task.get("source") == "auto":
                # Otomatik görevler için tamamlandı durumu yok (henüz)
                task_copy = task.copy()
                task_copy["tamamlandi"] = False
                task_copy["tamamlanma_zamani"] = None
                todo_list.append(task_copy)

        self._todo_tasks = todo_list

    def _populate_todo_table(self) -> None:
        """Yapılacaklar tablosunu doldur."""
        self._block_todo_item_changed = True
        self.todo_table.setRowCount(0)

        # Kullanıcı filtresi
        selected_user = ""
        if hasattr(self, "todo_user_filter") and self.todo_user_filter is not None:
            selected_user = self.todo_user_filter.currentData() or ""

        # Sıralama: tamamlanmamış önce, sonra tarihe göre
        rows = list(self._todo_tasks)
        rows.sort(key=lambda t: (
            t.get("tamamlandi", False),  # Tamamlanmamışlar önce
            t.get("date") or date.max,   # Tarihe göre
            t.get("tur_label") or ""
        ))

        for task in rows:
            # Çoklu kullanıcı desteği: virgülle ayrılmış listede seçili kullanıcıyı ara
            if selected_user:
                atanan_label = task.get("atanan_label", "")
                atanan_list = [u.strip() for u in atanan_label.split(",") if u.strip()]
                if selected_user not in atanan_list:
                    continue
            row_idx = self.todo_table.rowCount()
            self.todo_table.insertRow(row_idx)

            # Checkbox sütunu
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(
                Qt.ItemFlag.ItemIsSelectable |
                Qt.ItemFlag.ItemIsEnabled |
                Qt.ItemFlag.ItemIsUserCheckable
            )
            is_completed = task.get("tamamlandi", False)
            checkbox_item.setCheckState(
                Qt.CheckState.Checked if is_completed else Qt.CheckState.Unchecked
            )
            checkbox_item.setData(self._task_data_role, task)

            # Tarih
            task_date = task.get("date")
            date_q = (
                QDate(task_date.year, task_date.month, task_date.day)
                if isinstance(task_date, date)
                else QDate()
            )
            date_text = date_q.toString("dd.MM.yyyy") if date_q.isValid() else "Tarihsiz"
            date_item = QTableWidgetItem(date_text)

            # Konu
            konu_display = task.get("konu_text", "") or ""
            if task.get("type") == "DURUSMA":
                konu_display = "Duruşma"
            konu_item = QTableWidgetItem(konu_display)

            row_items = [
                checkbox_item,
                date_item,
                konu_item,
                QTableWidgetItem(task.get("bn", "")),
                QTableWidgetItem(task.get("muvekkil", "")),
                QTableWidgetItem(task.get("atanan_label", "")),
                QTableWidgetItem(task.get("icerik_text", "")),
            ]

            is_manual = task.get("source") == "manual"
            for col, item in enumerate(row_items):
                item.setData(self._task_data_role, task)
                flags = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
                if col == 0:  # Checkbox
                    flags |= Qt.ItemFlag.ItemIsUserCheckable
                elif col == 1 and is_manual:  # Tarih - sadece manuel görevler için
                    flags |= Qt.ItemFlag.ItemIsEditable
                elif col == 5:  # Atanan
                    flags |= Qt.ItemFlag.ItemIsEditable
                item.setFlags(flags)
                self.todo_table.setItem(row_idx, col, item)

            # Renk ve üstü çizili
            self._apply_todo_row_style(row_idx, task)

        self._block_todo_item_changed = False

    def _apply_todo_row_style(self, row: int, task: dict[str, Any]) -> None:
        """Yapılacaklar tablosunda satır stilini uygula.

        Sütunlar: [0: ✓, 1: Tarih, 2: Konu, 3: BN, 4: Müvekkil, 5: Atanan, 6: İçerik]
        - Tarih sütunu (col=1): tarihe göre renklendirilir
        - Konu sütunu (col=2): dava durumuna göre renklendirilir (dosyalar gibi)
        - Diğer sütunlar: tema rengi (transparent)
        - Tamamlandı: üstü çizili ve gri yazı
        """
        is_completed = task.get("tamamlandi", False)
        task_date = task.get("date")

        # Tarih sütunu için renk hesapla
        color_info = get_task_color_by_date(task_date, is_completed)
        date_bg = QColor(color_info["bg"])
        date_fg = QColor(color_info["fg"])

        # Konu sütunu için renk al
        status_bg = None
        status_fg = None
        # Önce görev türüne göre özel renk kontrol et
        task_type = task.get("type", "")
        if task_type == "DURUSMA" and not is_completed:
            status_bg = QColor("#c2185b")  # Pembe
            status_fg = QColor("#ffffff")  # Beyaz metin
        elif task_type == "TEBLIGAT" and not is_completed:
            status_bg = QColor("#9b59b6")  # Mor
            status_fg = QColor("#ffffff")  # Beyaz metin
        elif task_type == "ARABULUCULUK" and not is_completed:
            status_bg = QColor("#1abc9c")  # Turkuaz
            status_fg = QColor("#000000")  # Siyah metin (turkuaz açık renk)
        else:
            try:
                status_color = task.get("dava_durumu_color") or task.get("status_color")
                # Renk yoksa dava_durumu adından rengi al
                if not status_color:
                    dava_durumu = task.get("dava_durumu")
                    if dava_durumu and isinstance(dava_durumu, str) and dava_durumu.strip():
                        status_color = get_status_color(dava_durumu.strip())

                if status_color and not is_completed:
                    normalized = normalize_hex(status_color)
                    if normalized:
                        status_bg = QColor(f"#{normalized}")
                        status_fg = QColor(get_status_text_color(normalized))
            except Exception:
                pass  # Hata durumunda konu sütunu renksiz kalır

        for col in range(self.todo_table.columnCount()):
            item = self.todo_table.item(row, col)
            if item is not None:
                font = item.font()
                font.setStrikeOut(is_completed)

                if col == 1:  # Tarih sütunu
                    item.setBackground(date_bg)
                    item.setForeground(date_fg)
                    font.setBold(True)
                elif col == 2 and status_bg:  # Konu sütunu - dava durumu rengi
                    item.setBackground(status_bg)
                    item.setForeground(status_fg)
                    font.setBold(True)
                else:
                    # Diğer sütunlar tema renginde (transparent)
                    item.setData(Qt.ItemDataRole.BackgroundRole, None)
                    if is_completed:
                        item.setForeground(QColor("#888888"))
                    else:
                        item.setData(Qt.ItemDataRole.ForegroundRole, None)

                item.setFont(font)

    def _add_todo_task(self) -> None:
        """Yapılacaklar sekmesinden yeni görev ekle."""
        dialog = GorevDialog(parent=self, mode="manual", task={})
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        data = dialog.result_data()
        insert_manual_task(
            data["tarih"],
            data["konu"],
            data.get("aciklama"),
            data.get("atanan_kullanicilar"),
            self.current_user.get("username"),
        )
        self.refresh_tasks()

    def _edit_todo_task(self) -> None:
        """Yapılacaklar sekmesinden görevi düzenle."""
        row = self.todo_table.currentRow()
        if row < 0:
            return
        data = self.todo_table.item(row, 0).data(self._task_data_role)
        if not isinstance(data, dict):
            return

        if data.get("source") == "manual":
            dialog = GorevDialog(parent=self, mode="manual", task=data)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            # Dosya bağlantısı varsa atanan kullanıcıları dosya_atamalar'a kaydet
            dialog.save_assignees_if_linked_to_dosya()
            result = dialog.result_data()
            update_manual_task(
                int(data.get("manual_id") or data.get("id")),
                result["tarih"],
                result["konu"],
                result.get("aciklama"),
                result.get("atanan_kullanicilar"),
            )
        else:
            # Otomatik görev - normal düzenleme diyaloğuna yönlendir
            dialog = GorevDialog(parent=self, mode="auto", task=data)
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            # Dosya bağlantısı varsa atanan kullanıcıları dosya_atamalar'a kaydet
            dialog.save_assignees_if_linked_to_dosya()
            result = dialog.result_data()
            assigned_user = result.get("atanan_kullanicilar", "")
            key = data.get("task_id") or f"{data.get('dosya_id')}-{data.get('type')}"
            if key:
                self._auto_assignees[key] = assigned_user
            field = {
                "DURUSMA": "durusma_tarihi",
                "IS_TARIHI": "is_tarihi",
                "IS_TARIHI_2": "is_tarihi_2",
            }.get(data.get("type"))
            if field and data.get("dosya_id"):
                update_dosya_with_auto_timeline(
                    int(data.get("dosya_id")),
                    {field: result["tarih"]},
                    self.current_user.get("username", ""),
                )
        self.refresh_tasks()

    def _toggle_todo_complete(self) -> None:
        """Seçili görevi tamamlandı/tamamlanmadı olarak işaretle."""
        row = self.todo_table.currentRow()
        if row < 0:
            return
        data = self.todo_table.item(row, 0).data(self._task_data_role)
        if not isinstance(data, dict):
            return

        if data.get("source") != "manual":
            QMessageBox.information(
                self, "Bilgi",
                "Otomatik görevler butonla tamamlanamaz.\nDava durumunu güncelleyerek otomatik olarak tamamlanır."
            )
            return

        task_id = data.get("manual_id") or data.get("id")
        if not task_id:
            return

        current_state = data.get("tamamlandi", False)
        mark_task_complete(int(task_id), not current_state)
        self.refresh_tasks()

    def _delete_todo_task(self) -> None:
        """Yapılacaklar sekmesinden görevi sil."""
        row = self.todo_table.currentRow()
        if row < 0:
            return
        data = self.todo_table.item(row, 0).data(self._task_data_role)
        if not isinstance(data, dict) or data.get("source") != "manual":
            QMessageBox.information(self, "Silinemiyor", "Otomatik görevler buradan silinemez.")
            return
        # TEBLIGAT ve ARABULUCULUK görevleri kendi sekmelerinden silinmeli
        gorev_turu = data.get("type", "")
        if gorev_turu == "TEBLIGAT":
            QMessageBox.information(
                self,
                "Silinemiyor",
                "Tebligat görevleri sadece Tebligatlar sekmesinden silinebilir.",
            )
            return
        if gorev_turu == "ARABULUCULUK":
            QMessageBox.information(
                self,
                "Silinemiyor",
                "Arabuluculuk görevleri sadece Arabuluculuk sekmesinden silinebilir.",
            )
            return
        delete_manual_task(int(data["id"]))
        self.refresh_tasks()

    def _on_todo_item_double_clicked(self, index: QModelIndex = None) -> None:
        """Yapılacaklar tablosunda çift tıklama.

        Tarih sütununda (column 1) çift tıklayınca inline düzenleme açılır,
        diğer sütunlarda düzenleme dialogu açılır.
        """
        # Tarih sütununda inline düzenleme - dialog açma
        if index is not None and index.column() == 1:
            return  # Delegate otomatik olarak editor açacak
        self._edit_todo_task()

    def _on_todo_item_changed(self, item: QTableWidgetItem) -> None:
        """Yapılacaklar tablosunda öğe değiştiğinde."""
        if self._block_todo_item_changed or item is None:
            return

        task = item.data(self._task_data_role)
        if not isinstance(task, dict):
            return

        # Checkbox değişikliği
        if item.column() == 0:
            if task.get("source") != "manual":
                # Otomatik görev checkbox'ı değiştirilmemeye çalışıldı, geri al
                self._block_todo_item_changed = True
                item.setCheckState(Qt.CheckState.Unchecked)
                self._block_todo_item_changed = False
                return

            task_id = task.get("manual_id") or task.get("id")
            if task_id:
                is_checked = item.checkState() == Qt.CheckState.Checked
                mark_task_complete(int(task_id), is_checked)
                self.refresh_tasks()
            return

        # Tarih sütunu değişikliği
        if item.column() == 1:
            new_date_str = item.text().strip()
            # Sadece manuel görevler düzenlenebilir
            if task.get("source") != "manual":
                return
            task_id = task.get("manual_id") or task.get("id")
            if task_id:
                # Türkçe formatı ISO formatına çevir
                try:
                    parsed = QDate.fromString(new_date_str, "dd.MM.yyyy")
                    if parsed.isValid():
                        iso_date = parsed.toString("yyyy-MM-dd")
                    else:
                        iso_date = new_date_str
                except Exception:
                    iso_date = new_date_str

                update_manual_task(
                    int(task_id),
                    iso_date,
                    task.get("konu_text", ""),
                    task.get("icerik_text") or task.get("description"),
                    task.get("atanan_kullanicilar") or "",
                )
                self.refresh_tasks()
            return

        # Atanan sütunu değişikliği
        if item.column() == 5:
            new_value = item.text().strip()
            task["atanan_kullanicilar"] = new_value
            task["atanan_label"] = new_value
            if task.get("source") == "manual" and (task.get("manual_id") or task.get("id")):
                update_manual_task(
                    int(task.get("manual_id") or task.get("id")),
                    task.get("date_str", "") or (task.get("date") or ""),
                    task.get("konu_text", ""),
                    task.get("icerik_text") or task.get("description"),
                    new_value,
                )
            else:
                key = task.get("task_id") or f"{task.get('dosya_id')}-{task.get('type')}"
                if key:
                    self._auto_assignees[key] = new_value
            self.refresh_tasks()


class MainWindow(QMainWindow):
    def __init__(self, current_user):
        super().__init__()

        # Pencere ikonunu ayarla
        import os
        icon_path = _resource_path("assets/icon.png")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.current_user = current_user
        self.current_user_id = self.current_user.get("id")
        load_status_palette()
        role = self.current_user.get("role")
        self.permissions: dict[str, bool] = self.current_user.get("permissions", {}) or {}
        self.can_manage_assignments = role in ASSIGNMENT_EDIT_ROLES
        self.can_manage_users = bool(self.permissions.get("manage_users", False))
        self.can_hard_delete = bool(self.permissions.get("can_hard_delete", False))
        self.can_view_all_cases = bool(self.permissions.get("view_all_cases", False))
        self.only_own_records = not self.can_view_all_cases
        self.can_view_finance = self._role_can_view_finance(role)
        self.finance_model: FinanceTableModel | None = None
        self.finance_proxy: FinanceProxyModel | None = None
        self.finance_table_view: QTableView | None = None
        self.finance_search_input: QLineEdit | None = None
        self.finance_user_filter_combo: QComboBox | None = None
        self.finance_tab: QWidget | None = None
        self.finance_tab_index: int | None = None
        self.finance_tabs: QTabWidget | None = None
        self.finance_summary_bar: FinanceSummaryBar | None = None
        self.finance_clear_button: QToolButton | None = None
        self.finance_filter_buttons: dict[str, QPushButton] = {}
        self.finance_active_filter: str = "all"
        self.harici_model: "HariciFinanceTableModel" | None = None
        self.harici_proxy: FinanceProxyModel | None = None
        self.harici_table_view: QTableView | None = None
        self.harici_search_input: QLineEdit | None = None
        self.harici_clear_button: QToolButton | None = None
        self.harici_summary_bar: FinanceSummaryBar | None = None
        self.harici_filter_buttons: dict[str, QPushButton] = {}
        self.harici_active_filter: str = "all"
        self.finance_bound_widget: QWidget | None = None
        self.finance_harici_widget: QWidget | None = None
        self._finance_search_timer = QTimer(self)
        self._finance_search_timer.setSingleShot(True)
        self._finance_search_timer.setInterval(300)
        self._finance_search_timer.timeout.connect(
            self._apply_finance_search_filter
        )
        self._pending_finance_search_text = ""
        self._harici_search_timer = QTimer(self)
        self._harici_search_timer.setSingleShot(True)
        self._harici_search_timer.setInterval(300)
        self._harici_search_timer.timeout.connect(
            self._apply_harici_search_filter
        )
        self._pending_harici_search_text = ""
        self.setWindowTitle(f"TakibiEsasi - Hoşgeldiniz, {current_user['username']}")

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # Demo banner'ı ekle (eğer demo modundaysa)
        self.demo_banner = None
        self._setup_demo_banner(main_layout)

        self._dosyalar_header_configured = False

        self.dosyalar_tab = DosyalarTab(
            self.current_user,
            mode="main",
            only_own_records=self.only_own_records,
            can_manage_assignments=self.can_manage_assignments,
            parent=self,
        )
        self._connect_dosyalar_tab_actions(self.dosyalar_tab, include_filters=True)

        self._setup_shortcuts()

        self.archive_table_model = DosyaTableModel()
        self.archive_table_view = QTableView()
        self._setup_table_view(self.archive_table_view, self.archive_table_model)
        self.archive_table_view.doubleClicked.connect(self.prompt_unarchive)

        self.custom_tab_widgets: dict[DosyalarTab, int] = {}
        self.column_indices: dict[str, dict[str, int]] = {}
        self._configured_table_views: set[QTableView] = set()

        self.tab_widget = QTabWidget()
        self.tab_bar = LexTabBar()
        self.tab_bar.tabDoubleClicked.connect(self.on_tab_bar_double_clicked)
        self.tab_bar.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tab_bar.customContextMenuRequested.connect(
            self.on_tab_bar_context_menu
        )
        self.tab_widget.setTabBar(self.tab_bar)

        self.add_tab_button = QPushButton("+")
        self.add_tab_button.setFixedWidth(28)
        self.add_tab_button.setToolTip("Yeni sekme ekle")
        self.add_tab_button.clicked.connect(self.on_add_custom_tab_clicked)
        self.tab_widget.setCornerWidget(self.add_tab_button, Qt.Corner.TopRightCorner)

        self.tab_widget.addTab(self.dosyalar_tab, "Dosyalar")
        self.gorevler_tab = GorevlerTab(self.current_user, self)
        self.gorevler_tab_index = self.tab_widget.addTab(
            self.gorevler_tab, "Görevler"
        )
        self.tebligatlar_tab = TebligatlarTab(
            current_user=self.current_user,
            parent=self,
        )
        self.tebligatlar_tab_index = self.tab_widget.addTab(
            self.tebligatlar_tab, "Tebligatlar"
        )
        self.arabuluculuk_tab = ArabuluculukTab(
            current_user=self.current_user,
            parent=self,
        )
        self.arabuluculuk_tab_index = self.tab_widget.addTab(
            self.arabuluculuk_tab, "Arabuluculuk"
        )
        self.archive_tab_index: int | None = None
        self._finance_widths_loaded = False
        if self.can_view_finance:
            self._setup_finance_tab()
        self._load_existing_custom_tabs()
        main_layout.addWidget(self.tab_widget)

        self._register_column_indices("Dosyalar", self.dosyalar_tab.table_model)
        self._register_column_indices("Arşiv", self.archive_table_model)

        self.refresh_table()

        self.update_column_widths()
        self.showMaximized()

        # Auto-refresh sistemi
        self._setup_auto_refresh()

    def _setup_auto_refresh(self) -> None:
        """Otomatik veri güncelleme sistemini başlat.

        SQLite trigger + change_log sistemi kullanır.
        Timestamp takibi yapmaz, sadece change_log tablosunu kontrol eder.
        """
        self._auto_refresh_enabled = True
        self._auto_refresh_interval = 30000  # 30 saniye
        self._auto_refresh_paused = False

        # Worker ve thread referansları
        self._change_detector_thread: QThread | None = None
        self._change_detector_worker: ChangeDetectorWorker | None = None

        # Timer oluştur
        self._auto_refresh_timer = QTimer(self)
        self._auto_refresh_timer.timeout.connect(self._check_for_changes)
        self._auto_refresh_timer.start(self._auto_refresh_interval)

    def _check_for_changes(self) -> None:
        """Arka planda değişiklikleri kontrol et."""
        if not self._auto_refresh_enabled or self._auto_refresh_paused:
            return
        # Thread hala çalışıyorsa bekle
        if self._change_detector_thread is not None:
            try:
                if self._change_detector_thread.isRunning():
                    return
            except RuntimeError:
                # Thread silinmiş, referansı temizle
                self._change_detector_thread = None
                self._change_detector_worker = None

        self._change_detector_worker = ChangeDetectorWorker()
        self._change_detector_thread = QThread(self)
        self._change_detector_worker.moveToThread(self._change_detector_thread)
        self._change_detector_thread.started.connect(self._change_detector_worker.run)
        self._change_detector_worker.changesDetected.connect(self._on_changes_detected)
        self._change_detector_worker.finished.connect(self._change_detector_thread.quit)
        self._change_detector_worker.finished.connect(self._change_detector_worker.deleteLater)
        self._change_detector_thread.finished.connect(self._on_change_detector_finished)
        self._change_detector_thread.start()

    def _on_change_detector_finished(self) -> None:
        """Thread bittiğinde referansları temizle."""
        self._change_detector_thread = None
        self._change_detector_worker = None

    def _on_changes_detected(self, changes: dict) -> None:
        """Değişiklik tespit edildiğinde çağrılır.

        changes dict'i şu anahtarları içerir:
        - dosyalar: bool - dosyalar tablosunda değişiklik var mı
        - gorevler: bool - gorevler tablosunda değişiklik var mı
        - finans: bool - finans tablosunda değişiklik var mı
        """
        refreshed = []
        if changes.get("dosyalar"):
            self.refresh_table()
            refreshed.append("Dosyalar")
        if changes.get("gorevler"):
            self.gorevler_tab.refresh_tasks()
            refreshed.append("Görevler")
        if changes.get("finans") and self.can_view_finance:
            self.refresh_finance_table()
            refreshed.append("Finans")

        if refreshed:
            self.statusBar().showMessage(f"Güncellendi: {', '.join(refreshed)}", 3000)

    def pause_auto_refresh(self) -> None:
        """Otomatik güncellemeyi duraklat."""
        self._auto_refresh_paused = True

    def resume_auto_refresh(self) -> None:
        """Otomatik güncellemeyi devam ettir."""
        self._auto_refresh_paused = False

    def _setup_table_view(self, view: QTableView, model: DosyaTableModel) -> None:
        model.attach_view(view)
        view.setModel(model)
        view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        view.setSortingEnabled(True)
        view.setWordWrap(False)
        view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        view.setVerticalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        view.setHorizontalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        view.verticalHeader().setVisible(False)
        header = view.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(False)
            header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _install_header_menu(
            view,
            lambda col: model.headers[col]
            if 0 <= col < len(model.headers)
            else "",
        )
        _install_copy_shortcut(view)

    def clear_all_filters(self) -> None:
        cleared = False
        current_widget = self.tab_widget.currentWidget()
        if isinstance(current_widget, DosyalarTab):
            current_widget.reset_filters()
            cleared = True
        elif current_widget is self.finance_tab:
            cleared = self._clear_finance_filters()
        elif current_widget is self.tebligatlar_tab:
            handler = getattr(self.tebligatlar_tab, "clear_filters", None)
            if callable(handler):
                handler()
                cleared = True
        elif current_widget is self.arabuluculuk_tab:
            handler = getattr(self.arabuluculuk_tab, "clear_filters", None)
            if callable(handler):
                handler()
                cleared = True
        if cleared:
            self.statusBar().showMessage("Filtreler temizlendi", 5000)

    def _clear_finance_filters(self) -> bool:
        if self.finance_tabs is None:
            return False
        current = self.finance_tabs.currentWidget()
        cleared = False
        if current is self.finance_bound_widget:
            if self.finance_search_input is not None:
                self.finance_search_input.blockSignals(True)
                if self.finance_search_input.text():
                    cleared = True
                self.finance_search_input.clear()
                self.finance_search_input.blockSignals(False)
            self._finance_search_timer.stop()
            self._pending_finance_search_text = ""
            self.finance_search_changed("")
            if self.finance_user_filter_combo is not None:
                if self.finance_user_filter_combo.currentIndex() != 0:
                    cleared = True
                self.finance_user_filter_combo.blockSignals(True)
                self.finance_user_filter_combo.setCurrentIndex(0)
                self.finance_user_filter_combo.blockSignals(False)
                self.finance_user_filter_changed()
            if self.finance_active_filter != "all":
                cleared = True
            self.apply_finance_quick_filter("all", force=True)
            cleared = True
        elif current is self.finance_harici_widget:
            if self.harici_search_input is not None:
                self.harici_search_input.blockSignals(True)
                if self.harici_search_input.text():
                    cleared = True
                self.harici_search_input.clear()
                self.harici_search_input.blockSignals(False)
            self._harici_search_timer.stop()
            self._pending_harici_search_text = ""
            self.harici_search_changed("")
            if self.harici_active_filter != "all":
                cleared = True
            self.apply_harici_quick_filter("all", force=True)
            if self.harici_table_view is not None:
                self.harici_table_view.scrollToTop()
            cleared = True
        return cleared

    def _setup_shortcuts(self) -> None:
        self._shortcut_find = QShortcut(
            QKeySequence(QKeySequence.StandardKey.Find), self
        )
        self._shortcut_find.activated.connect(self._focus_active_search)
        self._shortcut_new = QShortcut(
            QKeySequence(QKeySequence.StandardKey.New), self
        )
        self._shortcut_new.activated.connect(self._shortcut_new_record)
        delete_sequence = QKeySequence(Qt.Key.Key_Delete)
        self._shortcut_delete = QShortcut(delete_sequence, self)
        self._shortcut_delete.activated.connect(self._shortcut_delete_record)
        self._shortcut_clear_filters = QShortcut(
            QKeySequence("Ctrl+Backspace"), self
        )
        self._shortcut_clear_filters.activated.connect(self.clear_all_filters)
        self._shortcut_clear_filters_alt = QShortcut(
            QKeySequence("Ctrl+L"), self
        )
        self._shortcut_clear_filters_alt.activated.connect(self.clear_all_filters)
        self._shortcut_open_return = QShortcut(QKeySequence(Qt.Key.Key_Return), self)
        self._shortcut_open_return.setContext(
            Qt.ShortcutContext.WidgetWithChildrenShortcut
        )
        self._shortcut_open_return.activated.connect(
            self._shortcut_open_selected_row
        )
        self._shortcut_open_enter = QShortcut(QKeySequence(Qt.Key.Key_Enter), self)
        self._shortcut_open_enter.setContext(
            Qt.ShortcutContext.WidgetWithChildrenShortcut
        )
        self._shortcut_open_enter.activated.connect(
            self._shortcut_open_selected_row
        )
        self._shortcut_refresh = QShortcut(QKeySequence(Qt.Key.Key_F5), self)
        self._shortcut_refresh.activated.connect(self._shortcut_refresh_data)
        self._shortcut_escape = QShortcut(QKeySequence(Qt.Key.Key_Escape), self)
        self._shortcut_escape.setContext(
            Qt.ShortcutContext.WidgetWithChildrenShortcut
        )
        self._shortcut_escape.activated.connect(self._handle_escape_shortcut)

    def _focus_active_search(self) -> None:
        current_widget = self.tab_widget.currentWidget()
        if isinstance(current_widget, DosyalarTab) and current_widget.search_input is not None:
            current_widget.search_input.setFocus()
            current_widget.search_input.selectAll()
            return
        if current_widget is self.finance_tab and self.finance_search_input is not None:
            self.finance_search_input.setFocus()
            self.finance_search_input.selectAll()

    def _shortcut_new_record(self) -> None:
        self.new_file()

    def _shortcut_delete_record(self) -> None:
        current_widget = self.tab_widget.currentWidget()
        if not isinstance(current_widget, DosyalarTab):
            return
        record, is_archive = self._get_selected_record()
        if not record:
            return
        dosya_id = record.get("id")
        if dosya_id is None:
            return
        try:
            dosya_int = int(dosya_id)
        except (TypeError, ValueError):
            return
        if is_archive:
            if not self.can_hard_delete:
                QMessageBox.information(
                    self,
                    "Bilgi",
                    "Arşivdeki kaydı silmek için yetkiniz yok.",
                )
                return
            reply = QMessageBox.question(
                self,
                "Kalıcı Sil",
                "Arşivdeki bu dosya kalıcı olarak silinsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            conn = get_connection()
            try:
                delete_case_hard(conn, dosya_int)
            except Exception as exc:  # pragma: no cover - GUI güvenliği
                QMessageBox.critical(self, "Hata", f"Kayıt silinemedi:\n{exc}")
                return
            finally:
                conn.close()
            log_action(self.current_user["id"], "delete_dosya_hard", dosya_int)
            self.refresh_table()
            return

        message = QMessageBox(self)
        message.setWindowTitle("Dosyayı Sil")
        message.setText("Seçili dosyayı arşive göndermek veya kalıcı silmek ister misiniz?")
        archive_btn = message.addButton("Arşive Gönder", QMessageBox.ButtonRole.DestructiveRole)
        delete_btn = None
        if self.can_hard_delete:
            delete_btn = message.addButton("Kalıcı Sil", QMessageBox.ButtonRole.ActionRole)
        message.addButton(QMessageBox.StandardButton.Cancel)
        message.exec()
        clicked = message.clickedButton()
        if clicked is archive_btn:
            try:
                set_archive_status(dosya_int, True)
            except Exception as exc:
                QMessageBox.critical(self, "Hata", f"Arşive gönderilemedi:\n{exc}")
                return
            log_action(self.current_user["id"], "archive_dosya", dosya_int)
            self.refresh_table()
            self.open_archive_tab()
        elif delete_btn is not None and clicked is delete_btn:
            conn = get_connection()
            try:
                delete_case_hard(conn, dosya_int)
            except Exception as exc:  # pragma: no cover - GUI güvenliği
                QMessageBox.critical(self, "Hata", f"Kayıt silinemedi:\n{exc}")
                return
            finally:
                conn.close()
            log_action(self.current_user["id"], "delete_dosya_hard", dosya_int)
            self.refresh_table()

    def _shortcut_open_selected_row(self) -> None:
        current_widget = self.tab_widget.currentWidget()
        if not isinstance(current_widget, DosyalarTab):
            return
        view = current_widget.table_view
        if not isinstance(view, QTableView):
            return
        focus_widget = QApplication.focusWidget()
        if focus_widget is None or not (
            focus_widget is view or view.isAncestorOf(focus_widget)
        ):
            return
        if view.state() == QAbstractItemView.State.EditingState:
            return
        index = view.currentIndex()
        if not index.isValid():
            return
        self.edit_row(index)

    def _shortcut_refresh_data(self) -> None:
        current_widget = self.tab_widget.currentWidget()
        if current_widget is self.finance_tab and self._refresh_finance_shortcut_target():
            return
        self.refresh_table()

    def _refresh_finance_shortcut_target(self) -> bool:
        if self.finance_tabs is None:
            return False
        current = self.finance_tabs.currentWidget()
        if current is self.finance_bound_widget:
            self.refresh_finance_table()
            return True
        if current is self.finance_harici_widget:
            self.reload_harici_table()
            return True
        return False

    def _handle_escape_shortcut(self) -> None:
        tab = getattr(self, "dosyalar_tab", None)
        focus_widget = QApplication.focusWidget()
        if focus_widget is None:
            return
        if isinstance(tab, DosyalarTab) and self._close_editor_if_active(
            tab.table_view, focus_widget
        ):
            return
        if self._close_editor_if_active(self.finance_table_view, focus_widget):
            return
        self._close_editor_if_active(self.harici_table_view, focus_widget)

    def _close_editor_if_active(
        self, view: QTableView | None, focus_widget: QWidget
    ) -> bool:
        if view is None:
            return False
        if not (focus_widget is view or view.isAncestorOf(focus_widget)):
            return False
        if view.state() != QAbstractItemView.State.EditingState:
            return False
        view.closeEditor(
            focus_widget,
            QAbstractItemDelegate.EndEditHint.RevertModelCache,
        )
        view.setFocus()
        return True

    def _filter_records_by_date_fields(
        self,
        records: List[dict],
        field_names: tuple[str, ...],
        token: AlertTokenRange,
    ) -> List[dict]:
        if not records:
            return []
        start, end = token.start, token.end
        filtered: List[dict] = []
        for record in records:
            for field in field_names:
                if not field:
                    continue
                record_date = coerce_to_date(record.get(field))
                if record_date is None:
                    continue
                if start <= record_date <= end:
                    filtered.append(record)
                    break
        return filtered

    def _connect_dosyalar_tab_actions(
        self, tab: "DosyalarTab", *, include_filters: bool = False
    ) -> None:
        tab.new_requested.connect(self.new_file)
        tab.edit_requested.connect(self.edit_file)
        tab.archive_requested.connect(self.open_archive_tab)
        tab.attachments_requested.connect(self.manage_attachments)
        tab.vekalet_requested.connect(self.open_vekalet_dialog)
        tab.settings_requested.connect(self.open_settings)
        tab.row_double_clicked.connect(self.edit_row)
        tab.clear_filters_requested.connect(self.clear_all_filters)
        tab.refresh_requested.connect(self._refresh_dosyalar_table)
        if include_filters:
            tab.filters_changed.connect(self.refresh_table)

    def _load_existing_custom_tabs(self) -> None:
        conn = get_connection()
        try:
            tabs = list_custom_tabs(conn)
            if not tabs:
                return
            current_index = self.tab_widget.currentIndex()
            for tab_info in tabs:
                tab_id = tab_info.get("id")
                if tab_id is None:
                    continue
                try:
                    allowed_ids = get_dosya_ids_for_tab(conn, tab_id)
                except Exception:
                    allowed_ids = set()
                self._add_custom_tab_widget(
                    tab_id,
                    tab_info.get("name", "Sekme"),
                    select=False,
                    allowed_ids=allowed_ids,
                )
            if current_index != -1:
                self.tab_widget.setCurrentIndex(current_index)
        finally:
            conn.close()

    def _add_custom_tab_widget(
        self,
        tab_id: int,
        title: str,
        *,
        select: bool = True,
        allowed_ids: Optional[set[int]] = None,
    ) -> tuple["DosyalarTab", int]:
        tab = DosyalarTab(
            self.current_user,
            mode="custom",
            custom_tab_id=tab_id,
            only_own_records=self.only_own_records,
            can_manage_assignments=self.can_manage_assignments,
            parent=self.tab_widget,
        )
        self._connect_dosyalar_tab_actions(tab)
        self._register_column_indices(title, tab.table_model)
        self.custom_tab_widgets[tab] = tab_id
        index = self.tab_widget.addTab(tab, title)
        if allowed_ids is not None:
            tab.proxy.set_allowed_ids(allowed_ids)
        if select:
            self.tab_widget.setCurrentIndex(index)
        if self.dosyalar_tab.table_model.records:
            tab.set_records(list(self.dosyalar_tab.table_model.records))
        self.update_column_widths()
        return tab, index

    def on_add_custom_tab_clicked(self) -> None:
        conn = get_connection()
        try:
            tab_id = create_custom_tab(conn, "Yeni Sekme")
        except Exception as exc:  # pragma: no cover - GUI safety
            conn.close()
            QMessageBox.critical(
                self,
                "Hata",
                f"Yeni sekme oluşturulamadı:\n{exc}",
            )
            return

        try:
            _, index = self._add_custom_tab_widget(
                tab_id,
                "Yeni Sekme",
                allowed_ids=set(),
            )
        except Exception as exc:  # pragma: no cover - GUI safety
            conn.close()
            QMessageBox.critical(
                self,
                "Hata",
                f"Sekme arayüze eklenemedi:\n{exc}",
            )
            return

        try:
            new_name, ok = QInputDialog.getText(
                self,
                "Sekme Adı",
                "Yeni sekme adını giriniz:",
                text="Yeni Sekme",
            )
            if ok:
                cleaned = new_name.strip()
                if cleaned:
                    if cleaned != "Yeni Sekme":
                        try:
                            rename_custom_tab(conn, tab_id, cleaned)
                        except ValueError as exc:
                            QMessageBox.warning(self, "Uyarı", str(exc))
                        except Exception as exc:  # pragma: no cover - GUI safety
                            QMessageBox.critical(
                                self,
                                "Hata",
                                f"Sekme adı güncellenemedi:\n{exc}",
                            )
                        else:
                            self.tab_widget.setTabText(index, cleaned)
                            self._rename_column_index_entry("Yeni Sekme", cleaned)
                else:
                    QMessageBox.warning(
                        self,
                        "Uyarı",
                        "Sekme adı boş olamaz. Varsayılan ad kullanılacak.",
                    )
        finally:
            conn.close()

    def on_tab_bar_double_clicked(self, index: int) -> None:
        widget = self.tab_widget.widget(index)
        if not isinstance(widget, DosyalarTab):
            return
        tab_id = self.custom_tab_widgets.get(widget)
        if tab_id is None:
            return
        current_name = self.tab_widget.tabText(index)
        new_name, ok = QInputDialog.getText(
            self,
            "Sekme Adı",
            "Sekme adını giriniz:",
            text=current_name,
        )
        if not ok:
            return
        cleaned = new_name.strip()
        if not cleaned:
            QMessageBox.warning(self, "Uyarı", "Sekme adı boş olamaz.")
            return
        if cleaned == current_name:
            return
        conn = get_connection()
        try:
            rename_custom_tab(conn, tab_id, cleaned)
        except ValueError as exc:
            QMessageBox.warning(self, "Uyarı", str(exc))
            return
        except Exception as exc:  # pragma: no cover - GUI safety
            QMessageBox.critical(
                self,
                "Hata",
                f"Sekme adı güncellenemedi:\n{exc}",
            )
            return
        finally:
            conn.close()
        self.tab_widget.setTabText(index, cleaned)
        self._rename_column_index_entry(current_name, cleaned)

    def on_tab_bar_context_menu(self, pos: QPoint) -> None:
        if not self.can_hard_delete:
            return
        index = self.tab_bar.tabAt(pos)
        if index < 0:
            return
        widget = self.tab_widget.widget(index)
        if not isinstance(widget, DosyalarTab):
            return
        tab_id = self.custom_tab_widgets.get(widget)
        if tab_id is None:
            return

        menu = QMenu(self)
        delete_action = menu.addAction("Sekmeyi Sil")
        chosen_action = menu.exec(self.tab_bar.mapToGlobal(pos))
        if chosen_action is delete_action:
            self._confirm_and_delete_custom_tab(index, widget, tab_id)

    def _confirm_and_delete_custom_tab(
        self, index: int, widget: "DosyalarTab", tab_id: int
    ) -> None:
        answer = QMessageBox.question(
            self,
            "Sekmeyi Sil",
            "Bu sekmeyi silmek istediğinize emin misiniz?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        conn = get_connection()
        try:
            delete_custom_tab(conn, tab_id)
        except Exception as exc:  # pragma: no cover - GUI safety
            QMessageBox.critical(
                self,
                "Hata",
                f"Sekme silinemedi:\n{exc}",
            )
            return
        finally:
            conn.close()

        tab_name = self.tab_widget.tabText(index)
        self.custom_tab_widgets.pop(widget, None)
        self.column_indices.pop(tab_name, None)
        self.tab_widget.removeTab(index)
        widget.deleteLater()
        self.update_column_widths()

    def _setup_finance_tab(self) -> None:
        self.finance_tab = QWidget()
        outer_layout = QVBoxLayout(self.finance_tab)

        self.finance_tabs = QTabWidget()
        outer_layout.addWidget(self.finance_tabs)

        # --- Dosyalara bağlı finans görünümü ---
        bound_widget = QWidget()
        self.finance_bound_widget = bound_widget
        bound_layout = QVBoxLayout(bound_widget)
        search_layout = QHBoxLayout()
        self.finance_search_input = QLineEdit()
        self.finance_search_input.setPlaceholderText("BN, Esas No veya Müvekkil ara...")
        self.finance_search_input.setClearButtonEnabled(True)
        search_layout.addWidget(self.finance_search_input)
        self.finance_clear_button = QToolButton()
        self.finance_clear_button.setObjectName("clearFiltersButton")
        self.finance_clear_button.setText("✕")
        self.finance_clear_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.finance_clear_button.setToolTip(
            "Filtreleri temizle (Ctrl+Backspace)"
        )
        self.finance_clear_button.clicked.connect(self.clear_all_filters)
        search_layout.addWidget(self.finance_clear_button)
        if self.current_user.get("role") in {"admin", "yonetici_avukat"}:
            self.finance_user_filter_combo = QComboBox()
            search_layout.addWidget(self.finance_user_filter_combo)
        else:
            self.finance_user_filter_combo = None
        search_layout.addStretch(1)
        bound_layout.addLayout(search_layout)

        quick_layout = QHBoxLayout()
        quick_layout.setSpacing(6)
        quick_layout.addWidget(QLabel("Hızlı Filtre:"))
        button_specs = [
            ("all", "Tümü"),
            ("overdue", "Gecikenler"),
            ("week", "Bu Hafta"),
            ("month", "Bu Ay"),
        ]
        self.finance_filter_buttons.clear()
        for key, title in button_specs:
            button = QPushButton(title)
            button.setCheckable(True)
            button.setObjectName("FinanceQuickFilter")
            button.setStyleSheet(
                """
                QPushButton#FinanceQuickFilter {
                    border: 1px solid #3f3f3f;
                    border-radius: 8px;
                    padding: 4px 12px;
                    background-color: #2b2b2b;
                    color: #e0e0e0;
                }
                QPushButton#FinanceQuickFilter:hover {
                    background-color: #343434;
                }
                QPushButton#FinanceQuickFilter:checked {
                    background-color: #355070;
                    border-color: #4f76d1;
                }
                """
            )
            button.clicked.connect(partial(self.on_finance_quick_filter_clicked, key))
            quick_layout.addWidget(button)
            self.finance_filter_buttons[key] = button
        export_button_style = """
            QPushButton#FinanceExportButton {
                border: 1px solid #3f3f3f;
                border-radius: 8px;
                padding: 4px 12px;
                background-color: #234038;
                color: #e0e0e0;
            }
            QPushButton#FinanceExportButton:hover {
                background-color: #295045;
            }
            QPushButton#FinanceExportButton:pressed {
                background-color: #1d3430;
            }
        """
        export_xlsx_button = QPushButton("XLSX")
        export_xlsx_button.setObjectName("FinanceExportButton")
        export_xlsx_button.clicked.connect(partial(self.export_finance_view, "xlsx"))
        export_xlsx_button.setStyleSheet(export_button_style)
        quick_layout.addWidget(export_xlsx_button)
        export_pdf_button = QPushButton("PDF")
        export_pdf_button.setObjectName("FinanceExportButton")
        export_pdf_button.setStyleSheet(export_button_style)
        export_pdf_button.clicked.connect(partial(self.export_finance_view, "pdf"))
        quick_layout.addWidget(export_pdf_button)
        quick_layout.addStretch(1)
        bound_layout.addLayout(quick_layout)
        if "all" in self.finance_filter_buttons:
            self.finance_filter_buttons["all"].setChecked(True)

        self.finance_model = FinanceTableModel()
        self.finance_proxy = FinanceProxyModel(self)
        self.finance_proxy.setSourceModel(self.finance_model)

        self.finance_table_view = QTableView()
        self.finance_table_view.setObjectName("financeTable")
        self.finance_table_view.setModel(self.finance_proxy)
        self.finance_table_view.setItemDelegate(FinanceRowDelegate(self.finance_table_view))
        self.finance_table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.finance_table_view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.finance_table_view.setSortingEnabled(True)
        self.finance_table_view.setWordWrap(False)
        self.finance_table_view.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.finance_table_view.setVerticalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.finance_table_view.setHorizontalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.finance_table_view.verticalHeader().setVisible(False)
        self.finance_table_view.doubleClicked.connect(self.open_finance_dialog)
        self.finance_table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.finance_table_view.customContextMenuRequested.connect(
            self.open_finance_context_menu
        )
        header = self.finance_table_view.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(False)
            header.setSectionResizeMode(
                QHeaderView.ResizeMode.Interactive
            )
            # Sütun genişliği değiştiğinde otomatik kaydet
            self._finance_column_save_timer = QTimer(self)
            self._finance_column_save_timer.setSingleShot(True)
            self._finance_column_save_timer.setInterval(500)
            self._finance_column_save_timer.timeout.connect(self.save_finance_column_widths)
            header.sectionResized.connect(self._on_finance_section_resized)
        _install_header_menu(
            self.finance_table_view,
            lambda col: self.finance_model.headers[col]
            if self.finance_model and 0 <= col < len(self.finance_model.headers)
            else "",
        )
        _install_copy_shortcut(self.finance_table_view)
        bound_layout.addWidget(self.finance_table_view, 1)

        self.finance_summary_bar = FinanceSummaryBar(bound_widget)
        bound_layout.addWidget(self.finance_summary_bar)

        self.finance_tabs.addTab(bound_widget, "Dosyalara Bağlı")

        self.finance_search_input.textChanged.connect(
            self._handle_finance_search_changed
        )
        if self.finance_user_filter_combo is not None:
            self.populate_finance_user_filter()
            self.finance_user_filter_combo.currentIndexChanged.connect(
                self.finance_user_filter_changed
            )

        # --- Harici finans görünümü ---
        harici_widget = QWidget()
        self.finance_harici_widget = harici_widget
        harici_layout = QVBoxLayout(harici_widget)
        harici_search_layout = QHBoxLayout()
        self.harici_search_input = QLineEdit()
        self.harici_search_input.setPlaceholderText("BN veya Müvekkil ara...")
        self.harici_search_input.setClearButtonEnabled(True)
        harici_search_layout.addWidget(self.harici_search_input)
        self.harici_clear_button = QToolButton()
        self.harici_clear_button.setObjectName("clearFiltersButton")
        self.harici_clear_button.setText("✕")
        self.harici_clear_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.harici_clear_button.setToolTip(
            "Filtreleri temizle (Ctrl+Backspace)"
        )
        self.harici_clear_button.clicked.connect(self.clear_all_filters)
        harici_search_layout.addWidget(self.harici_clear_button)
        self.btn_new_harici = QPushButton("Yeni Harici Finans")
        self.btn_new_harici.clicked.connect(self.on_new_harici_clicked)
        harici_search_layout.addWidget(self.btn_new_harici)
        harici_layout.addLayout(harici_search_layout)

        harici_quick_layout = QHBoxLayout()
        harici_quick_layout.setSpacing(6)
        harici_quick_layout.addWidget(QLabel("Hızlı Filtre:"))
        button_specs = [
            ("all", "Tümü"),
            ("overdue", "Gecikenler"),
            ("week", "Bu Hafta"),
            ("month", "Bu Ay"),
        ]
        self.harici_filter_buttons.clear()
        quick_button_style = """
            QPushButton#FinanceQuickFilter {
                border: 1px solid #3f3f3f;
                border-radius: 8px;
                padding: 4px 12px;
                background-color: #2b2b2b;
                color: #e0e0e0;
            }
            QPushButton#FinanceQuickFilter:hover {
                background-color: #343434;
            }
            QPushButton#FinanceQuickFilter:checked {
                background-color: #355070;
                border-color: #4f76d1;
            }
        """
        for key, title in button_specs:
            button = QPushButton(title)
            button.setCheckable(True)
            button.setObjectName("FinanceQuickFilter")
            button.setStyleSheet(quick_button_style)
            button.clicked.connect(partial(self.on_harici_quick_filter_clicked, key))
            harici_quick_layout.addWidget(button)
            self.harici_filter_buttons[key] = button
        harici_export_button = QPushButton("XLSX")
        harici_export_button.setObjectName("FinanceExportButton")
        harici_export_button.setStyleSheet(export_button_style)
        harici_export_button.clicked.connect(
            partial(self.export_harici_finance_view, "xlsx")
        )
        harici_quick_layout.addWidget(harici_export_button)
        harici_pdf_button = QPushButton("PDF")
        harici_pdf_button.setObjectName("FinanceExportButton")
        harici_pdf_button.setStyleSheet(export_button_style)
        harici_pdf_button.clicked.connect(
            partial(self.export_harici_finance_view, "pdf")
        )
        harici_quick_layout.addWidget(harici_pdf_button)
        harici_quick_layout.addStretch(1)
        harici_layout.addLayout(harici_quick_layout)

        if "all" in self.harici_filter_buttons:
            self.harici_filter_buttons["all"].setChecked(True)

        self.harici_model = HariciFinanceTableModel(self)
        self.harici_proxy = FinanceProxyModel(self)
        self.harici_proxy.setSourceModel(self.harici_model)
        self.harici_table_view = QTableView()
        self.harici_table_view.setModel(self.harici_proxy)
        self.harici_table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.harici_table_view.setSelectionMode(QTableView.SelectionMode.SingleSelection)
        self.harici_table_view.setSortingEnabled(True)
        self.harici_table_view.setWordWrap(False)
        self.harici_table_view.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.harici_table_view.setVerticalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.harici_table_view.setHorizontalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel
        )
        self.harici_table_view.verticalHeader().setVisible(False)
        self.harici_table_view.doubleClicked.connect(self.on_harici_double_clicked)
        self.harici_table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.harici_table_view.customContextMenuRequested.connect(
            self._harici_context_menu
        )
        harici_header = self.harici_table_view.horizontalHeader()
        if harici_header is not None:
            harici_header.setStretchLastSection(False)
            harici_header.setSectionResizeMode(
                QHeaderView.ResizeMode.Interactive
            )
        _install_header_menu(
            self.harici_table_view,
            lambda col: self.harici_model.headers[col]
            if self.harici_model and 0 <= col < len(self.harici_model.headers)
            else "",
        )
        _install_copy_shortcut(self.harici_table_view)

        # Sütun genişliklerini kaydet/yükle
        self._load_harici_column_widths()
        self._harici_column_save_timer = QTimer(self)
        self._harici_column_save_timer.setSingleShot(True)
        self._harici_column_save_timer.setInterval(500)
        self._harici_column_save_timer.timeout.connect(self._save_harici_column_widths)
        if harici_header is not None:
            harici_header.sectionResized.connect(self._on_harici_column_resized)

        harici_layout.addWidget(self.harici_table_view)

        self.harici_summary_bar = FinanceSummaryBar(harici_widget)
        harici_layout.addWidget(self.harici_summary_bar)

        self.finance_tabs.addTab(harici_widget, "Harici Finans")

        self.harici_search_input.textChanged.connect(
            self._handle_harici_search_changed
        )

        self.finance_tab_index = self.tab_widget.addTab(self.finance_tab, "Finansal Takip")
        self._finance_widths_loaded = False
        self.refresh_finance_table()
        self.reload_harici_table()
        if self.isVisible():
            self.load_finance_column_widths()

    def _register_column_indices(self, tab_name: str, model: DosyaTableModel) -> None:
        keys = list(model.keys)
        mapping: dict[str, int] = {"arrow": 0}
        for field in [
            "buro_takip_no",
            "dosya_esas_no",
            "muvekkil_adi",
            "karsi_taraf",
            "dosya_konusu",
            "mahkeme_adi",
            "durusma_tarihi",
            "dava_durumu",
            "is_tarihi",
            "aciklama",
            "tekrar_dava_durumu_2",
            "is_tarihi_2",
            "aciklama_2",
        ]:
            try:
                mapping[field] = keys.index(field)
            except ValueError:
                continue
        self.column_indices[tab_name] = mapping

    def _rename_column_index_entry(self, old_name: str, new_name: str) -> None:
        if old_name == new_name:
            return
        indices = self.column_indices.pop(old_name, None)
        if indices is not None:
            self.column_indices[new_name] = indices

    def _iter_table_views(self) -> list[tuple[str, QTableView]]:
        views: list[tuple[str, QTableView]] = [("Dosyalar", self.dosyalar_tab.table_view)]
        views.append(("Arşiv", self.archive_table_view))
        for tab in self.custom_tab_widgets:
            index = self.tab_widget.indexOf(tab)
            if index < 0:
                continue
            tab_name = self.tab_widget.tabText(index)
            views.append((tab_name, tab.table_view))
        return views

    def refresh_custom_tab_filters(self) -> None:
        if not self.custom_tab_widgets:
            return
        conn = get_connection()
        try:
            for tab, tab_id in self.custom_tab_widgets.items():
                try:
                    ids = get_dosya_ids_for_tab(conn, tab_id)
                except Exception:
                    ids = set()
                tab.proxy.set_allowed_ids(ids)
        finally:
            conn.close()

    def refresh_table(self):
        """Verileri yeniden yükler."""
        filters = self._collect_filters()
        hex6 = filters["hex6"]
        raw_search_text = filters["search_text"] or ""
        cleaned_search, token_filters = parse_alert_tokens(raw_search_text)
        search_text = cleaned_search or None
        open_only = filters["open_only"]
        other_filters = filters["other_filters"]
        assigned_user_id = filters["assigned_user_id"]
        t0 = _now()
        active_records = self._query_files(
            hex6,
            search_text=search_text,
            open_only=open_only,
            other_filters=other_filters,
            assigned_user_id=assigned_user_id,
            archived=False,
        )
        t1 = _now()
        active_records = self._apply_post_query_filters(active_records, token_filters)
        records_for_custom = list(active_records)
        t_build_start = _now()
        built_rows = self._build_model_rows(active_records)
        t2 = _now()
        self._apply_model(built_rows)
        t3 = _now()
        self._setup_header_if_needed()
        t4 = _now()
        self._apply_proxy_sort_filter()
        t5 = _now()
        self._final_view_adjustments()
        t6 = _now()
        row_count = len(active_records)
        print("[perf] dosyalar.DB         ", _ms(t0, t1), f"rows={row_count}")
        print("[perf] dosyalar.BUILD_ROWS ", _ms(t_build_start, t2))
        print("[perf] dosyalar.APPLY_MODEL", _ms(t2, t3))
        print("[perf] dosyalar.HEADER     ", _ms(t3, t4))
        print("[perf] dosyalar.PROXY_SORT ", _ms(t4, t5))
        print("[perf] dosyalar.VIEW_FINAL ", _ms(t5, t6))
        print("[perf] dosyalar.TOTAL      ", _ms(t0, t6))

        if self.custom_tab_widgets:
            for tab in self.custom_tab_widgets:
                tab.set_records(list(records_for_custom))
            self.refresh_custom_tab_filters()

        archived_records = self._query_files(
            hex6,
            search_text=search_text,
            open_only=False,
            other_filters=other_filters,
            assigned_user_id=assigned_user_id,
            archived=True,
        )
        archived_records = self._apply_post_query_filters(
            archived_records, token_filters
        )
        self.archive_table_model.set_records(archived_records)
        self.update_column_widths()
        if self.can_view_finance:
            self.refresh_finance_table()
        if getattr(self, "tebligatlar_tab", None) is not None:
            self.tebligatlar_tab.load_tebligatlar()
        print("[perf] summary ready")

    def _refresh_dosyalar_table(self) -> None:
        """Dosyalar tablosunu yenileme butonu için wrapper."""
        self.refresh_table()

    def _query_files(
        self,
        hex6: str | None,
        *,
        search_text: str | None,
        open_only: bool,
        other_filters: dict[str, object] | None,
        assigned_user_id: int | None,
        archived: bool,
    ) -> list[dict]:
        rows = fetch_dosyalar_by_color_hex(
            hex6,
            search_text=search_text,
            open_only=open_only,
            other_filters=other_filters,
            assigned_user_id=assigned_user_id,
            archived=archived,
        )
        # if False:
        #     rows = rows[:500]  # GEÇİCİ: LIMIT 500 testi
        return rows

    def _apply_post_query_filters(
        self, records: list[dict], token_filters: dict[str, AlertTokenRange]
    ) -> list[dict]:
        if not records:
            return []
        hearing_token = token_filters.get("hearing")
        if hearing_token is None:
            return list(records)
        return self._filter_records_by_date_fields(
            list(records),
            ("durusma_tarihi", "duruşma_tarihi", "hearing_date"),
            hearing_token,
        )

    def _build_model_rows(self, records: list[dict[str, Any]]) -> list[PreparedDosyaRow]:
        tab = getattr(self, "dosyalar_tab", None)
        if tab is None:
            temp_model = DosyaTableModel()
            return temp_model.prepare_records(records)
        return tab.build_model_rows(records)

    def _apply_model(self, built_rows: list[PreparedDosyaRow]) -> None:
        tab = getattr(self, "dosyalar_tab", None)
        if tab is None:
            return
        tab.apply_model_rows(built_rows)

    def _setup_header_if_needed(self) -> None:
        if getattr(self, "_dosyalar_header_configured", False):
            return
        tab = getattr(self, "dosyalar_tab", None)
        if tab is None:
            return
        header = tab.table_view.horizontalHeader() if tab.table_view else None
        if header is None:
            return
        # header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # Tümü Interactive
        # header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._dosyalar_header_configured = True

    def _apply_proxy_sort_filter(self) -> None:
        tab = getattr(self, "dosyalar_tab", None)
        if tab is None:
            return
        tab.apply_proxy_sort_filter()
        # if False:  # Proxy sort kapama testi
        #     if tab.table_view is not None:
        #         tab.table_view.setSortingEnabled(False)

    def _final_view_adjustments(self) -> None:
        tab = getattr(self, "dosyalar_tab", None)
        if tab is None or tab.table_view is None:
            return
        tab.table_view.viewport().update()

    def refresh_finance_table(self) -> None:
        if not self.can_view_finance or self.finance_model is None:
            return
        records = list_finance_overview()
        self.finance_model.set_records(records)
        if self.finance_proxy is not None:
            self.finance_proxy.invalidateFilter()
        if self.finance_user_filter_combo is not None:
            self.finance_user_filter_changed()
        else:
            self.update_finance_summary()
        self.apply_finance_quick_filter(self.finance_active_filter, force=True)
        self.refresh_finance_colors()
        self.reload_harici_table()

    def refresh_finance_colors(self) -> None:
        if not self.finance_table_view or not self.finance_model:
            return
        roles = [
            Qt.ItemDataRole.ForegroundRole,
            Qt.ItemDataRole.BackgroundRole,
        ]
        rows = self.finance_model.rowCount()
        cols = self.finance_model.columnCount()
        if rows > 0 and cols > 0:
            top_left = self.finance_model.index(0, 0)
            bottom_right = self.finance_model.index(rows - 1, cols - 1)
            self.finance_model.dataChanged.emit(top_left, bottom_right, roles)
        if self.finance_proxy:
            proxy_rows = self.finance_proxy.rowCount()
            proxy_cols = self.finance_proxy.columnCount()
            if proxy_rows > 0 and proxy_cols > 0:
                top_left = self.finance_proxy.index(0, 0)
                bottom_right = self.finance_proxy.index(proxy_rows - 1, proxy_cols - 1)
                self.finance_proxy.dataChanged.emit(top_left, bottom_right, roles)
        self.finance_table_view.viewport().update()

    def _handle_finance_search_changed(self, text: str) -> None:
        self._pending_finance_search_text = text
        if self._finance_search_timer.isActive():
            self._finance_search_timer.stop()
        self._finance_search_timer.start()

    def _apply_finance_search_filter(self) -> None:
        self.finance_search_changed(self._pending_finance_search_text)

    def finance_search_changed(self, text: str) -> None:
        if self.finance_proxy is not None:
            self.finance_proxy.set_search_text(text)
        self.update_finance_summary()

    def on_finance_quick_filter_clicked(self, key: str) -> None:
        self.apply_finance_quick_filter(key, force=True)

    def apply_finance_quick_filter(self, key: str, *, force: bool = False) -> None:
        if key not in self.finance_filter_buttons:
            key = "all"
        if not force and key == self.finance_active_filter:
            return
        self.finance_active_filter = key
        category_map = {
            "all": None,
            "overdue": "overdue",
            "week": "this_week",
            "month": "this_month",
        }
        if self.finance_proxy is not None:
            self.finance_proxy.set_due_category_filter(category_map.get(key))
        for name, button in self.finance_filter_buttons.items():
            button.blockSignals(True)
            button.setChecked(name == key)
            button.blockSignals(False)
        self.update_finance_summary()

    # -------------------- Harici Finans Sütun Genişlikleri --------------------

    def _load_harici_column_widths(self) -> None:
        """Harici finans tablosu için kayıtlı sütun genişliklerini yükle."""
        if self.harici_table_view is None:
            return
        try:
            settings = QSettings("MyCompany", "TakibiEsasi")
            widths = settings.value("harici_finans/col_widths", None)
            if widths:
                header = self.harici_table_view.horizontalHeader()
                for col, width in enumerate(widths):
                    if col < header.count():
                        try:
                            header.resizeSection(col, int(width))
                        except (TypeError, ValueError):
                            pass
        except Exception:
            pass

    def _save_harici_column_widths(self) -> None:
        """Harici finans tablosu sütun genişliklerini kaydet."""
        if self.harici_table_view is None:
            return
        try:
            header = self.harici_table_view.horizontalHeader()
            widths = [header.sectionSize(col) for col in range(header.count())]
            settings = QSettings("MyCompany", "TakibiEsasi")
            settings.setValue("harici_finans/col_widths", widths)
            settings.sync()
        except Exception:
            pass

    def _on_harici_column_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Harici finans sütun genişliği değiştiğinde kaydetmeyi tetikle."""
        if hasattr(self, "_harici_column_save_timer"):
            self._harici_column_save_timer.start()

    def on_harici_quick_filter_clicked(self, key: str) -> None:
        self.apply_harici_quick_filter(key, force=True)

    def apply_harici_quick_filter(self, key: str, *, force: bool = False) -> None:
        if key not in self.harici_filter_buttons:
            key = "all"
        if not force and key == self.harici_active_filter:
            return
        self.harici_active_filter = key
        category_map = {
            "all": None,
            "overdue": "overdue",
            "week": "this_week",
            "month": "this_month",
        }
        if self.harici_proxy is not None:
            self.harici_proxy.set_due_category_filter(category_map.get(key))
        for name, button in self.harici_filter_buttons.items():
            button.blockSignals(True)
            button.setChecked(name == key)
            button.blockSignals(False)
        self.update_harici_summary()

    def update_finance_summary(self) -> None:
        if self.finance_summary_bar is None:
            return
        if self.finance_proxy is None:
            self.finance_summary_bar.update_totals()
            return
        visible_ids: list[int] = []
        for row in range(self.finance_proxy.rowCount()):
            record = self.finance_proxy.record_at(row)
            if not record:
                continue
            finance_id = record.get("finans_id")
            if finance_id in (None, ""):
                finance_id = record.get("id") or record.get("dosya_id")
            try:
                visible_ids.append(int(finance_id))
            except (TypeError, ValueError):
                continue
        if not visible_ids:
            self.finance_summary_bar.update_totals()
            return
        try:
            totals = summarize_finance_by_ids(visible_ids)
        except Exception:
            totals = {"contract": 0, "collected": 0, "expense": 0, "balance": 0}
        self.finance_summary_bar.update_totals(**totals)

    def update_harici_summary(self) -> None:
        if self.harici_summary_bar is None:
            return
        proxy = self.harici_proxy
        if proxy is None:
            self.harici_summary_bar.update_totals()
            return
        visible_ids: list[int] = []
        for row in range(proxy.rowCount()):
            record = proxy.record_at(row)
            if not record:
                continue
            rec_id = record.get("id")
            try:
                visible_ids.append(int(rec_id))
            except (TypeError, ValueError):
                continue
        if not visible_ids:
            self.harici_summary_bar.update_totals()
            return
        try:
            totals = summarize_harici_finance_by_ids(visible_ids)
        except Exception:
            totals = {"contract": 0, "collected": 0, "expense": 0, "balance": 0}
        self.harici_summary_bar.update_totals(**totals)

    def open_finance_context_menu(self, position) -> None:
        if self.finance_table_view is None:
            return
        index = self.finance_table_view.indexAt(position)
        if not index.isValid():
            return
        self.finance_table_view.selectRow(index.row())
        menu = QMenu(self.finance_table_view)
        mark_action = menu.addAction("Taksiti ödenmiş say")
        partial_action = menu.addAction("Kısmi ödeme gir…")
        chosen = menu.exec(self.finance_table_view.viewport().mapToGlobal(position))
        if chosen == mark_action:
            self.mark_selected_installment_paid()
        elif chosen == partial_action:
            self.prompt_partial_payment()

    def _selected_finance_record(self) -> dict | None:
        if self.finance_table_view is None or self.finance_model is None:
            return None
        selection = self.finance_table_view.selectionModel()
        if selection is None or not selection.hasSelection():
            return None
        selected_rows = selection.selectedRows()
        if not selected_rows:
            return None
        index = selected_rows[0]
        if self.finance_proxy is not None:
            index = self.finance_proxy.mapToSource(index)
        return self.finance_model.record_at(index.row())

    def mark_selected_installment_paid(self) -> None:
        record = self._selected_finance_record()
        if not record:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir finans satırı seçin.")
            return
        finans_id = record.get("finans_id")
        if not finans_id:
            QMessageBox.warning(self, "Uyarı", "Seçili kayıt için finans bilgisi bulunamadı.")
            return
        try:
            result = mark_next_installment_paid(int(finans_id))
        except Exception as exc:  # pragma: no cover - GUI güvenliği
            QMessageBox.critical(self, "Hata", f"Taksit güncellenemedi:\n{exc}")
            return
        if not result:
            QMessageBox.information(self, "Bilgi", "Ödenecek taksit bulunamadı.")
            return
        amount_text = format_tl(result.get("tutar_cents", 0))
        QMessageBox.information(
            self,
            "Başarılı",
            f"En yakın taksit ödenmiş olarak işaretlendi.\nTutar: {amount_text}",
        )
        self.refresh_finance_table()

    def prompt_partial_payment(self) -> None:
        record = self._selected_finance_record()
        if not record:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir finans satırı seçin.")
            return
        finans_id = record.get("finans_id")
        if not finans_id:
            QMessageBox.warning(self, "Uyarı", "Seçili kayıt için finans bilgisi bulunamadı.")
            return
        dialog = PartialPaymentDialog(self)
        if dialog.exec() != int(QDialog.DialogCode.Accepted):
            return
        data = dialog.result_data()
        try:
            add_partial_payment(
                int(finans_id),
                data["tarih"],
                data["tutar_cents"],
                data["yontem"],
                data["aciklama"],
            )
        except ValueError as exc:
            QMessageBox.warning(self, "Uyarı", str(exc))
            return
        except Exception as exc:  # pragma: no cover - GUI güvenliği
            QMessageBox.critical(self, "Hata", f"Kısmi ödeme kaydedilemedi:\n{exc}")
            return
        QMessageBox.information(self, "Başarılı", "Kısmi ödeme kaydedildi.")
        self.refresh_finance_table()

    def finance_user_filter_changed(self) -> None:
        if self.finance_proxy is None or self.finance_user_filter_combo is None:
            return
        data = self.finance_user_filter_combo.currentData()
        user_id: int | None
        if data is None or data == "":
            user_id = None
        elif isinstance(data, int):
            user_id = data
        else:
            try:
                user_id = int(data)
            except (TypeError, ValueError):
                user_id = None
        self.finance_proxy.set_user_filter(user_id)
        self.update_finance_summary()

    def export_finance_view(self, fmt: str) -> None:
        if self.finance_table_view is None:
            return
        model = self.finance_table_view.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.information(
                self,
                "Bilgi",
                "Dışa aktarılacak kayıt bulunamadı.",
            )
            return

        # Seçim dialogu göster
        dialog = TableExportSelectionDialog(
            self.finance_table_view, self,
            title="Dışa Aktarılacak Finans Kayıtlarını Seç",
            display_columns=[0, 1, 2],
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        selected_rows = dialog.get_selected_row_indices()
        if not selected_rows:
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        if fmt == "pdf":
            title = "PDF olarak kaydet"
            filter_str = "PDF Dosyaları (*.pdf)"
            extension = ".pdf"
        else:
            title = "Excel olarak kaydet"
            filter_str = "Excel Dosyaları (*.xlsx)"
            extension = ".xlsx"
        default_name = f"finans_{timestamp}{extension}"
        path, _ = QFileDialog.getSaveFileName(
            self,
            title,
            default_name,
            filter_str,
        )
        if not path:
            return
        if not path.lower().endswith(extension):
            path += extension
        if fmt == "pdf":
            export_table_to_pdf_with_selection(
                self.finance_table_view,
                path,
                title="TakibiEsasi – Finansal Takip",
                subtitle=self._finance_filter_summary(),
                selected_rows=selected_rows,
            )
        else:
            export_table_to_excel_with_selection(self.finance_table_view, path, selected_rows)

    def export_harici_finance_view(self, fmt: str) -> None:
        view = self.harici_table_view
        if view is None:
            return
        model = view.model()
        if model is None or model.rowCount() == 0:
            QMessageBox.information(
                self,
                "Bilgi",
                "Dışa aktarılacak harici finans kaydı bulunamadı.",
            )
            return

        # Seçim dialogu göster
        dialog = TableExportSelectionDialog(
            view, self,
            title="Dışa Aktarılacak Harici Finans Kayıtlarını Seç",
            display_columns=[0, 1, 2],
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        selected_rows = dialog.get_selected_row_indices()
        if not selected_rows:
            return

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        if fmt == "pdf":
            title = "PDF olarak kaydet"
            filter_str = "PDF Dosyaları (*.pdf)"
            extension = ".pdf"
        else:
            title = "Excel olarak kaydet"
            filter_str = "Excel Dosyaları (*.xlsx)"
            extension = ".xlsx"
        default_name = f"harici_finans_{timestamp}{extension}"
        path, _ = QFileDialog.getSaveFileName(
            self,
            title,
            default_name,
            filter_str,
        )
        if not path:
            return
        if not path.lower().endswith(extension):
            path += extension
        if fmt == "pdf":
            export_table_to_pdf_with_selection(
                view,
                path,
                title="TakibiEsasi – Harici Finansal Takip",
                subtitle=self._harici_filter_summary(),
                selected_rows=selected_rows,
            )
        else:
            export_table_to_excel_with_selection(view, path, selected_rows)

    def _current_finance_quick_filter_label(self, *, harici: bool = False) -> str:
        if harici:
            buttons = self.harici_filter_buttons
            active = self.harici_active_filter
        else:
            buttons = self.finance_filter_buttons
            active = self.finance_active_filter
        button = buttons.get(active)
        if button is None:
            return ""
        return button.text()

    def _finance_filter_summary(self) -> str:
        parts: list[str] = []
        if self.finance_search_input is not None:
            text = self.finance_search_input.text().strip()
            if text:
                parts.append(f"Arama: {text}")
        if self.finance_user_filter_combo is not None:
            current_index = self.finance_user_filter_combo.currentIndex()
            if current_index > 0:
                parts.append(
                    f"Kullanıcı: {self.finance_user_filter_combo.currentText()}"
                )
        quick = self._current_finance_quick_filter_label()
        if quick:
            parts.append(f"Hızlı Filtre: {quick}")
        if not parts:
            return "Filtre: Tümü"
        return " | ".join(parts)

    def _harici_filter_summary(self) -> str:
        parts: list[str] = []
        if self.harici_search_input is not None:
            text = self.harici_search_input.text().strip()
            if text:
                parts.append(f"Arama: {text}")
        quick = self._current_finance_quick_filter_label(harici=True)
        if quick:
            parts.append(f"Hızlı Filtre: {quick}")
        if not parts:
            return "Filtre: Tümü"
        return " | ".join(parts)

    def open_finance_dialog(self, index):
        if not self.can_view_finance or self.finance_model is None or self.finance_proxy is None:
            return
        if not index.isValid():
            return
        source_index = self.finance_proxy.mapToSource(index)
        record = self.finance_model.record_at(source_index.row())
        if not record:
            return
        finans_id = record.get("finans_id")
        if not finans_id:
            QMessageBox.warning(self, "Uyarı", "Finans kaydı bulunamadı.")
            return
        dosya_id = record.get("dosya_id")
        dialog = FinanceDialog(
            self,
            dosya_id=int(dosya_id) if dosya_id else None,
            finans_id=int(finans_id),
            dosya_info=record,
            current_user=self.current_user,
        )
        dialog.exec()
        self.refresh_finance_table()

    def on_new_harici_clicked(self) -> None:
        conn = get_connection()
        try:
            harici_id = harici_create(conn)
        except Exception as exc:  # pragma: no cover - GUI safety
            conn.close()
            QMessageBox.critical(self, "Hata", f"Harici finans oluşturulamadı:\n{exc}")
            return

        dialog = FinansHariciQuickDialog(
            self,
            conn=conn,
            harici_id=harici_id,
        )
        result = dialog.exec()
        if result != int(QDialog.DialogCode.Accepted):
            try:
                with conn:
                    conn.execute(
                        "DELETE FROM finans_harici WHERE id=? AND (harici_muvekkil IS NULL OR harici_muvekkil='') AND (harici_bn IS NULL OR harici_bn='')",
                        (harici_id,),
                    )
            except Exception:
                # Sessizce yoksay – boş kayıt kalırsa bir sonraki aramada düzenlenebilir
                pass
        conn.close()
        if result == int(QDialog.DialogCode.Accepted):
            self.reload_harici_table(select_id=harici_id, scroll_to_bottom=True)
        else:
            self.reload_harici_table()

    def _handle_harici_search_changed(self, text: str) -> None:
        self._pending_harici_search_text = text
        if self._harici_search_timer.isActive():
            self._harici_search_timer.stop()
        self._harici_search_timer.start()

    def _apply_harici_search_filter(self) -> None:
        self.harici_search_changed(self._pending_harici_search_text)

    def harici_search_changed(self, text: str) -> None:
        if self.harici_proxy is not None:
            self.harici_proxy.set_search_text(text)
        self.update_harici_summary()

    def reload_harici_table(
        self, select_id: int | None = None, scroll_to_bottom: bool = False
    ) -> None:
        if self.harici_model is None or self.harici_table_view is None:
            return
        conn = get_connection()
        try:
            records = harici_get_master_list(conn)
        finally:
            conn.close()

        self.harici_model.set_records(records)
        if self.harici_proxy is not None:
            self.harici_proxy.invalidateFilter()
        self.apply_harici_quick_filter(self.harici_active_filter, force=True)
        self.update_harici_summary()

        if select_id is not None:
            row_index = self.harici_model.find_row_by_id(select_id)
            if row_index is not None:
                source_index = self.harici_model.index(row_index, 0)
                target_index = source_index
                if self.harici_proxy is not None:
                    target_index = self.harici_proxy.mapFromSource(source_index)
                if target_index.isValid():
                    self.harici_table_view.selectRow(target_index.row())
                    self.harici_table_view.scrollTo(target_index)
                    return
        if scroll_to_bottom:
            self.harici_table_view.scrollToBottom()

    def on_harici_double_clicked(self, index: QModelIndex) -> None:
        if self.harici_model is None or self.harici_table_view is None:
            return
        if not index.isValid():
            return
        source_index = index
        if self.harici_proxy is not None:
            source_index = self.harici_proxy.mapToSource(index)
        if not source_index.isValid():
            return
        harici_id = self.harici_model.row_id(source_index.row())
        if harici_id is None:
            return
        conn = get_connection()
        try:
            dialog = FinansHariciDialog(self, conn=conn, harici_id=harici_id, current_user=self.current_user)
            dialog.exec()
        finally:
            conn.close()
        self.reload_harici_table(select_id=harici_id)

    def _harici_context_menu(self, pos: QPoint) -> None:
        if self.harici_model is None or self.harici_table_view is None:
            return
        index = self.harici_table_view.indexAt(pos)
        if not index.isValid():
            return
        source_index = index
        if self.harici_proxy is not None:
            source_index = self.harici_proxy.mapToSource(index)
        if not source_index.isValid():
            return
        harici_id = self.harici_model.row_id(source_index.row())
        if harici_id is None:
            return
        menu = QMenu(self.harici_table_view)
        action = menu.addAction("BN/Müvekkil Bilgileri…")
        chosen = menu.exec(self.harici_table_view.viewport().mapToGlobal(pos))
        if chosen is not action:
            return
        conn = get_connection()
        try:
            row = conn.execute(
                "SELECT harici_bn, harici_muvekkil FROM finans_harici WHERE id=?",
                (harici_id,),
            ).fetchone()
        finally:
            conn.close()
        bn, muv = (row or ("", ""))
        conn = get_connection()
        try:
            dialog = FinansHariciQuickDialog(
                self,
                conn=conn,
                harici_id=harici_id,
                init_bn=bn or "",
                init_muv=muv or "",
            )
            if dialog.exec():
                self.reload_harici_table(select_id=harici_id)
        finally:
            conn.close()

    def _role_can_view_finance(self, role: str | None) -> bool:
        if role == "admin":
            return True
        return bool(self.permissions.get("can_view_finance", False))

    def _setup_demo_banner(self, main_layout: QVBoxLayout) -> None:
        """Demo modundaysa banner göster."""
        try:
            db_path = get_database_path()
            demo_manager = get_demo_manager(db_path)
            status = demo_manager.get_demo_status()

            if status["status"] == "demo_active":
                days_remaining = status.get("days_remaining", 0)
                self.demo_banner = DemoStatusWidget(days_remaining)
                self.demo_banner.license_clicked.connect(self._on_demo_license_clicked)
                main_layout.addWidget(self.demo_banner)

        except Exception as e:
            print(f"Demo banner hatası: {e}")

    def _on_demo_license_clicked(self) -> None:
        """Demo banner'daki lisans butonuna tıklandı."""
        try:
            from app.ui_activation_dialog import ActivationDialog
        except ModuleNotFoundError:
            from ui_activation_dialog import ActivationDialog

        dialog = ActivationDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Lisans aktifleştirildi, banner'ı kaldır
            if self.demo_banner:
                self.demo_banner.setParent(None)
                self.demo_banner.deleteLater()
                self.demo_banner = None
                QMessageBox.information(
                    self,
                    "Lisans Aktif",
                    "Lisansınız başarıyla aktifleştirildi!\n"
                    "Tüm özelliklere sınırsız erişiminiz var."
                )

    def _update_assignment_checkbox_state(self) -> None:
        self.dosyalar_tab.only_own_records = self.only_own_records
        self.dosyalar_tab.update_assignment_checkbox_state()
        for tab in self.custom_tab_widgets:
            tab.only_own_records = self.only_own_records
            tab.update_assignment_checkbox_state()

    def _refresh_permissions(self) -> None:
        role = self.current_user.get("role")
        if role:
            self.current_user["permissions"] = get_permissions_for_role(role)
        self.permissions = self.current_user.get("permissions", {}) or {}
        self.can_manage_users = bool(self.permissions.get("manage_users", False))
        self.can_view_all_cases = bool(self.permissions.get("view_all_cases", False))
        self.only_own_records = not self.can_view_all_cases
        self.can_manage_assignments = role in ASSIGNMENT_EDIT_ROLES if role else False
        previous_finance = getattr(self, "can_view_finance", False)
        self.can_view_finance = self._role_can_view_finance(role)
        if not self.can_view_finance and self.finance_tab_index is not None:
            if self.archive_tab_index is not None and self.archive_tab_index > self.finance_tab_index:
                self.archive_tab_index -= 1
            self.tab_widget.removeTab(self.finance_tab_index)
            self.finance_tab_index = None
            self.finance_model = None
            self.finance_proxy = None
            self.finance_table_view = None
            self.finance_search_input = None
            self.finance_tab = None
            self.finance_tabs = None
            self.harici_model = None
            self.harici_table_view = None
            self.harici_search_input = None
            self._finance_widths_loaded = False
        elif self.can_view_finance and self.finance_tab_index is None and previous_finance != self.can_view_finance:
            self._setup_finance_tab()
        self.dosyalar_tab.update_user_context(
            self.current_user,
            only_own_records=self.only_own_records,
            can_manage_assignments=self.can_manage_assignments,
        )
        for tab in self.custom_tab_widgets:
            tab.update_user_context(
                self.current_user,
                only_own_records=self.only_own_records,
                can_manage_assignments=self.can_manage_assignments,
            )

    def _collect_filters(self) -> dict[str, object | None]:
        return self.dosyalar_tab.collect_filters()

    def _current_view_and_model(self):
        if (
            self.archive_tab_index is not None
            and self.tab_widget.currentIndex() == self.archive_tab_index
        ):
            return self.archive_table_view, self.archive_table_model, True
        current_widget = self.tab_widget.currentWidget()
        if isinstance(current_widget, DosyalarTab):
            return current_widget.table_view, current_widget.table_model, False
        return self.dosyalar_tab.table_view, self.dosyalar_tab.table_model, False

    def _map_to_source_index(
        self, view: QTableView, index: QModelIndex
    ) -> QModelIndex:
        model = view.model()
        if isinstance(model, QSortFilterProxyModel):
            return model.mapToSource(index)
        return index

    def _get_selected_record(self):
        view, model, is_archive = self._current_view_and_model()
        selection_model = view.selectionModel()
        if selection_model is None or not selection_model.hasSelection():
            return None, is_archive
        index = selection_model.selectedRows()[0]
        if not index.isValid():
            return None, is_archive
        source_index = self._map_to_source_index(view, index)
        if not source_index.isValid() or source_index.row() >= len(model.records):
            return None, is_archive
        return model.records[source_index.row()], is_archive

    def open_archive_tab(self):
        if self.archive_tab_index is None:
            self.archive_tab_index = self.tab_widget.addTab(self.archive_table_view, "Arşiv")
        self.tab_widget.setCurrentIndex(self.archive_tab_index)
        self.update_column_widths()

    def prompt_unarchive(self, index):
        if not index.isValid() or index.row() >= len(self.archive_table_model.records):
            return
        record = self.archive_table_model.records[index.row()]
        dosya_id = record.get("id")
        if dosya_id is None:
            return
        reply = QMessageBox.question(
            self,
            "Arşiv",
            "Arşivden çıkarılsın mı?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            set_archive_status(dosya_id, False)
            log_action(self.current_user["id"], "unarchive_dosya", dosya_id)
            self.refresh_table()
            self.tab_widget.setCurrentIndex(0)

    def new_file(self):
        try:
            self.pause_auto_refresh()
            dialog = EditDialog(self, current_user=self.current_user)
            if dialog.exec():
                self.refresh_table()
                self.gorevler_tab.refresh_tasks()
                log_action(self.current_user["id"], "add_dosya", dialog.dosya_id)
        except Exception as exc:  # pragma: no cover - GUI safety
            QMessageBox.critical(
                self,
                "Hata",
                f"Yeni dosya oluşturulurken bir hata oluştu:\n{exc}",
            )
        finally:
            self.resume_auto_refresh()

    def edit_row(self, index):
        if not index.isValid():
            return
        view, model, _ = self._current_view_and_model()
        if view is None or model is None:
            return
        source_index = self._map_to_source_index(view, index)
        if not source_index.isValid():
            return
        if source_index.row() >= len(model.records):
            return
        record = model.records[source_index.row()]
        dosya_id = record.get("id")
        if dosya_id is None:
            return
        try:
            self.pause_auto_refresh()
            dialog = EditDialog(self, dosya_id=dosya_id, current_user=self.current_user)
            if dialog.exec():
                self.refresh_table()
                self.gorevler_tab.refresh_tasks()
                if getattr(dialog, "was_hard_deleted", False):
                    target_id = dialog.hard_deleted_id or dosya_id
                    log_action(self.current_user["id"], "delete_dosya_hard", target_id)
                    return
                action = "archive_dosya" if dialog.was_archived else "update_dosya"
                log_action(self.current_user["id"], action, dosya_id)
                if dialog.was_archived:
                    self.open_archive_tab()
        finally:
            self.resume_auto_refresh()

    def edit_file(self):
        try:
            record, _ = self._get_selected_record()
            if not record:
                QMessageBox.warning(
                    self,
                    "Uyarı",
                    "Lütfen önce bir satır seçin.",
                )
                return
            dosya_id = record.get("id")
            if dosya_id is None:
                return
            self.pause_auto_refresh()
            dialog = EditDialog(self, dosya_id=dosya_id, current_user=self.current_user)
            if dialog.exec():
                self.refresh_table()
                self.gorevler_tab.refresh_tasks()
                if getattr(dialog, "was_hard_deleted", False):
                    target_id = dialog.hard_deleted_id or dosya_id
                    log_action(self.current_user["id"], "delete_dosya_hard", target_id)
                    self.resume_auto_refresh()
                    return
                action = "archive_dosya" if dialog.was_archived else "update_dosya"
                log_action(self.current_user["id"], action, dosya_id)
                if dialog.was_archived:
                    self.open_archive_tab()
            self.resume_auto_refresh()
        except Exception as exc:  # pragma: no cover - GUI safety
            self.resume_auto_refresh()
            QMessageBox.critical(
                self,
                "Hata",
                f"Dosya düzenlenirken bir hata oluştu:\n{exc}",
            )

    def manage_attachments(self):
        record, _ = self._get_selected_record()
        if not record:
            QMessageBox.warning(self, "Uyarı", "Ekleri yönetmek için bir satır seçiniz.")
            return
        dosya_id = record.get("id")
        if dosya_id is None:
            return
        dialog = AttachmentsDialog(self, dosya_id=dosya_id)
        dialog.exec()

    def open_vekalet_dialog(self) -> None:
        dialog = VekaletDialog(self)
        dialog.exec()

    def export_data(self):
        _, model, is_archive = self._current_view_and_model()
        filters = self._collect_filters()
        current_records = list(model.records)
        if not current_records:
            QMessageBox.information(
                self,
                "Bilgi",
                "Dışa aktarmak için listede kayıt bulunamadı.",
            )
            return

        choice_box = QMessageBox(self)
        choice_box.setWindowTitle("Dışa Aktar")
        choice_box.setText("Sadece mevcut filtre ile listelenen kayıtları dışa aktarmak ister misiniz?")
        choice_box.setInformativeText("Hayır seçeneği tüm kayıtları dışa aktarır.")
        choice_box.setStandardButtons(
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No
            | QMessageBox.StandardButton.Cancel
        )
        choice_box.setDefaultButton(QMessageBox.StandardButton.Yes)
        response = choice_box.exec()
        if response == QMessageBox.StandardButton.Cancel:
            return

        if response == QMessageBox.StandardButton.Yes:
            rows = current_records
        else:
            assigned_filter: int | None = None
            assigned_candidate = filters.get("assigned_user_id")
            if self.only_own_records and self.current_user_id is not None:
                assigned_filter = int(self.current_user_id)
            elif isinstance(assigned_candidate, int):
                assigned_filter = assigned_candidate
            rows = get_all_dosyalar(
                archived=is_archive,
                assigned_user_id=assigned_filter,
            )

        export_options = ["CSV", "XLSX", "Word (DOCX)"]
        format_choice, ok = QInputDialog.getItem(
            self,
            "Dışa Aktar",
            "Format seçin:",
            export_options,
            0,
            False,
        )
        if not ok:
            return

        format_map = {
            "CSV": (".csv", "CSV Dosyaları (*.csv)", export_dosyalar_to_csv),
            "XLSX": (".xlsx", "Excel Dosyaları (*.xlsx)", export_dosyalar_to_xlsx),
            "Word (DOCX)": (".docx", "Word Dosyaları (*.docx)", export_dosyalar_to_docx),
        }

        ext, file_filter, export_func = format_map[format_choice]
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        default_name = f"export_{timestamp}{ext}"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Dışa Aktar",
            default_name,
            file_filter,
        )
        if not path:
            return
        if not path.lower().endswith(ext):
            path += ext

        try:
            export_func(path, rows)
            QMessageBox.information(self, "Başarılı", "Dışa aktarma tamamlandı.")
        except Exception as exc:  # pragma: no cover - GUI safety
            QMessageBox.critical(
                self,
                "Hata",
                f"Dışa aktarma işlemi sırasında hata oluştu.\n{exc}",
            )

    def populate_status_filter(self):
        self.dosyalar_tab.populate_status_filter()

    def populate_user_filter(self) -> None:
        self.dosyalar_tab.populate_user_filter()

    def populate_finance_user_filter(self) -> None:
        if self.finance_user_filter_combo is None:
            return
        current_data = self.finance_user_filter_combo.currentData()
        self.finance_user_filter_combo.blockSignals(True)
        self.finance_user_filter_combo.clear()
        self.finance_user_filter_combo.addItem("Tümü", None)
        users = [user for user in get_users() if user.get("active")]
        users.sort(key=lambda item: item.get("username", "").lower())
        for user in users:
            self.finance_user_filter_combo.addItem(user.get("username", ""), user.get("id"))
        if current_data is not None:
            index = self.finance_user_filter_combo.findData(current_data)
            if index >= 0:
                self.finance_user_filter_combo.setCurrentIndex(index)
        self.finance_user_filter_combo.blockSignals(False)
        self.finance_user_filter_changed()

    def resizeEvent(self, event):  # type: ignore[override]
        super().resizeEvent(event)
        self.update_column_widths()

    def update_column_widths(self):
        for _, view in self._iter_table_views():
            self._apply_default_column_widths(view)

    def _apply_default_column_widths(self, view: QTableView | None) -> None:
        if view is None:
            return
        if getattr(view, "_lex_skip_main_defaults", False):
            return
        if view in self._configured_table_views:
            return
        header = view.horizontalHeader()
        if header is None:
            return
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        for column, width in DEFAULT_DOSYALAR_COLUMN_WIDTHS.items():
            if column >= header.count():
                continue
            if column == COL_SELECTION:
                header.setSectionResizeMode(column, QHeaderView.ResizeMode.Fixed)
            header.resizeSection(column, width)
        self._configured_table_views.add(view)

    def _on_finance_section_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        """Finans tablosu sütun genişliği değiştiğinde timer'ı başlat."""
        if hasattr(self, "_finance_column_save_timer"):
            self._finance_column_save_timer.start()

    def load_finance_column_widths(self) -> None:
        if self.finance_table_view is None:
            self._finance_widths_loaded = True
            return

        header = self.finance_table_view.horizontalHeader()
        if header is None:
            self._finance_widths_loaded = True
            return

        settings = QSettings("MyCompany", "TakibiEsasi")
        stored = settings.value("finance/col_widths")
        widths: list[int] = []
        if isinstance(stored, (list, tuple)):
            for value in stored:
                try:
                    widths.append(int(value))
                except (TypeError, ValueError):
                    continue
        elif isinstance(stored, str):
            for part in stored.split(","):
                part = part.strip()
                if not part:
                    continue
                try:
                    widths.append(int(part))
                except ValueError:
                    continue
        elif stored is not None:
            try:
                widths.append(int(stored))
            except (TypeError, ValueError):
                pass

        for col, width in enumerate(widths):
            if 0 <= col < header.count() and width > 0:
                header.resizeSection(col, width)

        self._finance_widths_loaded = True

    def save_finance_column_widths(self) -> None:
        if self.finance_table_view is None:
            return

        header = self.finance_table_view.horizontalHeader()
        if header is None:
            return

        widths = [header.sectionSize(col) for col in range(header.count())]
        settings = QSettings("MyCompany", "TakibiEsasi")
        settings.setValue("finance/col_widths", widths)
        settings.sync()

    def _save_dosyalar_header_state(self) -> None:
        view = getattr(self.dosyalar_tab, "table_view", None)
        if not isinstance(view, QTableView):
            return
        header = view.horizontalHeader()
        if header is None:
            return
        try:
            widths = [header.sectionSize(col) for col in range(header.count())]
            settings = QSettings("MyCompany", "TakibiEsasi")
            settings.setValue("dosyalar/col_widths", widths)
            settings.sync()
        except Exception:
            pass

    def showEvent(self, event):  # type: ignore[override]
        super().showEvent(event)
        if self.can_view_finance and not self._finance_widths_loaded:
            self.load_finance_column_widths()

    def closeEvent(self, event):  # type: ignore[override]
        try:
            self._save_dosyalar_header_state()
            self.save_finance_column_widths()
        finally:
            super().closeEvent(event)

    def open_settings(self):
        role = self.current_user.get("role")
        permissions = self.current_user.get("permissions", {})
        dialog = SettingsDialog(
            self,
            main_window=self,
            user_id=self.current_user["id"],
            is_admin=role == "admin",
            can_edit_statuses=role == "admin",
            show_status_tab=role != "stajyer",
            can_manage_users=self.can_manage_users,
            can_manage_backups=permissions.get("can_manage_backups", role == "admin"),
        )
        if dialog.exec():
            self._refresh_permissions()
            EditDialog.load_status_names()
            self.populate_status_filter()
            self.populate_user_filter()
            if self.finance_user_filter_combo is not None:
                self.populate_finance_user_filter()
            self.refresh_table()

# END OF PATCH
