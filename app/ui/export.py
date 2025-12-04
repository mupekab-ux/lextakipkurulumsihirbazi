# -*- coding: utf-8 -*-
"""
Tablo export fonksiyonları - Excel ve PDF çıktısı.
"""

import os
import html
from datetime import datetime
from typing import List

from PyQt6.QtCore import Qt, QSizeF, QRectF, QMarginsF
from PyQt6.QtGui import QPainter, QTextDocument, QPageLayout, QPageSize
from PyQt6.QtWidgets import QTableView, QMessageBox
from PyQt6.QtPrintSupport import QPrinter

__all__ = [
    "gather_visible_columns",
    "export_table_to_excel",
    "export_table_to_pdf",
]


def gather_visible_columns(view: QTableView) -> List[int]:
    """Tablodaki görünür sütunları döndürür."""
    model = view.model()
    if model is None:
        return []
    header = view.horizontalHeader()
    column_count = model.columnCount()
    if header is None:
        return [
            column
            for column in range(column_count)
            if not view.isColumnHidden(column)
        ]
    visible: List[int] = []
    for visual_index in range(header.count()):
        logical_index = header.logicalIndex(visual_index)
        if logical_index < 0 or logical_index >= column_count:
            continue
        if view.isColumnHidden(logical_index):
            continue
        visible.append(logical_index)
    return visible


def export_table_to_excel(view: QTableView, filename: str) -> None:
    """Tablo içeriğini Excel dosyasına aktarır."""
    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return

    try:
        from openpyxl import Workbook
    except ImportError:
        QMessageBox.critical(view, "Hata", "openpyxl modülü yüklü değil.")
        return

    columns = gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return

    rows = model.rowCount()
    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active

        # Başlık satırı
        for col_idx, column in enumerate(columns, start=1):
            header_text = model.headerData(
                column,
                Qt.Orientation.Horizontal,
                Qt.ItemDataRole.DisplayRole,
            )
            sheet.cell(row=1, column=col_idx, value=str(header_text or ""))

        # Veri satırları
        for row in range(rows):
            for col_offset, column in enumerate(columns, start=1):
                index = model.index(row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                sheet.cell(row=row + 2, column=col_offset, value=str(data or ""))

        workbook.save(filename)
        QMessageBox.information(
            view,
            "Başarılı",
            f"Excel dosyası oluşturuldu:\n{filename}",
        )
    except PermissionError:
        QMessageBox.critical(
            view,
            "Hata",
            "Dosya yazılamadı. Dosya başka bir programda açık olabilir.",
        )
    except OSError as exc:
        QMessageBox.critical(
            view,
            "Hata",
            f"Dosya sistemi hatası:\n{exc}",
        )
    except Exception as exc:
        QMessageBox.critical(
            view,
            "Hata",
            f"Excel çıktısı oluşturulamadı:\n{exc}",
        )


def export_table_to_pdf(
    view: QTableView,
    filename: str,
    *,
    title: str,
    subtitle: str | None = None,
) -> None:
    """Tablo içeriğini PDF dosyasına aktarır."""
    model = view.model()
    if model is None:
        QMessageBox.warning(view, "Uyarı", "Aktarılacak veri bulunamadı.")
        return

    columns = gather_visible_columns(view)
    if not columns:
        QMessageBox.warning(view, "Uyarı", "Görünür sütun bulunamadı.")
        return

    rows = model.rowCount()
    try:
        directory = os.path.dirname(os.path.abspath(filename))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        # HTML oluştur
        parts = [
            "<html><head>",
            "<style>",
            "body { font-family: 'DejaVu Sans', 'Arial', sans-serif; font-size: 9pt; color: #111; }",
            ".header { margin-bottom: 8px; }",
            ".header .title { font-size: 12pt; font-weight: bold; }",
            ".header .subtitle { font-size: 10pt; color: #444; }",
            ".header .meta { font-size: 9pt; color: #555; }",
            "table { border-collapse: collapse; width: 100%; font-size: 8.5pt; }",
            "th, td { border: 1px solid #666; padding: 4px 6px; text-align: left; }",
            "th { background-color: #efefef; }",
            "</style></head><body>",
        ]

        timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
        parts.append("<div class='header'>")
        parts.append(f"<div class='title'>{html.escape(title)}</div>")
        if subtitle:
            parts.append(f"<div class='subtitle'>{html.escape(subtitle)}</div>")
        parts.append(f"<div class='meta'>Oluşturma: {html.escape(timestamp)}</div>")
        parts.append("</div>")

        # Tablo başlığı
        parts.append("<table><thead><tr>")
        for column in columns:
            header_text = model.headerData(
                column,
                Qt.Orientation.Horizontal,
                Qt.ItemDataRole.DisplayRole,
            )
            parts.append(f"<th>{html.escape(str(header_text or ''))}</th>")
        parts.append("</tr></thead><tbody>")

        # Tablo verileri
        for row in range(rows):
            parts.append("<tr>")
            for column in columns:
                index = model.index(row, column)
                data = index.data(Qt.ItemDataRole.DisplayRole)
                parts.append(f"<td>{html.escape(str(data or ''))}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table></body></html>")

        # Belge oluştur
        document = QTextDocument()
        document.setDocumentMargin(12)
        document.setHtml("".join(parts))

        # Boyut hesapla
        header = view.horizontalHeader()
        approx_width = 0
        if header is not None:
            for column in columns:
                approx_width += header.sectionSize(column)
        if approx_width <= 0:
            approx_width = max(len(columns), 1) * 120

        vheader = view.verticalHeader()
        row_height = 0
        if vheader is not None:
            row_height = vheader.defaultSectionSize()
        if row_height <= 0:
            row_height = 24
        approx_height = max((rows + 4) * row_height, row_height)

        # Yazıcı ayarları
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(filename)

        layout = QPageLayout(
            QPageSize(QPageSize.PageSizeId.A4),
            QPageLayout.Orientation.Landscape,
            QMarginsF(12, 12, 12, 12),
        )
        printer.setPageLayout(layout)

        # Ölçekleme
        page_rect = printer.pageRect(QPrinter.Unit.Point)
        page_width = float(page_rect.width())
        page_height = float(page_rect.height())
        scale_x = page_width / float(approx_width) if approx_width else 1.0
        scale_y = page_height / float(approx_height) if approx_height else 1.0
        scale = min(scale_x, scale_y, 1.0) * 0.95
        if scale <= 0:
            scale = 0.95

        text_width = page_width / scale
        document.setPageSize(QSizeF(text_width, page_height / scale))

        # PDF oluştur
        painter = QPainter()
        if not painter.begin(printer):
            raise RuntimeError("PDF yazıcısı başlatılamadı")
        try:
            painter.translate(page_rect.left(), page_rect.top())
            painter.scale(scale, scale)
            document.drawContents(
                painter,
                QRectF(0, 0, text_width, page_height / scale),
            )
        finally:
            painter.end()

        QMessageBox.information(
            view,
            "Başarılı",
            f"PDF dosyası oluşturuldu:\n{filename}",
        )
    except PermissionError:
        QMessageBox.critical(
            view,
            "Hata",
            "Dosya yazılamadı. Dosya başka bir programda açık olabilir.",
        )
    except OSError as exc:
        QMessageBox.critical(
            view,
            "Hata",
            f"Dosya sistemi hatası:\n{exc}",
        )
    except Exception as exc:
        QMessageBox.critical(
            view,
            "Hata",
            f"PDF çıktısı oluşturulamadı:\n{exc}",
        )
